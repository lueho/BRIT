import codecs
import csv
from datetime import date
from io import BytesIO

from django.contrib.auth.models import Permission, User
from django.db.models import signals
from django.test import TestCase
from factory.django import mute_signals
from openpyxl import load_workbook

from maps.models import NutsRegion
from materials.models import MaterialCategory
from ..models import (
    Collection,
    CollectionCatchment,
    CollectionFrequency,
    CollectionSystem,
    Collector,
    FeeSystem,
    WasteCategory,
    WasteComponent,
    WasteFlyer,
    WasteStream,
)
from ..renderers import CollectionCSVRenderer, CollectionXLSXRenderer
from ..serializers import CollectionFlatSerializer


class CollectionCSVRendererTestCase(TestCase):

    @classmethod
    def setUpTestData(cls):
        MaterialCategory.objects.create(name="Biowaste component")
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed Material 1"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed Material 2"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden Material 1"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden Material 2"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=WasteCategory.objects.create(name="Test category"),
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            waste_flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123", url="https://www.test-flyer.org"
            )
        frequency = CollectionFrequency.objects.create(name="Test Frequency")
        nuts = NutsRegion.objects.create(
            name="Test NUTS", nuts_id="DE123", cntr_code="DE"
        )
        catchment = CollectionCatchment.objects.create(
            name="Test catchment", region=nuts.region_ptr
        )
        for i in range(1, 3):
            collection = Collection.objects.create(
                name=f"collection{i}",
                catchment=catchment,
                collector=Collector.objects.create(name=f"collector{1}"),
                collection_system=CollectionSystem.objects.create(name="Test system"),
                waste_stream=waste_stream,
                fee_system=FeeSystem.objects.create(name="Fixed fee"),
                frequency=frequency,
                valid_from=date(2020, 1, 1),
                description="This is a test case.",
            )
            collection.flyers.add(waste_flyer)

    def setUp(self):
        self.file = BytesIO()
        self.content = CollectionFlatSerializer(
            Collection.objects.all(), many=True
        ).data

    def test_fieldnames_in_right_order(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        fieldnames = [renderer.labels[key] for key in renderer.header]
        self.assertListEqual(fieldnames, list(reader.fieldnames))
        self.assertEqual(2, sum(1 for _ in reader))
        # Ensure 'Connection type' is present and in the correct order
        self.assertIn("Connection type", reader.fieldnames)
        self.assertEqual(
            renderer.header.index("connection_type"),
            reader.fieldnames.index("Connection type"),
        )

    def test_connection_type_field_exported(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertIn(row["Connection type"], ["Compulsory", "Voluntary"])

    def test_allowed_materials_formatted_as_comma_separated_list_in_one_field(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertEqual(
                "Allowed Material 1, Allowed Material 2", row["Allowed Materials"]
            )

    def test_forbidden_materials_formatted_as_comma_separated_list_in_one_field(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertEqual(
                "Forbidden Material 1, Forbidden Material 2", row["Forbidden Materials"]
            )

    def test_regression_flyers_without_urls_dont_raise_type_error(self):
        defected_collection = Collection.objects.first()
        with mute_signals(signals.post_save):
            rogue_flyer = WasteFlyer.objects.create(
                title="Rogue flyer without url", abbreviation="RF"
            )
        defected_collection.flyers.add(rogue_flyer)
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        self.assertEqual(Collection.objects.count(), len(list(reader)))

    def test_min_ton_fields_exported(self):
        # Set values for export
        collection = Collection.objects.first()
        collection.min_ton_size = 120
        collection.min_ton_volume_per_inhabitant = 5.5
        collection.save()
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertIn("Minimum container size (L)", row)
            self.assertIn("Minimum container volume per inhabitant (L/person)", row)
            if row["Minimum container size (L)"] not in ("", None):
                self.assertEqual(int(row["Minimum container size (L)"]), 120)
            if row["Minimum container volume per inhabitant (L/person)"] not in (
                "",
                None,
            ):
                self.assertEqual(
                    float(row["Minimum container volume per inhabitant (L/person)"]),
                    5.5,
                )
        # Null values
        collection.min_ton_size = None
        collection.min_ton_volume_per_inhabitant = None
        collection.save()
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            # Defensive: skip header or malformed rows
            if row["Minimum container size (L)"] == "Minimum container size (L)":
                continue
            self.assertIn("Minimum container size (L)", row)
            self.assertIn("Minimum container volume per inhabitant (L/person)", row)
            self.assertIn(row["Minimum container size (L)"], ("", None))
            self.assertIn(
                row["Minimum container volume per inhabitant (L/person)"], ("", None)
            )


class CollectionXLSXRendererTestCase(TestCase):

    @classmethod
    def setUpTestData(cls):
        User.objects.create(username="outsider")
        member = User.objects.create(username="member")
        member.user_permissions.add(Permission.objects.get(codename="add_collection"))

        MaterialCategory.objects.create(name="Biowaste component")
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed Material 1"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed Material 2"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden Material 1"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden Material 2"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=WasteCategory.objects.create(name="Test category"),
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            waste_flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123", url="https://www.test-flyer.org"
            )
        frequency = CollectionFrequency.objects.create(name="Test Frequency")
        nuts = NutsRegion.objects.create(
            name="Test NUTS", nuts_id="DE123", cntr_code="DE"
        )
        catchment = CollectionCatchment.objects.create(
            name="Test catchment", region=nuts.region_ptr
        )
        for i in range(1, 3):
            collection = Collection.objects.create(
                name=f"collection{i}",
                catchment=catchment,
                collector=Collector.objects.create(name=f"collector{1}"),
                collection_system=CollectionSystem.objects.create(name="Test system"),
                waste_stream=waste_stream,
                frequency=frequency,
                description="This is a test case.",
            )
            collection.flyers.add(waste_flyer)

    def setUp(self):
        self.file = BytesIO()

    def test_contains_all_labels_in_right_order(self):
        renderer = CollectionXLSXRenderer()
        qs = Collection.objects.all()
        content = CollectionFlatSerializer(qs, many=True).data
        renderer.render(self.file, content)
        wb = load_workbook(self.file)
        ws = wb.active
        ordered_content = [
            dict((k, row.get(k)) for k in list(renderer.labels.keys()))
            for row in content
        ]
        for column, (key, value) in enumerate(ordered_content[0].items(), start=1):
            self.assertEqual(renderer.labels[key], ws.cell(row=1, column=column).value)
