"""
Test cases for Collection review action cascade functionality.

ARCHITECTURE NOTE:
Cascade is a VIEW-LEVEL feature, not a model-level feature.
Direct model method calls (e.g., collection.submit_for_review()) do NOT cascade.
Only views using CollectionReviewActionCascadeMixin trigger cascade.

These tests verify that the mixin correctly cascades review actions to
CollectionPropertyValues and AggregatedCollectionPropertyValues when used in
Collection-specific review action views.

Tests use RequestFactory to simulate view-level calls and test the mixin logic
in isolation, bypassing URL routing.
"""

import codecs
import csv
import time
from collections import namedtuple
from datetime import date
from io import BytesIO
from unittest.mock import patch

from celery import chord
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ImproperlyConfigured
from django.db.models import signals
from django.http.request import MultiValueDict, QueryDict
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from factory.django import mute_signals
from openpyxl import load_workbook

from case_studies.soilcom.derived_values import (
    backfill_derived_values,
    clear_derived_value_config_cache,
    compute_counterpart_value,
    convert_specific_to_total_mg,
    convert_total_to_specific,
    create_or_update_derived_cpv,
    delete_derived_cpv,
    get_derived_property_config,
    get_population_for_collection,
)
from case_studies.soilcom.models import (
    AggregatedCollectionPropertyValue,
    Collection,
    CollectionCatchment,
    CollectionFrequency,
    CollectionPropertyValue,
    CollectionSystem,
    Collector,
    FeeSystem,
    WasteCategory,
    WasteComponent,
    WasteFlyer,
    WasteStream,
)
from case_studies.soilcom.renderers import CollectionCSVRenderer, CollectionXLSXRenderer
from case_studies.soilcom.serializers import CollectionFlatSerializer
from case_studies.soilcom.signals import (
    sync_derived_cpv_on_delete,
    sync_derived_cpv_on_save,
)
from case_studies.soilcom.tasks import (
    check_wasteflyer_url,
    check_wasteflyer_urls,
    check_wasteflyer_urls_callback,
)
from case_studies.soilcom.views import (
    CollectionApproveItemView,
    CollectionRejectItemView,
    CollectionSubmitForReviewView,
    CollectionWithdrawFromReviewView,
)
from case_studies.soilcom.waste_atlas.viewsets import (
    POPULATION_ATTRIBUTE_ID,
    _amounts_for_2024,
    _resolved_population_attribute_id,
)
from maps.models import Attribute, NutsRegion, Region, RegionAttributeValue
from materials.models import MaterialCategory
from utils.object_management.models import ReviewAction
from utils.properties.models import Property, Unit

User = get_user_model()


class CollectionCascadeMixinTestCase(TestCase):
    """Test the CollectionReviewActionCascadeMixin functionality."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        # Create users
        cls.owner = User.objects.create_user(
            username="owner", email="owner@test.com", password="pass"
        )
        cls.other_owner = User.objects.create_user(
            username="other", email="other@test.com", password="pass"
        )
        cls.staff = User.objects.create_user(
            username="staff", email="staff@test.com", password="pass", is_staff=True
        )

        # Create base data
        cls.catchment = CollectionCatchment.objects.create(
            name="Test Catchment", publication_status="published"
        )
        cls.system = CollectionSystem.objects.create(
            name="Test System", publication_status="published"
        )
        cls.category = WasteCategory.objects.create(
            name="Test Category", publication_status="published"
        )
        cls.stream = WasteStream.objects.create(
            category=cls.category, publication_status="published"
        )

        # Create properties
        cls.prop1 = Property.objects.create(
            name="Property 1", publication_status="published"
        )
        cls.prop2 = Property.objects.create(
            name="Property 2", publication_status="published"
        )
        cls.unit1 = Unit.objects.create(name="Unit 1", publication_status="published")
        cls.unit2 = Unit.objects.create(name="Unit 2", publication_status="published")
        cls.prop1.allowed_units.add(cls.unit1)
        cls.prop2.allowed_units.add(cls.unit2)

        cls.factory = RequestFactory()

    def _create_collection(self, name, owner, status="private", valid_from=None):
        """Helper to create a collection."""
        return Collection.objects.create(
            name=name,
            catchment=self.catchment,
            collection_system=self.system,
            waste_stream=self.stream,
            valid_from=valid_from or date(2020, 1, 1),
            publication_status=status,
            owner=owner,
        )

    def _create_cpv(self, collection, prop, unit, owner, status="private", **kwargs):
        """Helper to create a collection property value."""
        return CollectionPropertyValue.objects.create(
            collection=collection,
            property=prop,
            unit=unit,
            owner=owner,
            publication_status=status,
            year=kwargs.get("year", 2020),
            average=kwargs.get("average", 10.0),
        )

    def _create_acpv(self, collections, prop, unit, owner, status="private", **kwargs):
        """Helper to create an aggregated collection property value."""
        acpv = AggregatedCollectionPropertyValue.objects.create(
            property=prop,
            unit=unit,
            owner=owner,
            publication_status=status,
            year=kwargs.get("year", 2020),
            average=kwargs.get("average", 20.0),
        )
        acpv.collections.set(collections)
        return acpv


class SubmitCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test submit_for_review cascade via CollectionSubmitForReviewView."""

    def test_submit_cascades_to_owner_cpvs(self):
        """Submit cascades to owner's private and declined CPVs."""
        collection = self._create_collection("C1", self.owner, status="private")
        cpv_private = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "private", year=2020
        )
        cpv_declined = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "declined", year=2021
        )
        cpv_published = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "published", year=2022
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        cpv_private.refresh_from_db()
        cpv_declined.refresh_from_db()
        cpv_published.refresh_from_db()

        self.assertEqual(cpv_private.publication_status, "review")
        self.assertEqual(cpv_declined.publication_status, "review")
        self.assertEqual(cpv_published.publication_status, "published")  # Unchanged

    def test_submit_includes_collaborator_cpvs(self):
        """Submit cascades to all CPVs on owner's collection, including collaborators' CPVs."""
        collection = self._create_collection("C1", self.owner, status="private")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "private"
        )
        # Other user contributed a CPV to owner's collection
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "private"
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both cascade because they're on the owner's collection
        self.assertEqual(cpv_owner.publication_status, "review")
        self.assertEqual(cpv_other.publication_status, "review")  # Also cascaded

    def test_submit_cascades_to_acpvs(self):
        """Submit cascades to owner's aggregated property values."""
        collection = self._create_collection("C1", self.owner, status="private")
        acpv = self._create_acpv(
            [collection], self.prop1, self.unit1, self.owner, "private"
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        acpv.refresh_from_db()
        self.assertEqual(acpv.publication_status, "review")

    def test_submit_cascades_across_version_chain(self):
        """Submit cascades to CPVs on all versions in the chain."""
        v1 = self._create_collection("V1", self.owner, "published", date(2020, 1, 1))
        v2 = self._create_collection("V2", self.owner, "published", date(2021, 1, 1))
        v3 = self._create_collection("V3", self.owner, "private", date(2022, 1, 1))
        v2.predecessors.add(v1)
        v3.predecessors.add(v2)

        cpv1 = self._create_cpv(v1, self.prop1, self.unit1, self.owner, "private")
        cpv2 = self._create_cpv(v2, self.prop1, self.unit1, self.owner, "declined")
        cpv3 = self._create_cpv(v3, self.prop1, self.unit1, self.owner, "private")

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = v3
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        cpv1.refresh_from_db()
        cpv2.refresh_from_db()
        cpv3.refresh_from_db()

        self.assertEqual(cpv1.publication_status, "review")
        self.assertEqual(cpv2.publication_status, "review")
        self.assertEqual(cpv3.publication_status, "review")


class WithdrawCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test withdraw_from_review cascade via CollectionWithdrawFromReviewView."""

    def test_withdraw_cascades_to_owner_cpvs(self):
        """Withdraw cascades to owner's CPVs in review."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv = self._create_cpv(collection, self.prop1, self.unit1, self.owner, "review")

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionWithdrawFromReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "withdraw_from_review"
        view.post_action_hook(request, "review")

        cpv.refresh_from_db()
        self.assertEqual(cpv.publication_status, "private")

    def test_withdraw_includes_collaborator_cpvs(self):
        """Withdraw cascades to all CPVs on owner's collection, including collaborators'."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "review"
        )
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionWithdrawFromReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "withdraw_from_review"
        view.post_action_hook(request, "review")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both cascade because they're on the owner's collection
        self.assertEqual(cpv_owner.publication_status, "private")
        self.assertEqual(cpv_other.publication_status, "private")  # Also cascaded


class ApproveCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test approve cascade via CollectionApproveItemView."""

    def test_approve_cascades_to_all_cpvs_in_review(self):
        """Approve cascades to ALL CPVs in review, regardless of owner."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "review"
        )
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionApproveItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "approve"
        view.post_action_hook(request, "review")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both should be approved
        self.assertEqual(cpv_owner.publication_status, "published")
        self.assertEqual(cpv_other.publication_status, "published")
        self.assertEqual(cpv_owner.approved_by, self.staff)
        self.assertEqual(cpv_other.approved_by, self.staff)

    def test_approve_cascades_to_acpvs(self):
        """Approve cascades to aggregated property values."""
        collection = self._create_collection("C1", self.owner, status="review")
        acpv = self._create_acpv(
            [collection], self.prop1, self.unit1, self.owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionApproveItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "approve"
        view.post_action_hook(request, "review")

        acpv.refresh_from_db()
        self.assertEqual(acpv.publication_status, "published")
        self.assertEqual(acpv.approved_by, self.staff)

    def test_approve_cascades_across_version_chain(self):
        """Approve cascades to CPVs on all versions."""
        v1 = self._create_collection("V1", self.owner, "published", date(2020, 1, 1))
        v2 = self._create_collection("V2", self.owner, "review", date(2021, 1, 1))
        v2.predecessors.add(v1)

        cpv1 = self._create_cpv(v1, self.prop1, self.unit1, self.owner, "review")
        cpv2 = self._create_cpv(v2, self.prop1, self.unit1, self.owner, "review")

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionApproveItemView()
        view.request = request
        view.object = v2
        view.action_attr_name = "approve"
        view.post_action_hook(request, "review")

        cpv1.refresh_from_db()
        cpv2.refresh_from_db()

        self.assertEqual(cpv1.publication_status, "published")
        self.assertEqual(cpv2.publication_status, "published")


class RejectCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test reject cascade via CollectionRejectItemView."""

    def test_reject_cascades_to_all_cpvs_in_review(self):
        """Reject cascades to ALL CPVs in review, regardless of owner."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "review"
        )
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionRejectItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "reject"
        view.post_action_hook(request, "review")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both should be rejected
        self.assertEqual(cpv_owner.publication_status, "declined")
        self.assertEqual(cpv_other.publication_status, "declined")

    def test_reject_cascades_to_acpvs(self):
        """Reject cascades to aggregated property values."""
        collection = self._create_collection("C1", self.owner, status="review")
        acpv = self._create_acpv(
            [collection], self.prop1, self.unit1, self.owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionRejectItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "reject"
        view.post_action_hook(request, "review")

        acpv.refresh_from_db()
        self.assertEqual(acpv.publication_status, "declined")

    def test_reject_cascades_across_version_chain(self):
        """Reject cascades to CPVs on all versions."""
        v1 = self._create_collection("V1", self.owner, "published", date(2020, 1, 1))
        v2 = self._create_collection("V2", self.owner, "review", date(2021, 1, 1))
        v2.predecessors.add(v1)

        cpv1 = self._create_cpv(v1, self.prop1, self.unit1, self.owner, "review")
        cpv2 = self._create_cpv(v2, self.prop1, self.unit1, self.owner, "review")

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionRejectItemView()
        view.request = request
        view.object = v2
        view.action_attr_name = "reject"
        view.post_action_hook(request, "review")

        cpv1.refresh_from_db()
        cpv2.refresh_from_db()

        self.assertEqual(cpv1.publication_status, "declined")
        self.assertEqual(cpv2.publication_status, "declined")


class ReviewSubmissionConsistencyTests(TestCase):
    """Regression tests ensuring review-submission timestamps remain consistent."""

    def setUp(self):
        self.owner = User.objects.create_user(username="owner")
        self.moderator = User.objects.create_user(username="moderator", is_staff=True)

        content_type = ContentType.objects.get_for_model(Collection)
        permission, _ = Permission.objects.get_or_create(
            codename="can_moderate_collection",
            content_type=content_type,
            defaults={"name": "Can moderate collections"},
        )
        self.moderator.user_permissions.add(permission)

        self.collection = Collection.objects.create(
            name="Test Collection",
            owner=self.owner,
            publication_status=Collection.STATUS_PRIVATE,
        )

    def test_initial_submission_consistency(self):
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        self.collection.refresh_from_db()
        self.assertIsNotNone(self.collection.submitted_at)

        actions = ReviewAction.objects.filter(
            content_type=ContentType.objects.get_for_model(Collection),
            object_id=self.collection.pk,
            action=ReviewAction.ACTION_SUBMITTED,
        )
        self.assertEqual(actions.count(), 1)

        latest_action = actions.order_by("-created_at", "-id").first()
        self.assertIsNotNone(latest_action)

        time_diff = abs(self.collection.submitted_at - latest_action.created_at)
        self.assertLess(time_diff.total_seconds(), 1.0)

    def test_resubmission_consistency(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()
        initial_submitted_at = self.collection.submitted_at
        initial_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )

        response = self.client.post(
            reverse(
                "object_management:withdraw_from_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()
        self.assertIsNone(self.collection.submitted_at)

        time.sleep(0.01)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()

        latest_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )

        self.assertIsNotNone(self.collection.submitted_at)
        self.assertIsNotNone(latest_action)
        time_diff = abs(self.collection.submitted_at - latest_action.created_at)
        self.assertLess(time_diff.total_seconds(), 1.0)

        self.assertGreater(self.collection.submitted_at, initial_submitted_at)
        self.assertGreater(latest_action.created_at, initial_action.created_at)

    def test_review_ui_shows_latest_submission(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        time.sleep(0.01)

        response = self.client.post(
            reverse(
                "object_management:withdraw_from_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        time.sleep(0.01)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()

        self.client.force_login(self.moderator)
        response = self.client.get(
            reverse(
                "object_management:review_item_detail",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            )
        )
        self.assertEqual(response.status_code, 200)

        self.assertIn("review_submitted_action", response.context)
        submitted_action = response.context["review_submitted_action"]
        self.assertIsNotNone(submitted_action)
        self.assertEqual(submitted_action.action, ReviewAction.ACTION_SUBMITTED)

        latest_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )
        self.assertEqual(submitted_action.pk, latest_action.pk)

        time_diff = abs(submitted_action.created_at - self.collection.submitted_at)
        self.assertLess(time_diff.total_seconds(), 1.0)

    def test_approval_preserves_submission_timestamp(self):
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()
        submitted_at_before = self.collection.submitted_at

        self.client.force_login(self.moderator)
        response = self.client.post(
            reverse(
                "object_management:approve_item",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()

        self.assertEqual(self.collection.submitted_at, submitted_at_before)

        latest_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )
        time_diff = abs(self.collection.submitted_at - latest_action.created_at)
        self.assertLess(time_diff.total_seconds(), 1.0)


@patch("case_studies.soilcom.tests.test_regression.check_wasteflyer_urls.apply")
@patch("case_studies.soilcom.tests.test_regression.chord")
class CheckWasteFlyerUrlsTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        with mute_signals(signals.post_save):
            for i in range(1, 5):
                WasteFlyer.objects.create(
                    title=f"Waste flyer {i}",
                    abbreviation=f"WF{i}",
                    url_valid=i % 2 == 0,
                )

    def setUp(self):
        self.flyer = WasteFlyer.objects.first

    def test_initial(self, mock_chord, mock_apply):
        mock_async_result = namedtuple("MockAsyncResult", ["status", "get"])
        mock_apply.return_value = mock_async_result(status="SUCCESS", get=lambda: None)
        self.assertEqual(4, WasteFlyer.objects.count())
        params = {
            "csrfmiddlewaretoken": [
                "Hm7MXB2NjRCOIpNbGaRKR87VCHM5KwpR1t4AdZFgaqKfqui1EJwhKKmkxFKDfL3h"
            ],
            "url_valid": ["False"],
            "page": ["2"],
        }
        qdict = QueryDict("", mutable=True)
        qdict.update(MultiValueDict(params))
        newparams = qdict.copy()
        newparams.pop("csrfmiddlewaretoken")
        newparams.pop("page")
        result = check_wasteflyer_urls.apply(args=[newparams])
        while result.status == "PENDING":
            self.assertEqual("PENDING", result.status)
        if result.status == "FAILURE":
            result.get()
        self.assertEqual("SUCCESS", result.status)

    def test_chord(self, mock_chord, mock_apply):
        mock_chord.return_value = lambda x: type(
            "task", (object,), {"task_id": "fake_task_id"}
        )
        mock_apply.side_effect = [
            type("task", (object,), {"status": "SUCCESS"})
            for _ in WasteFlyer.objects.all()
        ]
        callback = check_wasteflyer_urls_callback.s()
        header = [
            check_wasteflyer_url.s(flyer.pk) for flyer in WasteFlyer.objects.all()
        ]
        result = chord(header)(callback)
        self.assertEqual(result.task_id, "fake_task_id")


@patch("case_studies.soilcom.tasks.find_wayback_snapshot_for_year")
@patch("case_studies.soilcom.tasks.check_url")
class CheckWasteFlyerUrlWaybackFallbackTestCase(TestCase):
    def setUp(self):
        with mute_signals(signals.post_save):
            self.flyer = WasteFlyer.objects.create(
                title="Waste flyer",
                abbreviation="WF",
                url="https://example.com/dead-flyer.pdf",
            )

        self.collection = Collection.objects.create(valid_from=date(2021, 1, 1))
        self.collection.flyers.add(self.flyer)

    def test_replaces_broken_url_with_year_snapshot(self, mock_check_url, mock_wayback):
        original_url = self.flyer.url
        mock_check_url.return_value = False
        mock_wayback.return_value = (
            "https://web.archive.org/web/20211230153000/"
            "https://example.com/dead-flyer.pdf"
        )

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertTrue(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, mock_wayback.return_value)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_keeps_original_url_when_no_snapshot_exists(
        self, mock_check_url, mock_wayback
    ):
        original_url = self.flyer.url
        mock_check_url.return_value = False
        mock_wayback.return_value = None

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertFalse(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, original_url)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_replaces_live_url_with_year_snapshot(self, mock_check_url, mock_wayback):
        original_url = self.flyer.url
        mock_check_url.return_value = True
        mock_wayback.return_value = (
            "https://web.archive.org/web/20211230153000/"
            "https://example.com/dead-flyer.pdf"
        )

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertTrue(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, mock_wayback.return_value)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_keeps_live_url_when_no_snapshot_exists(self, mock_check_url, mock_wayback):
        original_url = self.flyer.url
        mock_check_url.return_value = True
        mock_wayback.return_value = None

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertTrue(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, original_url)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_skips_wayback_lookup_when_url_is_already_archived(
        self, mock_check_url, mock_wayback
    ):
        with mute_signals(signals.post_save):
            self.flyer.url = (
                "https://web.archive.org/web/20211230153000/"
                "https://example.com/dead-flyer.pdf"
            )
            self.flyer.save()
        mock_check_url.return_value = True

        check_wasteflyer_url(self.flyer.pk)

        mock_wayback.assert_not_called()

    def test_returns_false_when_wasteflyer_was_deleted(
        self, mock_check_url, mock_wayback
    ):
        flyer_pk = self.flyer.pk
        self.flyer.delete()

        result = check_wasteflyer_url(flyer_pk)

        self.assertFalse(result)
        mock_check_url.assert_not_called()
        mock_wayback.assert_not_called()


class CollectionCSVRendererTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        MaterialCategory.objects.create(name="Biowaste component")
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed Material 1"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed Material 2"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden Material 1"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden Material 2"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=WasteCategory.objects.create(name="Test category"),
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            waste_flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123", url="https://www.test-flyer.org"
            )
        frequency = CollectionFrequency.objects.create(name="Test Frequency")
        nuts = NutsRegion.objects.create(
            name="Test NUTS", nuts_id="DE123", cntr_code="DE"
        )
        catchment = CollectionCatchment.objects.create(
            name="Test catchment", region=nuts.region_ptr
        )
        for i in range(1, 3):
            collection = Collection.objects.create(
                name=f"collection{i}",
                catchment=catchment,
                collector=Collector.objects.create(name=f"collector{1}"),
                collection_system=CollectionSystem.objects.create(name="Test system"),
                waste_stream=waste_stream,
                fee_system=FeeSystem.objects.create(name="Fixed fee"),
                frequency=frequency,
                valid_from=date(2020, 1, 1),
                description="This is a test case.",
            )
            collection.flyers.add(waste_flyer)

    def setUp(self):
        self.file = BytesIO()
        self.content = CollectionFlatSerializer(
            Collection.objects.all(), many=True
        ).data

    def test_fieldnames_in_right_order(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        fieldnames = [renderer.labels[key] for key in renderer.header]
        self.assertListEqual(fieldnames, list(reader.fieldnames))
        self.assertEqual(2, sum(1 for _ in reader))
        self.assertIn("Connection type", reader.fieldnames)
        self.assertEqual(
            renderer.header.index("connection_type"),
            reader.fieldnames.index("Connection type"),
        )

    def test_connection_type_field_exported(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        valid_labels = [
            "Compulsory",
            "Voluntary",
            "Mandatory",
            "Mandatory with exception for home composters",
            "Not specified",
            "",
        ]
        for row in reader:
            self.assertIn(row["Connection type"], valid_labels)

    def test_allowed_materials_formatted_as_comma_separated_list_in_one_field(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertEqual(
                "Allowed Material 1, Allowed Material 2", row["Allowed Materials"]
            )

    def test_forbidden_materials_formatted_as_comma_separated_list_in_one_field(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertEqual(
                "Forbidden Material 1, Forbidden Material 2", row["Forbidden Materials"]
            )

    def test_regression_flyers_without_urls_dont_raise_type_error(self):
        defected_collection = Collection.objects.first()
        with mute_signals(signals.post_save):
            rogue_flyer = WasteFlyer.objects.create(
                title="Rogue flyer without url", abbreviation="RF"
            )
        defected_collection.flyers.add(rogue_flyer)
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        self.assertEqual(Collection.objects.count(), len(list(reader)))


class CollectionXLSXRendererTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        User.objects.create(username="outsider")
        member = User.objects.create(username="member")
        content_type = ContentType.objects.get_for_model(Collection)
        permission, _ = Permission.objects.get_or_create(
            codename="add_collection",
            content_type=content_type,
            defaults={"name": "Can add collection"},
        )
        member.user_permissions.add(permission)

        MaterialCategory.objects.create(name="Biowaste component")
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed Material 1"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed Material 2"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden Material 1"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden Material 2"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=WasteCategory.objects.create(name="Test category"),
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            waste_flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123", url="https://www.test-flyer.org"
            )
        frequency = CollectionFrequency.objects.create(name="Test Frequency")
        nuts = NutsRegion.objects.create(
            name="Test NUTS", nuts_id="DE123", cntr_code="DE"
        )
        catchment = CollectionCatchment.objects.create(
            name="Test catchment", region=nuts.region_ptr
        )
        for i in range(1, 3):
            collection = Collection.objects.create(
                name=f"collection{i}",
                catchment=catchment,
                collector=Collector.objects.create(name=f"collector{1}"),
                collection_system=CollectionSystem.objects.create(name="Test system"),
                waste_stream=waste_stream,
                frequency=frequency,
                description="This is a test case.",
            )
            collection.flyers.add(waste_flyer)

    def setUp(self):
        self.file = BytesIO()

    def test_contains_all_labels_in_right_order(self):
        renderer = CollectionXLSXRenderer()
        qs = Collection.objects.all()
        content = CollectionFlatSerializer(qs, many=True).data
        renderer.render(self.file, content)
        wb = load_workbook(self.file)
        ws = wb.active
        ordered_content = [
            {k: row.get(k) for k in list(renderer.labels.keys())} for row in content
        ]
        for column, (key, _value) in enumerate(ordered_content[0].items(), start=1):
            self.assertEqual(renderer.labels[key], ws.cell(row=1, column=column).value)


@override_settings(
    SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=None,
    SOILCOM_TOTAL_WASTE_PROPERTY_ID=None,
    SOILCOM_SPECIFIC_WASTE_UNIT_ID=None,
    SOILCOM_TOTAL_WASTE_UNIT_ID=None,
    SOILCOM_POPULATION_ATTRIBUTE_ID=None,
    SOILCOM_SPECIFIC_WASTE_PROPERTY_NAME="specific waste collected [test]",
    SOILCOM_TOTAL_WASTE_PROPERTY_NAME="total waste collected [test]",
    SOILCOM_SPECIFIC_WASTE_UNIT_NAME="kg/(cap.*a) [test]",
    SOILCOM_TOTAL_WASTE_UNIT_NAME="Mg/a [test]",
    SOILCOM_POPULATION_ATTRIBUTE_NAME="Population [test]",
)
class DerivedValuesTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.property_specific = Property.objects.create(
            name="specific waste collected [test]"
        )
        cls.property_total = Property.objects.create(
            name="total waste collected [test]"
        )
        cls.unit_specific = Unit.objects.create(name="kg/(cap.*a) [test]")
        cls.unit_total = Unit.objects.create(name="Mg/a [test]")
        cls.population_attribute = Attribute.objects.create(
            name="Population [test]",
            unit="cap",
        )

        cls.collection_system = CollectionSystem.objects.create(
            name="Collection system"
        )
        cls.waste_category = WasteCategory.objects.create(name="Waste category")
        cls.waste_stream = WasteStream.objects.create(
            name="Waste stream",
            category=cls.waste_category,
        )

    def setUp(self):
        clear_derived_value_config_cache()

    def tearDown(self):
        clear_derived_value_config_cache()

    def _create_collection(self, suffix, *, population=None):
        region = Region.objects.create(name=f"Region {suffix}", country="DE")
        catchment = CollectionCatchment.objects.create(
            name=f"Catchment {suffix}",
            region=region,
        )
        collection = Collection.objects.create(
            catchment=catchment,
            collection_system=self.collection_system,
            waste_stream=self.waste_stream,
            valid_from=date(2024, 1, 1),
            publication_status="published",
        )
        if population is not None:
            RegionAttributeValue.objects.create(
                name=f"Population {suffix}",
                region=region,
                attribute=self.population_attribute,
                date=date(2024, 1, 1),
                value=population,
            )
        return collection, catchment

    @staticmethod
    def _bulk_create_cpv(**kwargs):
        cpv = CollectionPropertyValue(**kwargs)
        CollectionPropertyValue.objects.bulk_create([cpv])
        return cpv

    def test_create_or_update_removes_stale_derived_when_manual_exists(self):
        collection, _catchment = self._create_collection("stale", population=2000)

        specific = self._bulk_create_cpv(
            name="specific source",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10,
            publication_status="published",
            is_derived=False,
        )

        _derived, action = create_or_update_derived_cpv(specific)
        self.assertEqual(action, "created")
        self.assertTrue(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

        self._bulk_create_cpv(
            name="manual total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=25,
            publication_status="published",
            is_derived=False,
        )

        _derived, action = create_or_update_derived_cpv(specific)
        self.assertEqual(action, "skipped")
        self.assertFalse(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

    def test_backfill_counts_created_updated_and_skipped(self):
        collection_with_population, _ = self._create_collection(
            "has-pop", population=1000
        )
        collection_without_population, _ = self._create_collection(
            "no-pop", population=None
        )

        self._bulk_create_cpv(
            name="create-source",
            collection=collection_with_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="update-source",
            collection=collection_with_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2025,
            average=20,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="existing-derived-total",
            collection=collection_with_population,
            property=self.property_total,
            unit=self.unit_total,
            year=2025,
            average=20,
            publication_status="published",
            is_derived=True,
        )
        self._bulk_create_cpv(
            name="skip-manual-specific",
            collection=collection_with_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2026,
            average=30,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="manual-total",
            collection=collection_with_population,
            property=self.property_total,
            unit=self.unit_total,
            year=2026,
            average=30,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="skip-no-pop",
            collection=collection_without_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=40,
            publication_status="published",
            is_derived=False,
        )

        dry_stats = backfill_derived_values(dry_run=True)
        self.assertEqual(dry_stats, {"created": 1, "updated": 1, "skipped": 3})

        write_stats = backfill_derived_values(dry_run=False)
        self.assertEqual(write_stats, {"created": 1, "updated": 1, "skipped": 3})

    def test_amounts_for_2024_falls_back_to_total_and_population(self):
        collection, catchment = self._create_collection(
            "atlas-fallback", population=2500
        )

        self._bulk_create_cpv(
            name="total-only",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=6.5,
            publication_status="published",
            is_derived=False,
        )

        amounts = _amounts_for_2024(
            year=2024,
            all_collection_ids={collection.pk},
            col_to_cid={collection.pk: catchment.pk},
            catchment_ids=[catchment.pk],
        )
        self.assertEqual(amounts[catchment.pk], round(6.5 * 1000 / 2500, 1))

    def test_compute_counterpart_value_returns_none_for_non_convertible_property(self):
        other_property = Property.objects.create(name="other property [test]")
        collection, _ = self._create_collection("other", population=1000)
        cpv = self._bulk_create_cpv(
            name="other-cpv",
            collection=collection,
            property=other_property,
            unit=self.unit_specific,
            year=2024,
            average=1.0,
            publication_status="published",
            is_derived=False,
        )
        self.assertIsNone(compute_counterpart_value(cpv))

    def test_compute_counterpart_value_returns_none_without_population(self):
        collection, _ = self._create_collection("no-pop-compute", population=None)
        cpv = self._bulk_create_cpv(
            name="specific-no-pop",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=12.3,
            publication_status="published",
            is_derived=False,
        )
        self.assertIsNone(compute_counterpart_value(cpv))

    def test_create_or_update_skips_when_input_is_already_derived(self):
        collection, _ = self._create_collection("already-derived", population=1000)
        derived_source = self._bulk_create_cpv(
            name="derived-source",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=True,
        )
        derived, action = create_or_update_derived_cpv(derived_source)
        self.assertIsNone(derived)
        self.assertEqual(action, "skipped")

    def test_conversion_helpers_guard_invalid_population(self):
        self.assertIsNone(convert_specific_to_total_mg(10, 0))
        self.assertIsNone(convert_specific_to_total_mg(10, -5))
        self.assertIsNone(convert_total_to_specific(10, 0))
        self.assertIsNone(convert_total_to_specific(10, -5))

    def test_get_derived_property_config_raises_for_invalid_configured_id(self):
        with override_settings(SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=999999):
            clear_derived_value_config_cache()
            with self.assertRaises(ImproperlyConfigured):
                get_derived_property_config()

    def test_get_derived_property_config_raises_for_ambiguous_names(self):
        Property.objects.create(name="specific waste collected [test]")
        with override_settings(SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=None):
            clear_derived_value_config_cache()
            with self.assertRaises(ImproperlyConfigured):
                get_derived_property_config()

    def test_signals_swallow_improperly_configured_and_do_not_raise(self):
        collection, _ = self._create_collection("signal-noise", population=1000)
        cpv = self._bulk_create_cpv(
            name="signal-cpv",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=15.0,
            publication_status="published",
            is_derived=False,
        )

        with override_settings(SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=999999):
            clear_derived_value_config_cache()
            # Should not raise; handler catches ImproperlyConfigured.
            sync_derived_cpv_on_save(sender=CollectionPropertyValue, instance=cpv)
            sync_derived_cpv_on_delete(sender=CollectionPropertyValue, instance=cpv)

    def test_population_attribute_resolution_uses_fallback_when_misconfigured(self):
        with override_settings(SOILCOM_POPULATION_ATTRIBUTE_ID=999999):
            clear_derived_value_config_cache()
            self.assertEqual(
                _resolved_population_attribute_id(), POPULATION_ATTRIBUTE_ID
            )

    def test_convert_specific_to_total_mg_happy_path(self):
        self.assertEqual(convert_specific_to_total_mg(50, 2000), 100.0)
        self.assertEqual(convert_specific_to_total_mg(10, 500), 5.0)

    def test_convert_total_to_specific_happy_path(self):
        self.assertEqual(convert_total_to_specific(100, 2000), 50.0)
        self.assertEqual(convert_total_to_specific(5, 500), 10.0)

    def test_conversion_helpers_ndigits_none_returns_exact_float(self):
        result = convert_specific_to_total_mg(7, 3000, ndigits=None)
        self.assertEqual(result, 7 * 3000 / 1000)
        result = convert_total_to_specific(7, 3000, ndigits=None)
        self.assertEqual(result, 7 * 1000 / 3000)

    def test_conversion_helpers_respect_ndigits(self):
        self.assertEqual(convert_specific_to_total_mg(7, 3000, ndigits=1), 21.0)
        self.assertEqual(convert_total_to_specific(6.5, 2500, ndigits=1), 2.6)

    def test_compute_counterpart_specific_to_total(self):
        collection, _ = self._create_collection("s2t", population=5000)
        cpv = self._bulk_create_cpv(
            name="specific-s2t",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=80.0,
            publication_status="published",
            is_derived=False,
        )
        result = compute_counterpart_value(cpv)
        self.assertIsNotNone(result)
        target_prop_id, target_unit_id, computed = result
        self.assertEqual(target_prop_id, self.property_total.pk)
        self.assertEqual(target_unit_id, self.unit_total.pk)
        self.assertEqual(computed, 400.0)

    def test_compute_counterpart_total_to_specific(self):
        collection, _ = self._create_collection("t2s", population=4000)
        cpv = self._bulk_create_cpv(
            name="total-t2s",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=120.0,
            publication_status="published",
            is_derived=False,
        )
        result = compute_counterpart_value(cpv)
        self.assertIsNotNone(result)
        target_prop_id, target_unit_id, computed = result
        self.assertEqual(target_prop_id, self.property_specific.pk)
        self.assertEqual(target_unit_id, self.unit_specific.pk)
        self.assertEqual(computed, 30.0)

    def test_create_derived_total_stores_correct_average(self):
        collection, _ = self._create_collection("val-s2t", population=2000)
        cpv = self._bulk_create_cpv(
            name="source-specific",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        derived, action = create_or_update_derived_cpv(cpv)
        self.assertEqual(action, "created")
        self.assertIsNotNone(derived)
        self.assertEqual(derived.average, 20.0)
        self.assertEqual(derived.property_id, self.property_total.pk)
        self.assertEqual(derived.unit_id, self.unit_total.pk)
        self.assertTrue(derived.is_derived)

    def test_create_derived_specific_from_total_stores_correct_average(self):
        collection, _ = self._create_collection("val-t2s", population=5000)
        cpv = self._bulk_create_cpv(
            name="source-total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=250.0,
            publication_status="published",
            is_derived=False,
        )
        derived, action = create_or_update_derived_cpv(cpv)
        self.assertEqual(action, "created")
        self.assertIsNotNone(derived)
        self.assertEqual(derived.average, 50.0)
        self.assertEqual(derived.property_id, self.property_specific.pk)
        self.assertEqual(derived.unit_id, self.unit_specific.pk)
        self.assertTrue(derived.is_derived)

    def test_create_or_update_updates_existing_derived_value(self):
        collection, _ = self._create_collection("update-val", population=1000)
        cpv = self._bulk_create_cpv(
            name="source-update",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        derived1, action1 = create_or_update_derived_cpv(cpv)
        self.assertEqual(action1, "created")
        self.assertEqual(derived1.average, 10.0)

        cpv.average = 20.0
        cpv.save()
        derived2, action2 = create_or_update_derived_cpv(cpv)
        self.assertEqual(action2, "updated")
        self.assertEqual(derived2.average, 20.0)
        self.assertEqual(derived1.pk, derived2.pk)

    def test_delete_derived_cpv_removes_counterpart(self):
        collection, _ = self._create_collection("del", population=1000)
        source = self._bulk_create_cpv(
            name="source-del",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        create_or_update_derived_cpv(source)
        self.assertTrue(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

        count = delete_derived_cpv(source)
        self.assertEqual(count, 1)
        self.assertFalse(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

    def test_delete_derived_cpv_does_not_remove_manual_counterpart(self):
        collection, _ = self._create_collection("del-manual", population=1000)
        source = self._bulk_create_cpv(
            name="source-del-m",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="manual-total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=99.0,
            publication_status="published",
            is_derived=False,
        )

        count = delete_derived_cpv(source)
        self.assertEqual(count, 0)
        self.assertTrue(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=False,
            ).exists()
        )

    def test_delete_derived_cpv_skips_when_source_is_derived(self):
        collection, _ = self._create_collection("del-skip", population=1000)
        derived_source = self._bulk_create_cpv(
            name="derived-del",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=True,
        )
        self.assertEqual(delete_derived_cpv(derived_source), 0)

    def test_backfill_dry_run_does_not_write(self):
        collection, _ = self._create_collection("dryrun", population=1000)
        self._bulk_create_cpv(
            name="dry-source",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        derived_before = CollectionPropertyValue.objects.filter(is_derived=True).count()
        backfill_derived_values(dry_run=True)
        derived_after = CollectionPropertyValue.objects.filter(is_derived=True).count()
        self.assertEqual(derived_before, derived_after)

    def test_backfill_writes_correct_values(self):
        collection, _ = self._create_collection("bf-val", population=2000)
        self._bulk_create_cpv(
            name="bf-specific",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=60.0,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="bf-total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2025,
            average=50.0,
            publication_status="published",
            is_derived=False,
        )

        backfill_derived_values(dry_run=False)

        derived_total = CollectionPropertyValue.objects.get(
            collection=collection,
            property=self.property_total,
            year=2024,
            is_derived=True,
        )
        self.assertEqual(derived_total.average, 120.0)

        derived_specific = CollectionPropertyValue.objects.get(
            collection=collection,
            property=self.property_specific,
            year=2025,
            is_derived=True,
        )
        self.assertEqual(derived_specific.average, 25.0)

    def test_get_population_for_collection_returns_exact_year(self):
        collection, _ = self._create_collection("pop-year", population=None)
        region = collection.catchment.region
        RegionAttributeValue.objects.create(
            name="Pop 2023",
            region=region,
            attribute=self.population_attribute,
            date=date(2023, 1, 1),
            value=3000,
        )
        RegionAttributeValue.objects.create(
            name="Pop 2024",
            region=region,
            attribute=self.population_attribute,
            date=date(2024, 6, 15),
            value=3500,
        )
        self.assertEqual(get_population_for_collection(collection, year=2024), 3500)
        self.assertEqual(get_population_for_collection(collection, year=2023), 3000)

    def test_get_population_for_collection_falls_back_to_most_recent(self):
        collection, _ = self._create_collection("pop-fallback", population=None)
        region = collection.catchment.region
        RegionAttributeValue.objects.create(
            name="Pop old",
            region=region,
            attribute=self.population_attribute,
            date=date(2020, 1, 1),
            value=1000,
        )
        RegionAttributeValue.objects.create(
            name="Pop newer",
            region=region,
            attribute=self.population_attribute,
            date=date(2022, 1, 1),
            value=2000,
        )
        self.assertEqual(get_population_for_collection(collection, year=2024), 2000)

    def test_get_population_for_collection_returns_none_without_data(self):
        collection, _ = self._create_collection("pop-none", population=None)
        self.assertIsNone(get_population_for_collection(collection, year=2024))
