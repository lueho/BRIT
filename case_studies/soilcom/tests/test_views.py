import codecs
import csv
import time
from collections import namedtuple
from datetime import date, timedelta
from io import BytesIO
from unittest.mock import patch
from urllib.parse import urlencode

from celery import chord
from django.contrib.auth.models import Group, Permission, User
from django.contrib.contenttypes.models import ContentType
from django.contrib.gis.geos import MultiPolygon, Polygon
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.db.models import signals
from django.db.models.signals import post_save
from django.forms.formsets import BaseFormSet
from django.http import JsonResponse
from django.http.request import MultiValueDict, QueryDict
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from factory.django import mute_signals
from openpyxl import load_workbook

from bibliography.models import Source
from case_studies.soilcom.derived_values import (
    backfill_derived_values,
    clear_derived_value_config_cache,
    compute_counterpart_value,
    convert_specific_to_total_mg,
    convert_total_to_specific,
    create_or_update_derived_cpv,
    delete_derived_cpv,
    get_derived_property_config,
    get_population_for_collection,
)
from case_studies.soilcom.importers import CollectionImporter
from case_studies.soilcom.models import (
    Collection,
    CollectionCatchment,
    CollectionPropertyValue,
    CollectionSystem,
    Collector,
    MaterialCategory,  # soilcom.MaterialCategory
    WasteCategory,
    WasteComponent,
    WasteFlyer,
    WasteStream,
    check_url_valid,
)
from case_studies.soilcom.renderers import CollectionCSVRenderer, CollectionXLSXRenderer
from case_studies.soilcom.serializers import CollectionFlatSerializer
from case_studies.soilcom.signals import (
    sync_derived_cpv_on_delete,
    sync_derived_cpv_on_save,
)
from case_studies.soilcom.tasks import (
    check_wasteflyer_url,
    check_wasteflyer_urls,
    check_wasteflyer_urls_callback,
    cleanup_orphaned_waste_streams,
)
from case_studies.soilcom.views import (
    CollectionApproveItemView,
    CollectionRejectItemView,
    CollectionSubmitForReviewView,
    CollectionWithdrawFromReviewView,
)
from case_studies.soilcom.waste_atlas.viewsets import (
    POPULATION_ATTRIBUTE_ID,
    _amounts_for_2024,
    _resolved_population_attribute_id,
)
from distributions.models import TemporalDistribution, Timestep
from maps.models import (
    Attribute,
    GeoDataset,
    MapConfiguration,
    MapLayerConfiguration,
    MapLayerStyle,
    NutsRegion,
    Region,
    RegionAttributeValue,
)
from materials.models import (
    Material,
    Sample,
    SampleSeries,
)
from utils.object_management.models import ReviewAction
from utils.properties.models import Property, Unit
from utils.tests.testcases import AbstractTestCases, ViewWithPermissionsTestCase

from .. import views
from ..forms import CollectionModelForm, WasteFlyerFormSet
from ..models import (
    AggregatedCollectionPropertyValue,
    CollectionCountOptions,
    CollectionFrequency,
    CollectionSeason,
    FeeSystem,
    SortingMethod,
)
from ..tasks import cleanup_orphaned_waste_flyers


def setUpModule():
    post_save.disconnect(check_url_valid, sender=WasteFlyer)


def tearDownModule():
    post_save.connect(check_url_valid, sender=WasteFlyer)


# ----------- Collector CRUD -------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectorCRUDViewsTestCase(AbstractTestCases.UserCreatedObjectCRUDViewTestCase):
    modal_detail_view = True
    modal_update_view = True
    modal_create_view = True

    model = Collector

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "collector-create"
    view_modal_create_name = "collector-create-modal"
    view_published_list_name = "collector-list"
    view_private_list_name = "collector-list-owned"
    view_detail_name = "collector-detail"
    view_modal_detail_name = "collector-detail-modal"
    view_update_name = "collector-update"
    view_modal_update_name = "collector-update-modal"
    view_delete_name = "collector-delete-modal"

    add_scope_query_param_to_list_urls = True

    create_object_data = {"name": "Test Collector"}
    update_object_data = {"name": "Updated Test Collector"}


# ----------- Collection System CRUD -----------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectionSystemCRUDViewsTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    modal_detail_view = True
    modal_update_view = True
    modal_create_view = True
    add_scope_query_param_to_list_urls = True

    model = CollectionSystem

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "collectionsystem-create"
    view_modal_create_name = "collectionsystem-create-modal"
    view_published_list_name = "collectionsystem-list"
    view_private_list_name = "collectionsystem-list-owned"
    view_detail_name = "collectionsystem-detail"
    view_modal_detail_name = "collectionsystem-detail-modal"
    view_update_name = "collectionsystem-update"
    view_modal_update_name = "collectionsystem-update-modal"
    view_delete_name = "collectionsystem-delete-modal"

    create_object_data = {"name": "Test Collection System"}
    update_object_data = {"name": "Updated Test Collection System"}


# ----------- Waste Category CRUD --------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class WasteCategoryCRUDViewsTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    modal_detail_view = True
    modal_update_view = True
    modal_create_view = True
    add_scope_query_param_to_list_urls = True

    model = WasteCategory

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "wastecategory-create"
    view_modal_create_name = "wastecategory-create-modal"
    view_published_list_name = "wastecategory-list"
    view_private_list_name = "wastecategory-list-owned"
    view_detail_name = "wastecategory-detail"
    view_modal_detail_name = "wastecategory-detail-modal"
    view_update_name = "wastecategory-update"
    view_modal_update_name = "wastecategory-update-modal"
    view_delete_name = "wastecategory-delete-modal"

    create_object_data = {"name": "Test Waste Category"}
    update_object_data = {"name": "Updated Test Waste Category"}


# ----------- Waste Component CRUD -------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class WasteComponentCRUDViewsTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    modal_detail_view = True
    modal_update_view = True
    modal_create_view = True
    add_scope_query_param_to_list_urls = True

    model = WasteComponent

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "wastecomponent-create"
    view_modal_create_name = "wastecomponent-create-modal"
    view_published_list_name = "wastecomponent-list"
    view_private_list_name = "wastecomponent-list-owned"
    view_detail_name = "wastecomponent-detail"
    view_modal_detail_name = "wastecomponent-detail-modal"
    view_update_name = "wastecomponent-update"
    view_modal_update_name = "wastecomponent-update-modal"
    view_delete_name = "wastecomponent-delete-modal"

    create_object_data = {"name": "Test Waste Component"}
    update_object_data = {"name": "Updated Test Waste Component"}

    @classmethod
    def create_related_objects(cls):
        MaterialCategory.objects.create(
            name="Biowaste component", publication_status="published"
        )
        return {}

    @classmethod
    def create_published_object(cls):
        # This method is overridden to give another name to the published object because of the unique name constraint
        data = cls.create_object_data.copy()
        data["name"] = f"{data['name']} (published)"
        data["publication_status"] = "published"
        data.update(cls.related_objects)
        return cls.model.objects.create(owner=cls.owner_user, **data)


# ----------- Fee System CRUD ------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class FeeSystemCRUDViewsTestCase(AbstractTestCases.UserCreatedObjectCRUDViewTestCase):
    model = FeeSystem
    add_scope_query_param_to_list_urls = True

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "feesystem-create"
    view_published_list_name = "feesystem-list"
    view_private_list_name = "feesystem-list-owned"
    view_detail_name = "feesystem-detail"
    view_update_name = "feesystem-update"
    view_delete_name = "feesystem-delete-modal"

    create_object_data = {"name": "Test Fee System"}
    update_object_data = {"name": "Updated Test Fee System"}


# ----------- WasteFlyer CRUD ------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class WasteFlyerCRUDViewsTestCase(AbstractTestCases.UserCreatedObjectCRUDViewTestCase):
    create_view = False
    modal_detail_view = True
    update_view = False
    delete_view = False
    add_scope_query_param_to_list_urls = True

    model = WasteFlyer

    view_dashboard_name = "wastecollection-explorer"
    view_published_list_name = "wasteflyer-list"
    view_private_list_name = "wasteflyer-list-owned"
    view_detail_name = "wasteflyer-detail"
    view_modal_detail_name = "wasteflyer-detail-modal"

    create_object_data = {"url": "https://www.crud-test-flyer.org"}

    def test_list_unpublished_contains_check_urls_button_for_authenticated_owner(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(reverse("wasteflyer-list-owned"), follow=True)
        self.assertContains(response, "Check URLs")

    def test_list_published_contains_check_urls_button_for_staff_user(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(reverse("wasteflyer-list"), follow=True)
        self.assertContains(response, "Check URLs")

    def test_detail_view_unpublished_contains_check_url_button_for_owner(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_detail_url(self.unpublished_object.pk))
        self.assertContains(response, "check url")

    def test_detail_view_published_contains_check_url_button_for_owner(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_detail_url(self.published_object.pk))
        self.assertContains(response, "check url")

    def test_detail_view_published_doesnt_contain_check_url_button_for_anonymous(self):
        response = self.client.get(self.get_detail_url(self.published_object.pk))
        self.assertNotContains(response, "Check URLs")

    def test_detail_view_published_doesnt_contain_check_url_button_for_non_owner(self):
        self.client.force_login(self.non_owner_user)
        response = self.client.get(self.get_detail_url(self.published_object.pk))
        self.assertNotContains(response, "Check URLs")

    def test_create_wasteflyer_without_author_succeeds(self):
        self.client.force_login(self.user_with_add_perm)
        post_data = {
            "url": "https://example.com/waste-flyer.pdf",
            "sourceauthors-TOTAL_FORMS": 1,
            "sourceauthors-INITIAL_FORMS": 0,
            "sourceauthors-0-id": "",
            "sourceauthors-0-source": "",
            "sourceauthors-0-author": "",
        }

        response = self.client.post(
            reverse("wasteflyer-create"), post_data, follow=True
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            WasteFlyer.objects.filter(
                url="https://example.com/waste-flyer.pdf"
            ).exists()
        )

    @patch("case_studies.soilcom.views.check_wasteflyer_url")
    def test_check_url_view_dispatches_wasteflyer_task(self, mock_task):
        mock_task.delay.return_value.task_id = "fake-task-id"
        self.client.force_login(self.owner_user)
        response = self.client.get(
            reverse("wasteflyer-check-url", kwargs={"pk": self.published_object.pk})
        )
        self.assertEqual(response.status_code, 200)
        mock_task.delay.assert_called_once_with(self.published_object.pk)


# ----------- Collection Frequency CRUD --------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectionFrequencyCRUDViewsTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    modal_detail_view = True
    modal_update_view = True
    add_scope_query_param_to_list_urls = True

    model = CollectionFrequency

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "collectionfrequency-create"
    view_published_list_name = "collectionfrequency-list"
    view_private_list_name = "collectionfrequency-list-owned"
    view_detail_name = "collectionfrequency-detail"
    view_modal_detail_name = "collectionfrequency-detail-modal"
    view_update_name = "collectionfrequency-update"
    view_modal_update_name = "collectionfrequency-update-modal"
    view_delete_name = "collectionfrequency-delete-modal"

    create_object_data = {"name": "Test Frequency"}
    update_object_data = {"name": "Updated Test Frequency"}

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.distribution = TemporalDistribution.objects.get(name="Months of the year")
        cls.january = Timestep.objects.get(name="January")
        cls.may = Timestep.objects.get(name="May")
        cls.june = Timestep.objects.get(name="June")
        cls.december = Timestep.objects.get(name="December")
        season_1 = CollectionSeason.objects.create(
            distribution=cls.distribution,
            first_timestep=cls.january,
            last_timestep=cls.may,
        )
        season_2 = CollectionSeason.objects.create(
            distribution=cls.distribution,
            first_timestep=cls.june,
            last_timestep=cls.december,
        )
        cls.options_1 = CollectionCountOptions.objects.create(
            frequency=cls.unpublished_object,
            season=season_1,
            standard=100,
            option_1=150,
        )
        cls.options_2 = CollectionCountOptions.objects.create(
            frequency=cls.unpublished_object, season=season_2, standard=150
        )

    def related_objects_post_data(self):
        data = {
            "name": "Test Frequency with Seasons",
            "type": "Fixed-Seasonal",
            "form-INITIAL_FORMS": 1,
            "form-TOTAL_FORMS": 2,
            "form-0-distribution": self.distribution.id,
            "form-0-first_timestep": self.january.id,
            "form-0-last_timestep": self.may.id,
            "form-0-standard": 150,
            "form-0-option_1": "",
            "form-0-option_2": "",
            "form-0-option_3": "",
            "form-1-distribution": self.distribution.id,
            "form-1-first_timestep": self.june.id,
            "form-1-last_timestep": self.december.id,
            "form-1-standard": 200,
            "form-1-option_1": "",
            "form-1-option_2": "",
            "form-1-option_3": "",
        }
        return data

    def test_get_formset_queryset_returns_whole_year_season(self):
        request = RequestFactory().get(self.get_create_url())
        request.user = self.staff_user
        view = views.FrequencyCreateView()
        view.setup(request)
        months = TemporalDistribution.objects.get(name="Months of the year")
        first = months.timestep_set.get(name="January")
        last = months.timestep_set.get(name="December")
        initial = list(
            CollectionSeason.objects.filter(
                distribution=months, first_timestep=first, last_timestep=last
            ).values("distribution", "first_timestep", "last_timestep")
        )
        self.assertListEqual(initial, view.get_formset_initial())

    def test_post_with_valid_data_creates_and_relates_seasons(self):
        self.client.force_login(self.staff_user)
        data = {
            "name": "Test Frequency with Seasons",
            "type": "Fixed-Seasonal",
            "form-INITIAL_FORMS": 1,
            "form-TOTAL_FORMS": 2,
            "form-0-distribution": self.distribution.id,
            "form-0-first_timestep": Timestep.objects.get(
                distribution=self.distribution, name="January"
            ).id,
            "form-0-last_timestep": Timestep.objects.get(
                distribution=self.distribution, name="April"
            ).id,
            "form-1-distribution": self.distribution.id,
            "form-1-first_timestep": Timestep.objects.get(
                distribution=self.distribution, name="May"
            ).id,
            "form-1-last_timestep": Timestep.objects.get(
                distribution=self.distribution, name="December"
            ).id,
        }
        response = self.client.post(self.get_create_url(), data)
        self.assertEqual(response.status_code, 302)
        frequency = CollectionFrequency.objects.get(name="Test Frequency with Seasons")
        seasons = [
            CollectionSeason.objects.get(
                first_timestep__name="January", last_timestep__name="April"
            ),
            CollectionSeason.objects.get(
                first_timestep__name="May", last_timestep__name="December"
            ),
        ]
        self.assertListEqual(
            seasons, list(frequency.seasons.order_by("first_timestep__order"))
        )

    def test_form_contains_all_initials(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_update_url(self.unpublished_object.pk))
        form = response.context["form"]
        expected = {
            "name": self.unpublished_object.name,
            "type": self.unpublished_object.type,
            "description": self.unpublished_object.description,
        }
        self.assertDictEqual(expected, form.initial)

    def test_formset_contains_all_initials(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_update_url(self.unpublished_object.pk))
        formset = response.context["formset"]
        expected = [
            {
                "distribution": self.distribution,
                "first_timestep": self.january,
                "last_timestep": self.may,
                "standard": 100,
                "option_1": 150,
                "option_2": None,
                "option_3": None,
            },
            {
                "distribution": self.distribution,
                "first_timestep": self.june,
                "last_timestep": self.december,
                "standard": 150,
                "option_1": None,
                "option_2": None,
                "option_3": None,
            },
        ]
        self.assertListEqual(expected, formset.initial)

    def test_post_http_302_options_are_changed_on_save(self):
        self.client.force_login(self.owner_user)
        url = self.get_update_url(self.unpublished_object.pk)
        data = self.related_objects_post_data()
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)
        self.options_1.refresh_from_db()
        self.options_2.refresh_from_db()
        self.assertEqual(150, self.options_1.standard)
        self.assertIsNone(self.options_1.option_1)
        self.assertIsNone(self.options_1.option_2)
        self.assertIsNone(self.options_1.option_3)
        self.assertEqual(200, self.options_2.standard)
        self.assertIsNone(self.options_2.option_1)
        self.assertIsNone(self.options_2.option_2)
        self.assertIsNone(self.options_2.option_3)


# ----------- CollectionPropertyValue CRUD -----------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectionPropertyValueCRUDViewsTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    public_list_view = False
    private_list_view = False

    model = CollectionPropertyValue

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "collectionpropertyvalue-create"
    view_detail_name = "collectionpropertyvalue-detail"
    view_update_name = "collectionpropertyvalue-update"
    view_delete_name = "collectionpropertyvalue-delete-modal"

    create_object_data = {"average": 123.5, "standard_deviation": 12.54, "year": 2025}
    update_object_data = {"average": 456.7, "standard_deviation": 23.45, "year": 2025}

    @classmethod
    def create_related_objects(cls):
        return {
            "collection": Collection.objects.create(
                owner=cls.owner_user,
                name="Test Collection",
                publication_status="published",
            ),
            "property": Property.objects.create(
                name="Test Property", publication_status="published"
            ),
            "unit": Unit.objects.create(
                name="Test Unit", publication_status="published"
            ),
        }

    def related_objects_post_data(self):
        data = super().related_objects_post_data()
        # Add formset management data for WasteFlyerFormSet
        data.update(
            {
                "form-TOTAL_FORMS": "0",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
            }
        )
        return data

    def get_delete_success_url(self, publication_status=None):
        return reverse(
            "collection-detail", kwargs={"pk": self.related_objects["collection"].pk}
        )

    def test_add_waste_flyer_via_update_view(self):
        """Test that WasteFlyer can be added through CPV update view formset."""
        self.client.force_login(self.owner_user)

        url = reverse(
            "collectionpropertyvalue-update", kwargs={"pk": self.unpublished_object.pk}
        )

        # Prepare POST data with WasteFlyer formset
        data = {
            "collection": self.related_objects["collection"].pk,
            "property": self.related_objects["property"].pk,
            "unit": self.related_objects["unit"].pk,
            "year": 2022,
            "average": 15,  # Update value to verify main form works
            "standard_deviation": "",
            # Formset data for WasteFlyer
            "form-TOTAL_FORMS": "1",
            "form-INITIAL_FORMS": "0",
            "form-MIN_NUM_FORMS": "0",
            "form-MAX_NUM_FORMS": "1000",
            "form-0-url": "http://example.com/flyer.pdf",
        }

        response = self.client.post(url, data)

        self.assertEqual(response.status_code, 302)

        self.unpublished_object.refresh_from_db()
        self.assertEqual(self.unpublished_object.average, 15)

        # Check if WasteFlyer was created
        flyers = WasteFlyer.objects.filter(url="http://example.com/flyer.pdf")
        self.assertTrue(flyers.exists(), "WasteFlyer should be created")

        flyer = flyers.first()

        # Check if WasteFlyer is linked to CPV via sources
        self.assertIn(
            flyer,
            self.unpublished_object.sources.all(),
            "WasteFlyer should be linked to CPV",
        )

    def test_update_preserves_bibliographic_sources_when_adding_waste_flyer(self):
        """Test that CPV update keeps bibliographic sources alongside WasteFlyers."""
        self.client.force_login(self.owner_user)
        bibliographic_source = Source.objects.create(
            owner=self.owner_user,
            type="article",
            title="Bibliographic Source",
            abbreviation="BibSource",
            publication_status="published",
        )

        url = reverse(
            "collectionpropertyvalue-update", kwargs={"pk": self.unpublished_object.pk}
        )
        data = {
            "collection": self.related_objects["collection"].pk,
            "property": self.related_objects["property"].pk,
            "unit": self.related_objects["unit"].pk,
            "year": 2022,
            "average": 42,
            "standard_deviation": "",
            "sources": [bibliographic_source.pk],
            "form-TOTAL_FORMS": "1",
            "form-INITIAL_FORMS": "0",
            "form-MIN_NUM_FORMS": "0",
            "form-MAX_NUM_FORMS": "1000",
            "form-0-url": "http://example.com/flyer-preserve.pdf",
        }

        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)

        self.unpublished_object.refresh_from_db()
        flyer = WasteFlyer.objects.get(url="http://example.com/flyer-preserve.pdf")
        self.assertSetEqual(
            set(self.unpublished_object.sources.values_list("pk", flat=True)),
            {bibliographic_source.pk, flyer.pk},
        )


# ----------- AggregatedCollectionPropertyValue CRUD -------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class AggregatedCollectionPropertyValueCRUDViewsTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    public_list_view = False
    private_list_view = False

    model = AggregatedCollectionPropertyValue

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "aggregatedcollectionpropertyvalue-create"
    view_detail_name = "aggregatedcollectionpropertyvalue-detail"
    view_update_name = "aggregatedcollectionpropertyvalue-update"
    view_delete_name = "aggregatedcollectionpropertyvalue-delete-modal"

    create_object_data = {"year": 2025, "average": 123.5, "standard_deviation": 12.54}
    update_object_data = {"year": 2025, "average": 456.7, "standard_deviation": 23.45}

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.related_collections = [
            Collection.objects.create(
                name="Test Collection 1", publication_status="published"
            ),
            Collection.objects.create(
                name="Test Collection 2", publication_status="published"
            ),
        ]
        cls.published_object.collections.set(cls.related_collections)
        cls.unpublished_object.collections.set(cls.related_collections)

    @classmethod
    def create_related_objects(cls):
        return {
            "property": Property.objects.create(
                name="Test Property", publication_status="published"
            ),
            "unit": Unit.objects.create(
                name="Test Unit", publication_status="published"
            ),
        }

    def related_objects_post_data(self):
        data = super().related_objects_post_data()
        data.update(
            {
                "collections": [
                    Collection.objects.create(
                        name="Test Collection 3", publication_status="published"
                    ).pk,
                    Collection.objects.create(
                        name="Test Collection 4", publication_status="published"
                    ).pk,
                ],
                # Add formset management data for WasteFlyerFormSet
                "form-TOTAL_FORMS": "0",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
            }
        )
        return data

    def get_delete_success_url(self, publication_status=None):
        related_ids = [collection.id for collection in self.related_collections]
        base_url = reverse("collection-list")
        query_string = urlencode([("id", rid) for rid in related_ids])
        return f"{base_url}?{query_string}"

    def test_update_preserves_bibliographic_sources_when_adding_waste_flyer(self):
        """Test that ACPV update keeps bibliographic sources alongside WasteFlyers."""
        self.client.force_login(self.owner_user)
        bibliographic_source = Source.objects.create(
            owner=self.owner_user,
            type="article",
            title="Aggregated Bibliographic Source",
            abbreviation="AggBibSource",
            publication_status="published",
        )

        url = reverse(
            "aggregatedcollectionpropertyvalue-update",
            kwargs={"pk": self.unpublished_object.pk},
        )
        data = {
            "collections": [collection.pk for collection in self.related_collections],
            "property": self.related_objects["property"].pk,
            "unit": self.related_objects["unit"].pk,
            "year": 2025,
            "average": 77,
            "standard_deviation": "",
            "sources": [bibliographic_source.pk],
            "form-TOTAL_FORMS": "1",
            "form-INITIAL_FORMS": "0",
            "form-MIN_NUM_FORMS": "0",
            "form-MAX_NUM_FORMS": "1000",
            "form-0-url": "http://example.com/aggregated-flyer-preserve.pdf",
        }

        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)

        self.unpublished_object.refresh_from_db()
        flyer = WasteFlyer.objects.get(
            url="http://example.com/aggregated-flyer-preserve.pdf"
        )
        self.assertSetEqual(
            set(self.unpublished_object.sources.values_list("pk", flat=True)),
            {bibliographic_source.pk, flyer.pk},
        )


# ----------- Collection Catchment CRUD --------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectionCatchmentCRUDViewsTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    model = CollectionCatchment
    model_add_permission = "add_collectioncatchment"

    view_dashboard_name = "wastecollection-explorer"
    view_published_list_name = "collectioncatchment-list"
    view_private_list_name = "collectioncatchment-list-owned"
    view_create_name = "collectioncatchment-create"
    view_detail_name = "collectioncatchment-detail"
    view_update_name = "collectioncatchment-update"
    view_delete_name = "collectioncatchment-delete-modal"

    allow_create_for_any_authenticated_user = True
    add_scope_query_param_to_list_urls = True

    create_object_data = {"name": "Test Catchment"}
    update_object_data = {"name": "Updated Test Catchment"}

    @classmethod
    def create_related_objects(cls):
        return {
            "region": Region.objects.create(
                name="Test Region", publication_status="published"
            )
        }

    # -----------------------
    # CreateView Test Cases
    # -----------------------

    def test_create_view_post_as_authenticated_without_permission(self):
        self.skipTest("Post method is not implemented for this view.")

    def test_create_view_post_as_authenticated_with_permission(self):
        self.skipTest("Post method is not implemented for this view.")

    def test_create_view_post_as_staff_user(self):
        self.skipTest("Post method is not implemented for this view.")


# ----------- Collection CRUD ------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectionCRUDViewsTestCase(AbstractTestCases.UserCreatedObjectCRUDViewTestCase):
    modal_detail_view = True

    model = Collection

    view_dashboard_name = "wastecollection-explorer"
    view_create_name = "collection-create"
    view_published_list_name = "collection-list"
    view_private_list_name = "collection-list-owned"
    view_detail_name = "collection-detail"
    view_modal_detail_name = "collection-detail-modal"
    view_update_name = "collection-update"
    view_delete_name = "collection-delete-modal"

    add_scope_query_param_to_list_urls = True

    create_object_data = {
        "name": "Test Collection",
        "description": "The original collection",
        "connection_type": "VOLUNTARY",
        "valid_from": date.today(),
        "valid_until": date.today() + timedelta(days=365),
    }
    update_object_data = {
        "name": "Updated Test Collection",
        "connection_type": "VOLUNTARY",
        "description": "This has been updated",
    }

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        with mute_signals(signals.post_save):
            cls.flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123", url="https://www.test-flyer.org"
            )
            cls.flyer2 = WasteFlyer.objects.create(
                abbreviation="WasteFlyer234", url="https://www.rest-flyer.org"
            )
        cls.unpublished_object.flyers.add(cls.flyer)
        cls.unpublished_object.flyers.add(cls.flyer2)
        cls.published_object.flyers.add(cls.flyer)
        cls.published_object.flyers.add(cls.flyer2)

    @classmethod
    def create_related_objects(cls):
        MaterialCategory.objects.create(
            name="Biowaste component", publication_status="published"
        )
        catchment = CollectionCatchment.objects.create(
            name="Test catchment", publication_status="published"
        )
        collector = Collector.objects.create(
            name="Test collector", publication_status="published"
        )
        collection_system = CollectionSystem.objects.create(
            name="Test system", publication_status="published"
        )
        waste_category = WasteCategory.objects.create(
            name="Test category", publication_status="published"
        )
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed material 1", publication_status="published"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed material 2", publication_status="published"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden material 1", publication_status="published"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden material 2", publication_status="published"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            publication_status="published",
            category=waste_category,
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        frequency = CollectionFrequency.objects.create(
            name="Test Frequency", publication_status="published"
        )
        Collection.objects.create(
            catchment=catchment,
            collector=collector,
            collection_system=collection_system,
            waste_stream=waste_stream,
            frequency=frequency,
            valid_from=date.today() - timedelta(days=365),
            valid_until=date.today() - timedelta(days=1),
            description="Predecessor Collection 1",
            publication_status="published",
        )
        Collection.objects.create(
            catchment=catchment,
            collector=collector,
            collection_system=collection_system,
            waste_stream=waste_stream,
            frequency=frequency,
            valid_from=date.today() - timedelta(days=365),
            valid_until=date.today() - timedelta(days=1),
            description="Predecessor Collection 2",
            publication_status="published",
        )
        # Create a bunch of unused outdated collections
        for i in range(12):
            Collection.objects.create(
                catchment=catchment,
                collector=collector,
                collection_system=collection_system,
                waste_stream=waste_stream,
                frequency=frequency,
                valid_from=date.today() - timedelta(days=365),
                valid_until=date.today() - timedelta(days=1),
                description=f"Oudated Collection {i}",
                publication_status="published",
            )
        return {
            "catchment": catchment,
            "collector": collector,
            "collection_system": collection_system,
            "waste_stream": waste_stream,
            "frequency": frequency,
        }

    @classmethod
    def create_published_object(cls):
        collection = super().create_published_object()
        collection.add_predecessor(
            Collection.objects.get(description="Predecessor Collection 1")
        )
        collection.add_predecessor(
            Collection.objects.get(description="Predecessor Collection 2")
        )
        return collection

    def related_objects_post_data(self):
        data = super().related_objects_post_data()
        data.update(
            {
                "waste_category": WasteCategory.objects.get(name="Test category").pk,
                "valid_from": date.today(),
                "form-TOTAL_FORMS": 0,
                "form-INITIAL_FORMS": 0,
            }
        )
        return data

    def get_current_list_url(self, publication_status=None):
        """
        Returns the URL for the current list view based on the valid_on filter.
        """
        return self.get_list_url(publication_status=publication_status)

    def test_post_get_formset_kwargs_fetches_correct_parent_object(self):
        request = RequestFactory().post(self.get_create_url())
        request.user = self.staff_user
        view = views.CollectionCreateView()
        view.setup(request)
        view.object = self.unpublished_object
        formset_kwargs = view.get_formset_kwargs()
        self.assertEqual(formset_kwargs["parent_object"], self.unpublished_object)

    def test_get_formset_has_correct_queryset(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(self.get_create_url())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["formset"].forms), 1)

    def test_post_with_missing_data_errors(self):
        self.client.force_login(self.staff_user)
        response = self.client.post(
            self.get_create_url(), data={"connection_rate_year": 123}
        )
        self.assertEqual(response.status_code, 200)

        error_msg = "This field is required."
        self.assertTrue(error_msg in response.context["form"].errors["catchment"])
        self.assertTrue(error_msg in response.context["form"].errors["collector"])
        self.assertTrue(
            error_msg in response.context["form"].errors["collection_system"]
        )
        self.assertTrue(error_msg in response.context["form"].errors["waste_category"])

    def test_post_with_valid_form_data(self):
        initial_count = Collection.objects.count()
        self.client.force_login(self.staff_user)
        response = self.client.post(
            self.get_create_url(),
            data={
                "catchment": CollectionCatchment.objects.first().id,
                "collector": Collector.objects.first().id,
                "collection_system": CollectionSystem.objects.first().id,
                "waste_category": WasteCategory.objects.first().id,
                "connection_type": "VOLUNTARY",
                "allowed_materials": [
                    self.allowed_material_1.id,
                    self.allowed_material_2.id,
                ],
                "forbidden_materials": [
                    self.forbidden_material_1.id,
                    self.forbidden_material_2.id,
                ],
                "frequency": self.related_objects["frequency"].id,
                "valid_from": date(2020, 1, 1),
                "description": "This is a test case that should pass!",
                "form-INITIAL_FORMS": "0",
                "form-TOTAL_FORMS": "2",
                "form-0-url": "https://www.test-flyer.org",
                "form-0-id": "",
                "form-1-url": "",
                "form-1-id": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Collection.objects.count(), initial_count + 1)

    def test_post_with_unspecified_allowed_materials_creates_generic_waste_stream(self):
        self.client.force_login(self.staff_user)
        data = {
            "catchment": CollectionCatchment.objects.first().id,
            "collector": Collector.objects.first().id,
            "collection_system": CollectionSystem.objects.first().id,
            "waste_category": WasteCategory.objects.first().id,
            "connection_type": "VOLUNTARY",
            "allowed_materials": [],
            "forbidden_materials": [
                self.forbidden_material_1.id,
                self.forbidden_material_2.id,
            ],
            "frequency": self.related_objects["frequency"].id,
            "valid_from": date(2020, 1, 1),
            "description": "This is a test case that should pass!",
            "form-INITIAL_FORMS": "0",
            "form-TOTAL_FORMS": "2",
            "form-0-url": "https://www.test-flyer.org",
            "form-0-id": "",
            "form-1-url": "",
            "form-1-id": "",
        }
        initial_count = Collection.objects.count()
        response = self.client.post(self.get_create_url(), data=data, follow=True)
        self.assertEqual(Collection.objects.count(), initial_count + 1)
        new_collection = Collection.objects.get(description=data["description"])
        self.assertRedirects(
            response, reverse("collection-detail", kwargs={"pk": new_collection.pk})
        )
        self.assertFalse(new_collection.waste_stream.allowed_materials.exists())

    def test_template_contains_predecessor_collections(self):
        response = self.client.get(self.get_detail_url(self.published_object.pk))
        self.assertContains(
            response,
            Collection.objects.get(description="Predecessor Collection 1").name,
        )
        self.assertContains(
            response,
            Collection.objects.get(description="Predecessor Collection 2").name,
        )

    def test_template_contains_flyer_url(self):
        response = self.client.get(self.get_detail_url(self.published_object.pk))
        self.assertContains(response, "https://www.test-flyer.org")

    def test_detail_context_includes_material_lists(self):
        response = self.client.get(self.get_detail_url(self.published_object.pk))

        self.assertIn("allowed_materials", response.context)
        self.assertIn("forbidden_materials", response.context)
        self.assertIn(self.allowed_material_1, response.context["allowed_materials"])
        self.assertIn(self.allowed_material_2, response.context["allowed_materials"])
        self.assertIn(
            self.forbidden_material_1, response.context["forbidden_materials"]
        )
        self.assertIn(
            self.forbidden_material_2, response.context["forbidden_materials"]
        )

    def test_detail_context_includes_samples_list(self):
        sample = Sample.objects.create(
            name="Detail Sample",
            material=self.allowed_material_1,
            publication_status="published",
        )
        self.published_object.samples.add(sample)

        response = self.client.get(self.get_detail_url(self.published_object.pk))

        self.assertIn("samples", response.context)
        self.assertIn(sample, response.context["samples"])
        self.assertContains(response, "Detail Sample")

    def test_uses_custom_template(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_update_url(self.unpublished_object.pk))
        self.assertTemplateUsed(response, "soilcom/collection_form.html")

    def test_get_get_formset_kwargs(self):
        request = RequestFactory().get(self.get_update_url(self.unpublished_object.pk))
        request.user = self.owner_user
        view = views.CollectionUpdateView()
        view.setup(request)
        view.kwargs = {"pk": self.unpublished_object.pk}
        view.object = self.unpublished_object
        view.relation_field_name = "flyers"
        expected_formset_kwargs = {
            "initial": [{"url": self.flyer.url}, {"url": self.flyer2.url}],
            "parent_object": self.unpublished_object,
            "owner": self.owner_user,
            "relation_field_name": view.relation_field_name,
        }
        self.assertDictEqual(expected_formset_kwargs, view.get_formset_kwargs())

    def test_post_get_formset_kwargs(self):
        kwargs = {"pk": self.unpublished_object.pk}
        data = {
            "form-INITIAL_FORMS": "0",
            "form-TOTAL_FORMS": "2",
            "form-0-url": "https://www.test-flyer.org",
            "form-1-url": "https://www.best-flyer.org",
        }
        request = RequestFactory().post(
            self.get_update_url(self.unpublished_object.pk), data=data
        )
        request.user = self.owner_user
        view = views.CollectionUpdateView()
        view.setup(request)
        view.kwargs = kwargs
        view.object = self.unpublished_object
        view.relation_field_name = "flyers"
        query_dict = QueryDict("", mutable=True)
        query_dict.update(data)
        expected_formset_kwargs = {
            "parent_object": self.unpublished_object,
            "owner": self.owner_user,
            "initial": [{"url": self.flyer.url}, {"url": self.flyer2.url}],
            "data": query_dict,
            "relation_field_name": view.relation_field_name,
        }
        self.assertDictEqual(expected_formset_kwargs, view.get_formset_kwargs())

    def test_get_get_formset(self):
        request = RequestFactory().get(self.get_update_url(self.unpublished_object.pk))
        request.user = self.owner_user
        view = views.CollectionUpdateView()
        view.setup(request)
        view.kwargs = {"pk": self.unpublished_object.pk}
        view.object = self.unpublished_object
        formset = view.get_formset()
        self.assertIsInstance(formset, WasteFlyerFormSet)
        self.assertEqual(2, formset.initial_form_count())

    def test_post_get_formset(self):
        data = {
            "form-INITIAL_FORMS": "0",
            "form-TOTAL_FORMS": "2",
            "form-0-url": "https://www.test-flyer.org",
            "form-1-url": "https://www.best-flyer.org",
        }
        request = RequestFactory().post(
            self.get_update_url(self.unpublished_object.pk), data=data
        )
        request.user = self.owner_user
        view = views.CollectionUpdateView()
        view.setup(request)
        view.kwargs = {"pk": self.unpublished_object.pk}
        view.object = self.unpublished_object
        formset = view.get_formset()
        self.assertIsInstance(formset, WasteFlyerFormSet)

    def test_context_contains_form_and_formset(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_update_url(self.unpublished_object.pk))
        self.assertIsInstance(response.context["form"], CollectionModelForm)
        self.assertIsInstance(response.context["formset"], BaseFormSet)

    def test_update_post_with_missing_data_errors(self):
        self.client.force_login(self.owner_user)
        response = self.client.post(
            self.get_update_url(self.unpublished_object.pk),
            data={"connection_rate_year": 123},
        )
        self.assertEqual(response.status_code, 200)

        error_msg = "This field is required."
        self.assertTrue(error_msg in response.context["form"].errors["catchment"])
        self.assertTrue(error_msg in response.context["form"].errors["collector"])
        self.assertTrue(
            error_msg in response.context["form"].errors["collection_system"]
        )
        self.assertTrue(error_msg in response.context["form"].errors["waste_category"])

    def test_update_post_with_valid_form_data(self):
        self.client.force_login(self.owner_user)
        response = self.client.post(
            self.get_update_url(self.unpublished_object.pk),
            data={
                "catchment": CollectionCatchment.objects.first().id,
                "collector": Collector.objects.first().id,
                "collection_system": CollectionSystem.objects.first().id,
                "waste_category": WasteCategory.objects.first().id,
                "connection_type": "VOLUNTARY",
                "allowed_materials": [
                    self.allowed_material_1.id,
                    self.allowed_material_2.id,
                ],
                "forbidden_materials": [
                    self.forbidden_material_1.id,
                    self.forbidden_material_2.id,
                ],
                "frequency": self.related_objects["frequency"].id,
                "valid_from": date(2020, 1, 1),
                "description": "This is a test case that should pass!",
                "form-INITIAL_FORMS": "0",
                "form-TOTAL_FORMS": "2",
                "form-0-url": "https://www.test-flyer.org",
            },
        )
        self.assertEqual(response.status_code, 302)

    def test_post_without_allowed_materials(self):
        self.client.force_login(self.owner_user)
        response = self.client.post(
            self.get_update_url(self.unpublished_object.pk),
            data={
                "catchment": CollectionCatchment.objects.first().id,
                "collector": Collector.objects.first().id,
                "collection_system": CollectionSystem.objects.first().id,
                "waste_category": WasteCategory.objects.first().id,
                "connection_type": "VOLUNTARY",
                "allowed_materials": [
                    self.allowed_material_1.id,
                    self.allowed_material_2.id,
                ],
                "frequency": self.related_objects["frequency"].id,
                "valid_from": date(2020, 1, 1),
                "description": "This is a test case that should pass!",
                "form-INITIAL_FORMS": "2",
                "form-TOTAL_FORMS": "2",
                "form-0-url": "https://www.test-flyer.org",
            },
        )
        self.assertEqual(response.status_code, 302)

    def test_associated_flyers_are_displayed_as_initial_values_in_formset(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_update_url(self.unpublished_object.pk))
        expected_initial = [flyer.url for flyer in self.unpublished_object.flyers.all()]
        real_initial = [
            form.initial["url"] for form in response.context["formset"].initial_forms
        ]
        self.assertListEqual(expected_initial, real_initial)

    def test_new_flyers_are_created_from_unknown_urls(self):
        self.client.force_login(self.owner_user)
        data = {
            "catchment": self.unpublished_object.catchment.id,
            "collector": self.unpublished_object.collector.id,
            "collection_system": self.unpublished_object.collection_system.id,
            "waste_category": self.unpublished_object.waste_stream.category.id,
            "connection_type": "VOLUNTARY",
            "allowed_materials": [
                m.id
                for m in self.unpublished_object.waste_stream.allowed_materials.all()
            ],
            "forbidden_materials": [
                m.id
                for m in self.unpublished_object.waste_stream.forbidden_materials.all()
            ],
            "frequency": self.related_objects["frequency"].id,
            "valid_from": date(2020, 1, 1),
            "description": self.unpublished_object.description,
            "form-INITIAL_FORMS": "1",
            "form-TOTAL_FORMS": "2",
            "form-0-url": "https://www.best-flyer.org",
            "form-1-url": "https://www.fest-flyer.org",
        }
        with mute_signals(signals.post_save):
            self.client.post(self.get_update_url(self.unpublished_object.pk), data=data)
        flyer = WasteFlyer.objects.get(url="https://www.fest-flyer.org")
        self.assertIsInstance(flyer, WasteFlyer)

    def test_regression_post_with_valid_data_doesnt_delete_unchanged_flyers(self):
        self.client.force_login(self.owner_user)
        data = {
            "catchment": self.unpublished_object.catchment.id,
            "collector": self.unpublished_object.collector.id,
            "collection_system": self.unpublished_object.collection_system.id,
            "waste_category": self.unpublished_object.waste_stream.category.id,
            "connection_type": "VOLUNTARY",
            "allowed_materials": [
                m.id
                for m in self.unpublished_object.waste_stream.allowed_materials.all()
            ],
            "forbidden_materials": [
                m.id
                for m in self.unpublished_object.waste_stream.forbidden_materials.all()
            ],
            "frequency": self.unpublished_object.frequency.id,
            "valid_from": self.unpublished_object.valid_from,
            "description": self.unpublished_object.description,
            "form-INITIAL_FORMS": "1",
            "form-TOTAL_FORMS": "2",
            "form-0-url": self.unpublished_object.flyers.first().url,
            "form-0-id": self.unpublished_object.flyers.first().id,
            "form-1-url": "https://www.fest-flyer.org",
            "form-1-id": "",
        }
        # The flyer with the url rest-flyer.org should be removed from the unpublished collection  by this. Remove it from the
        # published collection as well before the request so that the process for deleting orphaned flyers is evoked.
        rest_flyer = WasteFlyer.objects.get(url="https://www.rest-flyer.org")
        self.published_object.flyers.remove(rest_flyer)
        with (
            mute_signals(post_save),
            patch(
                "case_studies.soilcom.forms.cleanup_orphaned_waste_flyers.delay"
            ) as mock_cleanup,
        ):
            response = self.client.post(
                self.get_update_url(self.unpublished_object.pk), data=data
            )
        mock_cleanup.assert_called_once()
        cleanup_orphaned_waste_flyers()
        self.assertEqual(response.status_code, 302)
        self.assertIn(
            WasteFlyer.objects.get(url="https://www.fest-flyer.org"),
            self.unpublished_object.flyers.all(),
        )
        self.assertIn(
            WasteFlyer.objects.get(url="https://www.test-flyer.org"),
            self.unpublished_object.flyers.all(),
        )
        self.assertEqual(WasteFlyer.objects.count(), 2)


class CollectionCopyViewTestCase(ViewWithPermissionsTestCase):
    member_permissions = "add_collection"
    url_name = "collection-copy"

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        cls.material_category = MaterialCategory.objects.create(
            name="Biowaste component", publication_status="published"
        )
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed material 1", publication_status="published"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed material 2", publication_status="published"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden material 1", publication_status="published"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden material 2", publication_status="published"
        )
        cls.waste_category = WasteCategory.objects.create(
            name="Test category", publication_status="published"
        )
        cls.waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=cls.waste_category,
        )
        cls.waste_stream.allowed_materials.add(cls.allowed_material_1)
        cls.waste_stream.allowed_materials.add(cls.allowed_material_2)
        cls.waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        cls.waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            cls.flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123",
                url="https://www.test-flyer.org",
                publication_status="published",
            )
            cls.flyer2 = WasteFlyer.objects.create(
                abbreviation="WasteFlyer234",
                url="https://www.fest-flyer.org",
                publication_status="published",
            )
        cls.frequency = CollectionFrequency.objects.create(
            name="Test Frequency", publication_status="published"
        )
        cls.collection_catchment = CollectionCatchment.objects.create(
            name="Test catchment", publication_status="published"
        )
        cls.collector = Collector.objects.create(
            name="Test collector", publication_status="published"
        )
        cls.collection_system = CollectionSystem.objects.create(
            name="Test system", publication_status="published"
        )
        cls.collection = Collection.objects.create(
            name="collection1",
            catchment=cls.collection_catchment,
            collector=cls.collector,
            collection_system=cls.collection_system,
            waste_stream=cls.waste_stream,
            frequency=cls.frequency,
            description="This is a test case.",
            publication_status="published",
        )
        cls.collection.flyers.add(cls.flyer)
        cls.collection.flyers.add(cls.flyer2)

    def test_get_http_302_redirect_for_anonymous(self):
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 302)

    def test_get_http_403_forbidden_for_non_group_members(self):
        self.client.force_login(self.outsider)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 403)

    def test_get_http_200_ok_for_group_members(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 200)

    def test_form_contains_exactly_one_submit_button(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertContains(response, 'type="submit"', count=1, status_code=200)

    def test_get_object(self):
        request = RequestFactory().get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        request.user = self.member
        view = views.CollectionCopyView()
        view.setup(request)
        view.kwargs = {"pk": self.collection.id}
        self.assertEqual(view.get_object(), self.collection)

    def test_get_get_formset_kwargs_fetches_initial_and_parent_object(self):
        request = RequestFactory().get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        view = views.CollectionCopyView()
        view.setup(request)
        view.kwargs = {"pk": self.collection.id}
        view.object = view.get_object()
        view.relation_field_name = "flyers"
        expected = {
            "initial": [{"url": self.flyer.url}, {"url": self.flyer2.url}],
            "parent_object": self.collection,
            "relation_field_name": view.relation_field_name,
        }
        self.assertDictEqual(expected, view.get_formset_kwargs())

    def test_get_get_formset_initial_fetches_urls_of_related_flyers(self):
        request = RequestFactory().get(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        request.user = self.member
        view = views.CollectionCopyView()
        view.setup(request)
        view.kwargs = {"pk": self.collection.pk}
        view.object = view.get_object()
        expected = [
            {"url": "https://www.test-flyer.org"},
            {"url": "https://www.fest-flyer.org"},
        ]
        self.assertListEqual(expected, view.get_formset_initial())

    def test_get_formset_has_correct_queryset(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            len(response.context["formset"].forms),
            len(self.collection.flyers.all()) + 1,
        )

    def test_post_http_302_redirect_for_anonymous(self):
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 302)

    def test_post_http_403_forbidden_for_non_group_members(self):
        self.client.force_login(self.outsider)
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 403)

    def test_post_http_302_redirect_for_member(self):
        self.client.force_login(self.member)
        data = {
            "catchment": self.collection_catchment.id,
            "collector": Collector.objects.create(
                name="New Test Collector", publication_status="published"
            ).id,
            "collection_system": self.collection_system.id,
            "waste_category": self.waste_stream.category.id,
            "connection_type": "VOLUNTARY",
            "allowed_materials": [
                self.allowed_material_1.id,
                self.allowed_material_2.id,
            ],
            "forbidden_materials": [
                self.forbidden_material_1.id,
                self.forbidden_material_2.id,
            ],
            "frequency": self.frequency.id,
            "valid_from": date(2022, 1, 1),
            "description": "This is a test case that should pass!",
            "form-INITIAL_FORMS": "0",
            "form-TOTAL_FORMS": "0",
        }
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.id}), data=data
        )
        self.assertEqual(response.status_code, 302)

    def test_post_creates_new_copy(self):
        self.client.force_login(self.member)
        self.assertEqual(Collection.objects.count(), 1)
        get_response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        initial = get_response.context["form"].initial
        data = {
            "catchment": initial["catchment"],
            "collector": initial["collector"],
            "collection_system": initial["collection_system"],
            "waste_category": initial["waste_category"],
            "connection_type": "VOLUNTARY",
            "allowed_materials": initial["allowed_materials"],
            "forbidden_materials": initial["forbidden_materials"],
            "frequency": initial["frequency"],
            "valid_from": initial["valid_from"],
            "description": initial["description"],
            "form-INITIAL_FORMS": "0",
            "form-TOTAL_FORMS": "0",
        }
        post_response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.pk}), data=data
        )
        self.assertEqual(post_response.status_code, 302)
        self.assertEqual(Collection.objects.count(), 2)

    def test_post_copy_is_still_associated_with_unchanged_original_flyers(self):
        self.client.force_login(self.member)
        self.assertEqual(Collection.objects.count(), 1)
        get_response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        initial = get_response.context["form"].initial
        data = {
            "catchment": initial["catchment"],
            "collector": initial["collector"],
            "collection_system": initial["collection_system"],
            "waste_category": initial["waste_category"],
            "connection_type": "VOLUNTARY",
            "allowed_materials": initial["allowed_materials"],
            "forbidden_materials": initial["forbidden_materials"],
            "frequency": initial["frequency"],
            "valid_from": initial["valid_from"],
            "valid_until": "",
            "description": "This is the copy.",
            "form-INITIAL_FORMS": "1",
            "form-TOTAL_FORMS": "1",
            "form-0-url": self.flyer.url,
            "form-0-id": self.flyer.id,
        }
        self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.pk}), data=data
        )
        copy = Collection.objects.get(description="This is the copy.")
        self.assertEqual(copy.flyers.count(), 1)
        flyer = copy.flyers.first()
        self.assertEqual(flyer.url, self.flyer.url)


class CollectionCreateNewVersionViewTestCase(ViewWithPermissionsTestCase):
    member_permissions = "add_collection"
    url_name = "collection-new-version"

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        cls.material_category = MaterialCategory.objects.create(
            name="Biowaste component", publication_status="published"
        )
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed material 1", publication_status="published"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed material 2", publication_status="published"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden material 1", publication_status="published"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden material 2", publication_status="published"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=WasteCategory.objects.create(
                name="Test category", publication_status="published"
            ),
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            cls.flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123",
                url="https://www.test-flyer.org",
                publication_status="published",
            )
            cls.flyer2 = WasteFlyer.objects.create(
                abbreviation="WasteFlyer234",
                url="https://www.fest-flyer.org",
                publication_status="published",
            )
        cls.frequency = CollectionFrequency.objects.create(
            name="Test Frequency", publication_status="published"
        )
        cls.catchment = CollectionCatchment.objects.create(
            name="Test catchment", publication_status="published"
        )
        cls.collector = Collector.objects.create(
            name="Test collector", publication_status="published"
        )
        cls.collection_system = CollectionSystem.objects.create(
            name="Test system", publication_status="published"
        )
        cls.waste_stream = WasteStream.objects.create(
            category=WasteCategory.objects.create(
                name="Test category", publication_status="published"
            ),
            publication_status="published",
        )
        cls.collection = Collection.objects.create(
            name="collection1",
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.collection_system,
            waste_stream=waste_stream,
            frequency=cls.frequency,
            description="This is a test case.",
            publication_status="published",
        )
        cls.collection.flyers.add(cls.flyer)
        cls.collection.flyers.add(cls.flyer2)

    def test_get_http_302_redirect_for_anonymous(self):
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 302)

    def test_get_http_403_forbidden_for_non_group_members(self):
        self.client.force_login(self.outsider)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 403)

    def test_get_http_200_ok_for_group_members(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 200)

    def test_post_http_302_redirect_for_anonymous(self):
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 302)

    def test_post_http_403_forbidden_for_non_group_members(self):
        self.client.force_login(self.outsider)
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        self.assertEqual(response.status_code, 403)

    def test_post_http_302_redirect_for_member(self):
        self.client.force_login(self.member)
        data = {
            "catchment": self.catchment.id,
            "collector": self.collector.id,
            "collection_system": self.collection_system.id,
            "waste_category": self.waste_stream.category.id,
            "connection_type": "VOLUNTARY",
            "allowed_materials": [
                self.allowed_material_1.id,
                self.allowed_material_2.id,
            ],
            "forbidden_materials": [
                self.forbidden_material_1.id,
                self.forbidden_material_2.id,
            ],
            "frequency": self.frequency.id,
            "valid_from": date(2022, 1, 1),
            "description": "This is a test case that should pass!",
            "form-INITIAL_FORMS": "0",
            "form-TOTAL_FORMS": "0",
        }
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.id}), data=data
        )
        self.assertEqual(response.status_code, 302)


# ----------- Collection utils -----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectionAutocompleteViewTestCase(ViewWithPermissionsTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        Collection.objects.create(
            catchment=CollectionCatchment.objects.create(name="Hamburg"),
            publication_status="published",
        )
        Collection.objects.create(
            catchment=CollectionCatchment.objects.create(name="Berlin"),
            publication_status="published",
        )

    def test_get_http_200_ok_for_anonymous(self):
        response = self.client.get(reverse("collection-autocomplete"))
        self.assertEqual(response.status_code, 200)

    def test_get_returns_json_response(self):
        response = self.client.get(reverse("collection-autocomplete"))
        self.assertIsInstance(response, JsonResponse)

    def test_get_returns_only_collections_with_names_containing_filter_string(self):
        response = self.client.get(reverse("collection-autocomplete") + "?q=Ham")
        self.assertContains(response, "Hamburg")
        self.assertNotContains(response, "Berlin")


class CollectionAddPropertyValueViewTestCase(ViewWithPermissionsTestCase):
    member_permissions = "add_collectionpropertyvalue"
    url_name = "collection-add-property"

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.collection = Collection.objects.create(
            name="Test Collection", publication_status="published", owner=cls.member
        )
        cls.unit = Unit.objects.create(name="Test Unit", publication_status="published")
        cls.prop = Property.objects.create(
            name="Test Property", publication_status="published"
        )
        cls.prop.allowed_units.add(cls.unit)

    def test_get_http_302_redirect_for_anonymous(self):
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        self.assertEqual(response.status_code, 302)

    def test_get_http_403_forbidden_for_outsiders(self):
        self.client.force_login(self.outsider)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        self.assertEqual(response.status_code, 403)

    def test_get_http_200_ok_for_members(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        self.assertEqual(response.status_code, 200)

    def test_form_contains_exactly_one_submit_button(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        self.assertContains(response, 'type="submit"', count=1, status_code=200)

    def test_get_initial_has_collection(self):
        request = RequestFactory().get(
            reverse(self.url_name, kwargs={"pk": self.collection.id})
        )
        request.user = self.member
        view = views.CollectionAddPropertyValueView()
        view.setup(request)
        view.kwargs = {"pk": self.collection.id}
        view.dispatch(request, pk=self.collection.id)
        initial = view.get_initial()
        anchor = self.collection.version_anchor or self.collection
        expected = {
            "collection": anchor.pk,
        }
        self.assertDictEqual(expected, initial)

    def test_post_http_302_redirect_for_anonymous(self):
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        self.assertEqual(response.status_code, 302)

    def test_post_http_403_forbidden_for_outsiders(self):
        self.client.force_login(self.outsider)
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.pk})
        )
        self.assertEqual(response.status_code, 403)

    def test_post_http_302_redirect_for_members_with_minimal_data(self):
        self.client.force_login(self.member)
        data = {
            "collection": self.collection.pk,
            "property": self.prop.pk,
            "unit": self.unit.pk,
            "year": 2022,
            "average": 123.5,
            "standard_deviation": 12.6,
            # Add formset management data for WasteFlyerFormSet
            "form-TOTAL_FORMS": "0",
            "form-INITIAL_FORMS": "0",
            "form-MIN_NUM_FORMS": "0",
            "form-MAX_NUM_FORMS": "1000",
        }
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.collection.pk}), data=data
        )
        self.assertEqual(response.status_code, 302)


class CollectionAddAggregatedPropertyValueViewTestCase(ViewWithPermissionsTestCase):
    member_permissions = "add_aggregatedcollectionpropertyvalue"
    url_name = "collectioncatchment-add-aggregatedpropertyvalue"

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.catchment = CollectionCatchment.objects.create(
            publication_status="published"
        )
        Collection.objects.create(
            catchment=CollectionCatchment.objects.create(
                parent=cls.catchment, publication_status="published"
            ),
            publication_status="published",
        )
        Collection.objects.create(
            catchment=CollectionCatchment.objects.create(
                parent=cls.catchment, publication_status="published"
            ),
            publication_status="published",
        )
        cls.unit = Unit.objects.create(name="Test Unit", publication_status="published")
        cls.prop = Property.objects.create(
            name="Test Property", publication_status="published"
        )
        cls.prop.allowed_units.add(cls.unit)

    def test_get_http_302_redirect_for_anonymous(self):
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.catchment.pk})
        )
        self.assertEqual(response.status_code, 302)

    def test_get_http_403_forbidden_for_outsiders(self):
        self.client.force_login(self.outsider)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.catchment.pk})
        )
        self.assertEqual(response.status_code, 403)

    def test_get_http_200_ok_for_members(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.catchment.pk})
        )
        self.assertEqual(response.status_code, 200)

    def test_form_contains_exactly_one_submit_button(self):
        self.client.force_login(self.member)
        response = self.client.get(
            reverse(self.url_name, kwargs={"pk": self.catchment.pk})
        )
        self.assertContains(response, 'type="submit"', count=1, status_code=200)

    def test_get_initial_has_collections(self):
        request = RequestFactory().get(
            reverse(self.url_name, kwargs={"pk": self.catchment.id})
        )
        view = views.CollectionCatchmentAddAggregatedPropertyView()
        view.setup(request)
        view.kwargs = {"pk": self.catchment.id}
        initial = view.get_initial()
        expected = {
            "collections": self.catchment.downstream_collections,
        }
        self.assertIn("collections", initial)
        self.assertQuerySetEqual(
            expected["collections"].order_by("id"),
            initial["collections"].order_by("id"),
        )

    def test_post_http_302_redirect_for_anonymous(self):
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.catchment.pk})
        )
        self.assertEqual(response.status_code, 302)

    def test_post_http_403_forbidden_for_outsiders(self):
        self.client.force_login(self.outsider)
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.catchment.pk})
        )
        self.assertEqual(response.status_code, 403)

    def test_post_http_302_redirect_for_members_with_minimal_data(self):
        self.client.force_login(self.member)
        data = {
            "collections": [
                collection.pk for collection in self.catchment.downstream_collections
            ],
            "property": self.prop.pk,
            "unit": self.unit.pk,
            "year": 2022,
            "average": 123.5,
            "standard_deviation": 12.6,
        }
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.catchment.pk}), data=data
        )
        self.assertEqual(response.status_code, 302)


class CollectionWasteSamplesViewTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    create_view = False
    public_list_view = False
    private_list_view = False
    detail_view = False
    delete_view = False

    model = Collection

    view_update_name = "collection-wastesamples"
    update_success_url_name = "collection-wastesamples"

    create_object_data = {"name": "Test Collection"}

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        material = Material.objects.create(
            name="Test Material", publication_status="published"
        )
        series = SampleSeries.objects.create(
            name="Test Series", material=material, publication_status="published"
        )
        cls.unpublished_object.samples.add(
            Sample.objects.create(
                name="Test Sample 1", material=material, series=series
            )
        )
        cls.sample = Sample.objects.create(
            name="Test Sample 2",
            material=material,
            series=series,
            publication_status="published",
        )

    def compile_update_post_data(self):
        data = {"sample": self.sample.pk, "submit": "Add"}
        return data

    def test_form_contains_one_submit_button_for_each_form(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_update_url(self.unpublished_object.pk))
        self.assertContains(response, 'type="submit"', count=2, status_code=200)
        self.assertContains(response, 'value="Add"')
        self.assertContains(response, 'value="Remove"')

    def test_unpublished_post_success_and_http_302_redirect_on_submit_of_add_form(self):
        self.client.force_login(self.owner_user)
        sample = Sample.objects.get(name="Test Sample 2")
        data = {"sample": sample.pk, "submit": "Add"}
        response = self.client.post(
            self.get_update_url(self.unpublished_object.pk), data, follow=True
        )
        self.assertRedirects(response, self.get_update_url(self.unpublished_object.pk))

    def test_unpublished_post_success_and_http_302_redirect_on_submit_of_remove_form(
        self,
    ):
        self.client.force_login(self.owner_user)
        sample = Sample.objects.get(name="Test Sample 1")
        data = {"sample": sample.pk, "submit": "Remove"}
        response = self.client.post(
            self.get_update_url(self.unpublished_object.pk), data, follow=True
        )
        self.assertRedirects(response, self.get_update_url(self.unpublished_object.pk))


class CollectionPredecessorsViewTestCase(
    AbstractTestCases.UserCreatedObjectCRUDViewTestCase
):
    create_view = False
    public_list_view = False
    private_list_view = False
    detail_view = False
    delete_view = False

    model = Collection

    view_update_name = "collection-predecessors"

    update_success_url_name = "collection-predecessors"

    create_object_data = {"name": "Test Collection"}
    update_object_data = {"name": "Test Collection"}

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.predecessor_collection = Collection.objects.create(
            name="Test Predecessor",
            catchment=cls.related_objects["catchment"],
            collector=cls.related_objects["collector"],
            collection_system=cls.related_objects["collection_system"],
            waste_stream=cls.related_objects["waste_stream"],
            publication_status="published",
        )

    @classmethod
    def create_related_objects(cls):
        return {
            "catchment": CollectionCatchment.objects.create(
                name="Test Catchment", publication_status="published"
            ),
            "collector": Collector.objects.create(
                name="Test Collector", publication_status="published"
            ),
            "collection_system": CollectionSystem.objects.create(
                name="Test System", publication_status="published"
            ),
            "waste_stream": WasteStream.objects.create(
                name="Test Waste Stream",
                category=WasteCategory.objects.create(
                    name="Test Category", publication_status="published"
                ),
            ),
        }

    def compile_update_post_data(self):
        # This is just for the standard access tests. Both forms function is tested separately in specific functions.
        return {"predecessor": self.predecessor_collection.pk, "submit": "Add"}

    def test_form_contains_one_submit_button_for_each_form(self):
        self.client.force_login(self.owner_user)
        response = self.client.get(self.get_update_url(self.unpublished_object.pk))
        self.assertContains(response, 'type="submit"', count=2, status_code=200)
        self.assertContains(response, 'value="Add"')
        self.assertContains(response, 'value="Remove"')

    def test_post_success_and_http_302_redirect_on_submit_of_add_form(self):
        self.client.force_login(self.owner_user)
        data = {"predecessor": self.predecessor_collection.pk, "submit": "Add"}
        response = self.client.post(
            self.get_update_url(self.unpublished_object.pk), data=data, follow=True
        )
        self.assertRedirects(
            response, self.get_update_success_url(pk=self.unpublished_object.pk)
        )

    def test_post_success_and_http_302_redirect_on_submit_of_remove_form(self):
        self.client.force_login(self.owner_user)
        self.unpublished_object.add_predecessor(self.predecessor_collection)
        data = {"predecessor": self.predecessor_collection.pk, "submit": "Remove"}
        response = self.client.post(
            self.get_update_url(self.unpublished_object.pk), data=data, follow=True
        )
        self.assertRedirects(
            response, self.get_update_success_url(pk=self.unpublished_object.pk)
        )


class WasteCollectionPublishedMapViewTestCase(ViewWithPermissionsTestCase):
    member_permissions = (
        "add_collection",
        "view_collection",
        "change_collection",
        "delete_collection",
    )
    url = reverse("WasteCollection")

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        style = MapLayerStyle.objects.create(name="default")
        layer = MapLayerConfiguration.objects.create(
            name="default", layer_type="features", style=style
        )
        map_config = MapConfiguration.objects.create(name="default")
        map_config.layers.add(layer)
        region = Region.objects.create(name="Test Region")
        cls.dataset, _ = GeoDataset.objects.get_or_create(
            model_name="WasteCollection",
            defaults={
                "name": "Waste Collections Europe",
                "description": "Waste Collection Systems of Europe",
                "region": region,
            },
        )
        cls.dataset.map_configuration = map_config
        cls.dataset.save()
        catchment = CollectionCatchment.objects.create(
            name="Test Catchment", region=region
        )
        cls.collection = Collection.objects.create(
            name="Test Collection",
            catchment=catchment,
            publication_status=Collection.STATUS_PUBLISHED,
        )

    def test_http_200_ok_for_anonymous(self):
        response = self.client.get(self.url, follow=True)
        self.assertEqual(response.status_code, 200)

    def test_http_200_ok_for_member(self):
        self.client.force_login(self.member)
        response = self.client.get(self.url, follow=True)
        self.assertEqual(response.status_code, 200)

    def test_uses_correct_template(self):
        self.client.force_login(self.member)
        response = self.client.get(self.url, follow=True)
        self.assertTemplateUsed(response, "waste_collection_map.html")

    def test_create_collection_option_visible_for_member(self):
        self.client.force_login(self.member)
        response = self.client.get(self.url, follow=True)
        self.assertContains(response, Collection.create_url())

    def test_create_collection_option_not_available_for_outsider(self):
        self.client.force_login(self.outsider)
        response = self.client.get(self.url, follow=True)
        self.assertNotContains(response, Collection.create_url())

    def test_copy_collection_option_visible_for_member(self):
        self.client.force_login(self.member)
        response = self.client.get(self.url, follow=True)
        self.assertContains(response, "Copy selected collection")

    def test_copy_collection_option_not_available_for_outsider(self):
        self.client.force_login(self.outsider)
        response = self.client.get(self.url, follow=True)
        self.assertNotContains(response, "Copy selected collection")

    def test_update_collection_option_visible_for_staff(self):
        self.client.force_login(self.staff)
        response = self.client.get(self.url, follow=True)
        self.assertContains(response, "Edit selected collection")

    def test_update_collection_option_not_available_for_outsider(self):
        self.client.force_login(self.outsider)
        response = self.client.get(self.url, follow=True)
        # Button is present for authenticated users but is disabled for outsiders
        self.assertContains(response, 'id="btn-collection-update"')
        self.assertContains(response, "Edit selected collection")
        self.assertContains(
            response, 'class="btn btn-outline-secondary disabled w-100 mt-2"'
        )
        self.assertContains(response, 'aria-disabled="true"')
        self.assertContains(response, 'href="#"')

    def test_collection_dashboard_option_visible_for_member(self):
        self.client.force_login(self.member)
        response = self.client.get(self.url, follow=True)
        self.assertContains(response, reverse("wastecollection-explorer"))

    def test_collection_dashboard_option_not_available_for_outsider(self):
        self.client.force_login(self.outsider)
        response = self.client.get(self.url, follow=True)
        self.assertContains(response, reverse("wastecollection-explorer"))

    def test_range_slider_static_files_are_embedded(self):
        self.client.force_login(self.member)
        response = self.client.get(self.url, follow=True)
        self.assertContains(response, "range_slider.min.js")


class CollectionReviewProcessWithPredecessorsTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        # Ensure a MaterialCategory (materials.models) for "Biowaste component" exists
        # This might be used by signals or other underlying logic.
        MaterialCategory.objects.get_or_create(
            name="Biowaste component", publication_status="published"
        )

        # Create a test user with necessary permissions
        cls.user = User.objects.create_user(
            username="testuser", is_staff=True, is_superuser=True
        )

        # Create common related objects (Collector, System, Catchment)
        # Assuming these models have an 'owner' field like others.
        cls.collector = Collector.objects.create(name="Test Collector", owner=cls.user)
        cls.collection_system = CollectionSystem.objects.create(
            name="Test System", owner=cls.user
        )
        cls.catchment = CollectionCatchment.objects.create(
            name="Test Catchment", owner=cls.user
        )

        # Create WasteCategory (soilcom.models)
        cls.waste_category = WasteCategory.objects.create(
            name="Test Waste Category", owner=cls.user
        )

        # Create MaterialCategory (soilcom.models) for "Biowaste component"
        # This is distinct from the MaterialCategory above.
        cls.soilcom_biowaste_material_category, _ = (
            MaterialCategory.objects.get_or_create(
                name="Biowaste component", defaults={"owner": cls.user}
            )
        )

        # Create WasteComponent and associate with the soilcom.MaterialCategory
        cls.waste_component = WasteComponent.objects.create(
            name="Test Component", owner=cls.user, type="material"
        )
        cls.waste_component.categories.add(cls.soilcom_biowaste_material_category)

        # Create WasteStream
        cls.waste_stream = WasteStream.objects.create(
            name="Test Stream", category=cls.waste_category, owner=cls.user
        )

        cls.waste_stream.allowed_materials.add(cls.waste_component)

        # Create MaterialCategory and Material (materials.models)
        cls.material_category = MaterialCategory.objects.create(
            name="Test Material Category", owner=cls.user
        )
        cls.material = Material.objects.create(
            name="Test Material",
            owner=cls.user,
        )
        cls.material.categories.add(cls.material_category)

        # Add the Material to the WasteStream's allowed materials
        cls.waste_stream.allowed_materials.add(cls.material)
        cls.waste_stream.save()  # Save WasteStream after M2M modifications

        # Create a base published collection for use in tests
        # Assumes Collection.catchment is a ManyToManyField to CollectionCatchment
        cls.published_collection = Collection.objects.create(
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.collection_system,
            waste_stream=cls.waste_stream,
            valid_from=date.today() - timedelta(days=30),
            valid_until=date.today() + timedelta(days=30),
            publication_status=Collection.STATUS_PUBLISHED,
            owner=cls.user,
            connection_type="VOLUNTARY",
        )

    def setUp(self):
        """
        Prepare environment for each test method.
        Re-fetch mutable objects from DB to ensure test isolation with TransactionTestCase.
        Log in the test user.
        """
        # Re-fetch user to ensure a fresh object (though often not strictly needed for user itself if not modified)
        self.client.force_login(self.user)

        # Re-fetch objects that might be modified by tests or require a pristine state
        self.collector = Collector.objects.get(pk=self.collector.pk)
        self.collection_system = CollectionSystem.objects.get(
            pk=self.collection_system.pk
        )
        self.catchment = CollectionCatchment.objects.get(pk=self.catchment.pk)
        self.waste_category = WasteCategory.objects.get(pk=self.waste_category.pk)
        self.waste_stream = WasteStream.objects.get(pk=self.waste_stream.pk)

        # Crucially, refresh the published_collection as its attributes (e.g., valid_until, status) can change
        self.published_collection = Collection.objects.get(
            pk=self.published_collection.pk
        )

    def test_predecessor_visibility_during_versioning(self):
        """Test that the predecessor remains visible until the new version is published."""
        # 1. Create a new version of the collection via POST request
        # The form data should accurately reflect the fields expected by the view/form.

        form_data = {
            "catchment": self.catchment.pk,
            "collector": self.collector.pk,
            "collection_system": self.collection_system.pk,
            "waste_category": self.waste_category.pk,
            "connection_type": "VOLUNTARY",
            "waste_stream": self.waste_stream.pk,
            "valid_from": date.today(),
            "form-TOTAL_FORMS": 1,
            "form-INITIAL_FORMS": 0,
            "form-0-url": "https://www.test-flyer.org",
            "form-0-id": "",
        }
        response = self.client.post(
            reverse(
                "collection-new-version", kwargs={"pk": self.published_collection.pk}
            ),
            form_data,
            follow=True,
        )

        self.assertEqual(response.status_code, 200, "Failed to create new version")

        new_version = Collection.objects.get(
            predecessors=self.published_collection,
            valid_from=form_data["valid_from"],
            owner=self.user,
        )

        self.assertIsNotNone(
            new_version, "New version was not created or could not be found."
        )

        # 2. Verify new version is private, predecessor is still published
        self.assertEqual(new_version.publication_status, Collection.STATUS_PRIVATE)
        self.published_collection.refresh_from_db()
        self.assertEqual(
            self.published_collection.publication_status, Collection.STATUS_PUBLISHED
        )

        # 3. Verify predecessor is visible in list view
        response = self.client.get(reverse("collection-list"), follow=True)
        self.assertEqual(response.status_code, 200)
        collections = response.context["object_list"]
        self.assertIn(self.published_collection, collections)
        self.assertNotIn(new_version, collections)

        # 4. Private object cannot be approved before submission for review.
        with self.assertRaises(ValidationError):
            new_version.approve(user=self.user)

        # 5. Submit new version for review
        new_version.submit_for_review()
        new_version.refresh_from_db()

        # 6. Publish the new version
        new_version.approve(
            user=self.user
        )  # Assumes 'approve' method handles status change
        new_version.refresh_from_db()

        # 7. After publishing the new version, predecessor should be archived
        self.published_collection.refresh_from_db()
        self.assertEqual(
            self.published_collection.publication_status, Collection.STATUS_ARCHIVED
        )

        # 7. Verify new version is now published
        self.assertEqual(new_version.publication_status, Collection.STATUS_PUBLISHED)

        # 8. Verify both versions appear in the list view (if business logic requires)
        response = self.client.get(reverse("collection-list"), follow=True)
        collections = response.context["object_list"]
        self.assertIn(new_version, collections)
        self.assertNotIn(self.published_collection, collections)

        # 9. Verify predecessor/successor relationships
        self.assertIn(self.published_collection, new_version.predecessors.all())
        self.assertIn(new_version, self.published_collection.successors.all())

    def test_approve_only_allowed_from_review(self):
        """Test that approve() raises ValidationError if called on a non-review status."""
        for status in [
            Collection.STATUS_PRIVATE,
            Collection.STATUS_PUBLISHED,
            Collection.STATUS_ARCHIVED,
        ]:
            collection = Collection.objects.create(
                catchment=self.catchment,
                collector=self.collector,
                collection_system=self.collection_system,
                waste_stream=self.waste_stream,
                valid_from=date.today() + timedelta(days=100),
                valid_until=date.today() + timedelta(days=120),
                publication_status=status,
                owner=self.user,
                connection_type="VOLUNTARY",
            )
            with self.assertRaises(ValidationError) as cm:
                collection.approve(user=self.user)
            self.assertIn("Only objects in review can be approved", str(cm.exception))


class WasteFlyerListCheckUrlsViewTestCase(ViewWithPermissionsTestCase):
    member_permissions = "change_wasteflyer"

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        with mute_signals(signals.post_save):
            for i in range(1, 5):
                WasteFlyer.objects.create(
                    title=f"Waste flyer {i}",
                    abbreviation=f"WF{i}",
                    url_valid=i % 2 == 0,
                )

    @patch("case_studies.soilcom.views.check_wasteflyer_urls.delay")
    def test_get_http_200_ok_for_members(self, mock_delay):
        mock_task = mock_delay.return_value
        mock_task.task_id = "scheduler-task-id"
        self.client.force_login(self.member)
        response = self.client.get(reverse("wasteflyer-list-check-urls"))
        self.assertEqual(response.status_code, 200)
        mock_delay.assert_called_once()
        self.assertEqual(mock_delay.call_args.args[1], self.member.pk)
        self.assertEqual(response.json()["task_id"], "scheduler-task-id")


class WasteFlyerListCheckUrlsProgressViewTestCase(ViewWithPermissionsTestCase):
    member_permissions = "change_wasteflyer"

    @patch("case_studies.soilcom.views.AsyncResult")
    def test_get_uses_callback_task_state_after_scheduler_success(
        self, mock_async_result
    ):
        scheduler_result = type(
            "SchedulerResult",
            (object,),
            {"state": "SUCCESS", "result": "callback-task-id", "info": None},
        )()
        callback_result = type(
            "CallbackResult",
            (object,),
            {"state": "PENDING", "info": {"progress": 50}},
        )()
        mock_async_result.side_effect = [scheduler_result, callback_result]

        self.client.force_login(self.member)
        response = self.client.get(
            reverse(
                "wasteflyer-list-check-urls-progress",
                kwargs={"task_id": "scheduler-task-id"},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            response.content,
            {"state": "PENDING", "details": {"progress": 50}},
        )
        self.assertEqual(mock_async_result.call_count, 2)
        mock_async_result.assert_any_call("scheduler-task-id")
        mock_async_result.assert_any_call("callback-task-id")


# --- Regression tests -------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------


class CollectionFilterWithCatchmentAndPropertiesRegressionTest(
    ViewWithPermissionsTestCase
):
    """
    Regression test for the following bug: When filtering for a catchment of type 'custom', the following error
    occurred:
    ProgrammingError
    column "maps_geopolygon.geom" must appear in the GROUP BY clause or be used in an aggregate function
    LINE 1: ...s_geopolygon"."geom"::bytea HAVING (((((ST_Within("maps_geop...
    """

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        cls.region_geom_poly = Polygon(((0, 0), (0, 10), (10, 10), (10, 0), (0, 0)))
        cls.region_geom_multipoly = MultiPolygon(cls.region_geom_poly)
        cls.region = Region.objects.create(
            name="Test Region for GroupBy Test", geom=cls.region_geom_multipoly
        )
        cls.catchment = CollectionCatchment.objects.create(
            name="Test Catchment for GroupBy Test",
            region=cls.region,
            geom=cls.region.geom,
        )

        cls.collector = Collector.objects.create(name="Test Collector for GroupBy Test")
        cls.collection_system = CollectionSystem.objects.create(
            name="Test System for GroupBy Test"
        )
        cls.waste_category = WasteCategory.objects.create(
            name="Bio-waste for GroupBy Test"
        )

        MaterialCategory.objects.get_or_create(name="Biowaste component")

        cls.waste_component = WasteComponent.objects.create(
            name="Food scraps for GroupBy Test"
        )
        cls.waste_stream = WasteStream.objects.create(
            name="Organic household waste for GroupBy Test", category=cls.waste_category
        )
        cls.waste_stream.allowed_materials.add(cls.waste_component)

        cls.collection1 = Collection.objects.create(
            name="Test Collection 1 for GroupBy Test",
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.collection_system,
            waste_stream=cls.waste_stream,
            publication_status="published",
        )
        # Ensure this is the only published collection and has no published successors
        # (No successors created here)
        # If any other collections are created in the future, ensure they are not published or not successors of this one.

        # Property and Value to trigger the problematic annotation
        # Using get_or_create for Property to avoid issues if tests are run multiple times
        # and the "Connection rate" property might already exist from a previous run.
        cls.connection_rate_property, _ = Property.objects.get_or_create(
            name="Connection rate", defaults={"unit": "%"}
        )
        CollectionPropertyValue.objects.create(
            collection=cls.collection1,
            property=cls.connection_rate_property,
            average=75.0,
        )

        cls.list_url = reverse("collection-list")

    def test_geojson_with_connection_rate_filter_reproduces_programming_error(self):
        """
        Tests that the geojson endpoint with connection_rate_min filter and spatial query
        reproduces the ProgrammingError related to GROUP BY.
        """

        query_params = {
            "csrfmiddlewaretoken": "6FT2ft5HlbqXpjRUolmjklscDKHf9SSOVO1BMxxP5yxjE2iI9BXc7AoNIBSrpRP5",
            "catchment": self.catchment.pk,
            "collector": "",
            "collection_system": "",
            "filter": "Filter",
            "connection_type": "",
            "connection_rate_min": "0",
            "connection_rate_max": "100",
            "connection_rate_is_null": "true",
            "seasonal_frequency": "",
            "optional_frequency": "",
            "fee_system": "",
            "min_bin_size_min": "0",
            "min_bin_size_max": "120",
            "min_bin_size_is_null": "true",
            "required_bin_capacity_min": "0",
            "required_bin_capacity_max": "120",
            "required_bin_capacity_is_null": "true",
            "required_bin_capacity_reference": "",
            "collections_per_year_min": "0",
            "collections_per_year_max": "104",
            "collections_per_year_is_null": "true",
            "spec_waste_collected_min": "0",
            "spec_waste_collected_max": "516",
            "spec_waste_collected_is_null": "true",
        }

        self.client.force_login(self.member)
        response = self.client.get(self.list_url, query_params)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(
            response.context["filter"].data["catchment"], str(self.catchment.pk)
        )

        self.assertEqual(response.context["filter"].qs.count(), 1)

        self.assertEqual(
            response.context["filter"].qs.first().catchment, self.catchment
        )


class CollectionAddPropertyValueAnchoringTestCase(ViewWithPermissionsTestCase):
    member_permissions = ["add_collectionpropertyvalue"]
    url_name = "collection-add-property"

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.catchment = CollectionCatchment.objects.create(name="C")
        cls.system = CollectionSystem.objects.create(name="S")
        cls.category = WasteCategory.objects.create(name="Cat")
        cls.stream = WasteStream.objects.create(category=cls.category)
        cls.root = Collection.objects.create(
            name="Root",
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2020, 1, 1),
            publication_status="published",
            owner=cls.member,
        )
        cls.succ = Collection.objects.create(
            name="Succ",
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2021, 1, 1),
            publication_status="published",
            owner=cls.member,
        )
        cls.succ.predecessors.add(cls.root)

        cls.prop = Property.objects.create(
            name="AnchorProp", publication_status="published"
        )
        cls.unit = Unit.objects.create(
            name="AnchorUnit", publication_status="published"
        )
        cls.prop.allowed_units.add(cls.unit)

    def test_member_post_attaches_to_anchor(self):
        self.client.force_login(self.member)
        data = {
            "collection": self.succ.pk,
            "property": self.prop.pk,
            "unit": self.unit.pk,
            "year": 2022,
            "average": 12.3,
            "form-TOTAL_FORMS": "0",
            "form-INITIAL_FORMS": "0",
            "form-MIN_NUM_FORMS": "0",
            "form-MAX_NUM_FORMS": "1000",
        }
        response = self.client.post(
            reverse(self.url_name, kwargs={"pk": self.succ.pk}), data=data
        )
        self.assertEqual(response.status_code, 302)

        cpv = CollectionPropertyValue.objects.get(
            property=self.prop, unit=self.unit, year=2022
        )
        self.assertEqual(cpv.collection_id, (self.succ.version_anchor or self.succ).pk)


class CollectionPropertyValueUpdateReanchorTestCase(ViewWithPermissionsTestCase):
    member_permissions = ["change_collectionpropertyvalue"]

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.catchment = CollectionCatchment.objects.create(name="C")
        cls.system = CollectionSystem.objects.create(name="S")
        cls.category = WasteCategory.objects.create(name="Cat")
        cls.stream = WasteStream.objects.create(category=cls.category)
        cls.root = Collection.objects.create(
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2020, 1, 1),
            publication_status="published",
            owner=cls.member,
        )
        cls.succ = Collection.objects.create(
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2021, 1, 1),
            publication_status="published",
            owner=cls.member,
        )
        cls.succ.predecessors.add(cls.root)

        cls.prop = Property.objects.create(
            name="ReanchorProp", publication_status="published"
        )
        cls.unit = Unit.objects.create(
            name="ReanchorUnit", publication_status="published"
        )
        cls.prop.allowed_units.add(cls.unit)

        cls.cpv = CollectionPropertyValue.objects.create(
            owner=cls.member,
            collection=cls.succ,
            property=cls.prop,
            unit=cls.unit,
            year=2020,
            average=5,
            publication_status="private",
        )

    def test_update_reanchors_to_version_anchor(self):
        self.client.force_login(self.member)
        response = self.client.post(
            reverse("collectionpropertyvalue-update", kwargs={"pk": self.cpv.pk}),
            data={
                "collection": self.succ.pk,
                "property": self.prop.pk,
                "unit": self.unit.pk,
                "year": 2020,
                "average": 6,
                "form-TOTAL_FORMS": "0",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
            },
        )
        self.assertEqual(response.status_code, 302)

        cpv = CollectionPropertyValue.objects.get(pk=self.cpv.pk)
        self.assertEqual(cpv.collection_id, self.root.version_anchor.pk)
        self.assertEqual(cpv.average, 6)


class CollectionPropertyValueDeleteAnchorSemanticsTestCase(ViewWithPermissionsTestCase):
    member_permissions = ["delete_collectionpropertyvalue"]

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.catchment = CollectionCatchment.objects.create(name="C")
        cls.system = CollectionSystem.objects.create(name="S")
        cls.category = WasteCategory.objects.create(name="Cat")
        cls.stream = WasteStream.objects.create(category=cls.category)
        cls.root = Collection.objects.create(
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2020, 1, 1),
            publication_status="published",
            owner=cls.member,
        )
        cls.succ = Collection.objects.create(
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2021, 1, 1),
            publication_status="published",
            owner=cls.member,
        )
        cls.succ.predecessors.add(cls.root)

        cls.prop = Property.objects.create(
            name="DelProp", publication_status="published"
        )
        cls.unit = Unit.objects.create(name="DelUnit", publication_status="published")
        cls.prop.allowed_units.add(cls.unit)

        cls.anchor_value = CollectionPropertyValue.objects.create(
            owner=cls.member,
            collection=cls.root,
            property=cls.prop,
            unit=cls.unit,
            year=2020,
            average=10,
            publication_status="private",
        )
        cls.child_value = CollectionPropertyValue.objects.create(
            owner=cls.member,
            collection=cls.succ,
            property=cls.prop,
            unit=cls.unit,
            year=2020,
            average=11,
            publication_status="private",
        )

    def test_delete_child_also_deletes_anchor_duplicate(self):
        self.client.force_login(self.member)
        response = self.client.post(
            reverse(
                "collectionpropertyvalue-delete-modal",
                kwargs={"pk": self.child_value.pk},
            )
        )
        self.assertEqual(response.status_code, 302)

        self.assertFalse(
            CollectionPropertyValue.objects.filter(
                pk__in=[self.child_value.pk, self.anchor_value.pk]
            ).exists()
        )


class CollectionDetailChainAwareValuesTestCase(ViewWithPermissionsTestCase):
    url_name = "collection-detail"

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.catchment = CollectionCatchment.objects.create(name="C")
        cls.system = CollectionSystem.objects.create(name="S")
        cls.category = WasteCategory.objects.create(name="Cat")
        cls.stream = WasteStream.objects.create(category=cls.category)
        cls.root = Collection.objects.create(
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2020, 1, 1),
            publication_status="published",
        )
        cls.succ = Collection.objects.create(
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            valid_from=date(2021, 1, 1),
            publication_status="published",
        )
        cls.succ.predecessors.add(cls.root)

        cls.prop = Property.objects.create(
            name="ViewProp", publication_status="published"
        )
        cls.unit = Unit.objects.create(name="ViewUnit", publication_status="published")
        cls.prop.allowed_units.add(cls.unit)
        cls.anchor_value = CollectionPropertyValue.objects.create(
            collection=cls.root,
            property=cls.prop,
            unit=cls.unit,
            year=2020,
            average=77,
            publication_status="published",
        )

    def test_detail_context_includes_chain_values(self):
        response = self.client.get(reverse(self.url_name, kwargs={"pk": self.succ.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertIn("collection_property_values", response.context)
        vals = response.context["collection_property_values"]
        self.assertTrue(any(v.pk == self.anchor_value.pk for v in vals))


class CollectionReviewDetailPropertiesTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.staff = User.objects.create(username="moderator", is_staff=True)

        cls.catchment = CollectionCatchment.objects.create(name="RC")
        cls.system = CollectionSystem.objects.create(name="RS")
        cls.category = WasteCategory.objects.create(name="RCat")
        cls.stream = WasteStream.objects.create(category=cls.category)
        cls.collection = Collection.objects.create(
            name="R",
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            publication_status="published",
        )

        cls.prop = Property.objects.create(
            name="ReviewProp", publication_status="published"
        )
        cls.unit = Unit.objects.create(
            name="ReviewUnit", publication_status="published"
        )
        cls.prop.allowed_units.add(cls.unit)

        cls.cpv = CollectionPropertyValue.objects.create(
            collection=cls.collection,
            property=cls.prop,
            unit=cls.unit,
            year=2020,
            average=77,
            publication_status="published",
        )

        cls.agg = AggregatedCollectionPropertyValue.objects.create(
            property=cls.prop,
            unit=cls.unit,
            year=2021,
            average=88,
            publication_status="published",
        )
        cls.agg.collections.add(cls.collection)

        cls.ct_id = ContentType.objects.get_for_model(Collection).pk

    def test_review_detail_shows_cpv_and_aggregated(self):
        self.client.force_login(self.staff)
        url = reverse(
            "object_management:review_item_detail",
            kwargs={"content_type_id": self.ct_id, "object_id": self.collection.pk},
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        body = response.content.decode()
        self.assertIn("77", body)
        self.assertIn("ReviewUnit", body)
        self.assertIn("aggregated", body)


class CollectionDetailOnlyPublishedCpvsTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.catchment = CollectionCatchment.objects.create(name="PC")
        cls.system = CollectionSystem.objects.create(name="PS")
        cls.category = WasteCategory.objects.create(name="PCat")
        cls.stream = WasteStream.objects.create(category=cls.category)
        cls.collection = Collection.objects.create(
            name="PublishedCollection",
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            publication_status="published",
        )

        cls.prop = Property.objects.create(
            name="VisibilityProp", publication_status="published"
        )
        cls.published_unit = Unit.objects.create(
            name="PublishedUnit", publication_status="published"
        )
        cls.private_unit = Unit.objects.create(
            name="PrivateUnit", publication_status="published"
        )
        cls.prop.allowed_units.add(cls.published_unit, cls.private_unit)

        cls.published_value = CollectionPropertyValue.objects.create(
            collection=cls.collection,
            property=cls.prop,
            unit=cls.published_unit,
            year=2020,
            average=12,
            publication_status="published",
        )

        cls.private_value = CollectionPropertyValue.objects.create(
            collection=cls.collection,
            property=cls.prop,
            unit=cls.private_unit,
            year=2021,
            average=24,
            publication_status="private",
        )

    def test_public_detail_only_shows_published_cpvs(self):
        response = self.client.get(
            reverse("collection-detail", kwargs={"pk": self.collection.pk})
        )
        self.assertEqual(response.status_code, 200)
        body = response.content.decode()
        self.assertIn("PublishedUnit", body)
        self.assertNotIn("PrivateUnit", body)

    def test_staff_detail_only_shows_published_cpvs(self):
        staff = User.objects.create(username="staff-user", is_staff=True)
        self.client.force_login(staff)
        response = self.client.get(
            reverse("collection-detail", kwargs={"pk": self.collection.pk})
        )
        self.assertEqual(response.status_code, 200)
        context_vals = response.context["collection_property_values"]
        self.assertTrue(all(v.publication_status == "published" for v in context_vals))
        self.assertNotIn("PrivateUnit", response.content.decode())


class CollectionReviewDetailPreviewTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.staff = User.objects.create(username="reviewer", is_staff=True)

        cls.catchment = CollectionCatchment.objects.create(name="RC")
        cls.system = CollectionSystem.objects.create(name="RS")
        cls.category = WasteCategory.objects.create(name="RCat")
        cls.stream = WasteStream.objects.create(category=cls.category)
        cls.collection = Collection.objects.create(
            name="ReviewCollection",
            catchment=cls.catchment,
            collection_system=cls.system,
            waste_stream=cls.stream,
            publication_status="review",
        )

        cls.prop = Property.objects.create(
            name="PreviewProp", publication_status="published"
        )
        cls.unit = Unit.objects.create(
            name="PreviewUnit", publication_status="published"
        )
        cls.other_unit = Unit.objects.create(
            name="OtherUnit", publication_status="published"
        )
        cls.prop.allowed_units.add(cls.unit, cls.other_unit)

        cls.cpv_published = CollectionPropertyValue.objects.create(
            collection=cls.collection,
            property=cls.prop,
            unit=cls.unit,
            year=2020,
            average=10,
            publication_status="published",
        )

        cls.cpv_review = CollectionPropertyValue.objects.create(
            collection=cls.collection,
            property=cls.prop,
            unit=cls.unit,
            year=2020,
            average=12,
            publication_status="review",
        )

        cls.cpv_private = CollectionPropertyValue.objects.create(
            collection=cls.collection,
            property=cls.prop,
            unit=cls.other_unit,
            year=2021,
            average=5,
            publication_status="private",
        )

        cls.agg_prop = Property.objects.create(
            name="PreviewAggProp", publication_status="published"
        )
        cls.agg_unit = Unit.objects.create(
            name="AggUnit", publication_status="published"
        )
        cls.agg_prop.allowed_units.add(cls.agg_unit)

        cls.agg_published = AggregatedCollectionPropertyValue.objects.create(
            property=cls.agg_prop,
            unit=cls.agg_unit,
            year=2019,
            average=40,
            publication_status="published",
        )
        cls.agg_published.collections.add(cls.collection)

        cls.agg_review = AggregatedCollectionPropertyValue.objects.create(
            property=cls.agg_prop,
            unit=cls.agg_unit,
            year=2019,
            average=45,
            publication_status="review",
        )
        cls.agg_review.collections.add(cls.collection)

        cls.agg_private = AggregatedCollectionPropertyValue.objects.create(
            property=cls.agg_prop,
            unit=cls.agg_unit,
            year=2021,
            average=55,
            publication_status="private",
        )
        cls.agg_private.collections.add(cls.collection)

        cls.ct_id = ContentType.objects.get_for_model(Collection).pk

        from materials.models import Material

        cls.allowed_material = Material.objects.create(name="Allowed Material")
        cls.forbidden_material = Material.objects.create(name="Forbidden Material")
        cls.stream.allowed_materials.add(cls.allowed_material)
        cls.stream.forbidden_materials.add(cls.forbidden_material)

    def test_review_preview_shows_published_and_review_only(self):
        self.client.force_login(self.staff)
        url = reverse(
            "object_management:review_item_detail",
            kwargs={"content_type_id": self.ct_id, "object_id": self.collection.pk},
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        cpvs = response.context["collection_property_values"]
        self.assertEqual([v.pk for v in cpvs], [self.cpv_review.pk])
        self.assertTrue(
            all(v.publication_status in {"published", "review"} for v in cpvs)
        )

        agg_vals = response.context["aggregated_collection_property_values"]
        self.assertEqual([v.pk for v in agg_vals], [self.agg_review.pk])
        self.assertTrue(
            all(v.publication_status in {"published", "review"} for v in agg_vals)
        )

        body = response.content.decode()
        self.assertIn("12", body)
        self.assertIn("45", body)
        self.assertNotIn("Private", body)

    def test_review_detail_shows_allowed_and_forbidden_materials(self):
        self.client.force_login(self.staff)
        url = reverse(
            "object_management:review_item_detail",
            kwargs={"content_type_id": self.ct_id, "object_id": self.collection.pk},
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        allowed_materials = response.context["allowed_materials"]
        forbidden_materials = response.context["forbidden_materials"]

        self.assertEqual(len(allowed_materials), 1)
        self.assertEqual(len(forbidden_materials), 1)
        self.assertEqual(allowed_materials[0].name, "Allowed Material")
        self.assertEqual(forbidden_materials[0].name, "Forbidden Material")

        body = response.content.decode()
        self.assertIn("Allowed Materials", body)
        self.assertIn("Forbidden Materials", body)
        self.assertIn("Allowed Material", body)
        self.assertIn("Forbidden Material", body)


class WasteAtlasMapViewsTestCase(TestCase):
    """Tests for Waste Atlas map pages rendered via template views."""

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username="atlas-user", password="secret")
        waste_atlas_group, _ = Group.objects.get_or_create(name="waste_atlas")
        cls.user.groups.add(waste_atlas_group)

    def setUp(self):
        self.client.force_login(self.user)

    def test_italy_orga_level_map_defaults_to_it_and_english_labels(self):
        """Italy orga-level map defaults to country IT and English text."""
        response = self.client.get(reverse("waste-atlas-orga-level-italy-map"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="IT" selected')
        self.assertContains(response, "Administrative level of waste collection")
        self.assertContains(response, "Map overview")
        self.assertContains(response, "No data")

    def test_country_specific_orga_level_maps_default_to_expected_country(self):
        """Country-specific orga-level maps default to expected country and year."""
        map_defaults = {
            "waste-atlas-orga-level-italy-map": ("IT", "2024"),
            "waste-atlas-orga-level-sweden-map": ("SE", "2024"),
            "waste-atlas-orga-level-denmark-map": ("DK", "2023"),
            "waste-atlas-orga-level-netherlands-map": ("NL", "2024"),
            "waste-atlas-orga-level-belgium-map": ("BE", "2024"),
        }

        for url_name, (expected_country, expected_year) in map_defaults.items():
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, f'value="{expected_country}" selected')
                self.assertContains(response, f'value="{expected_year}" selected')
                self.assertContains(
                    response, "Administrative level of waste collection"
                )
                self.assertContains(response, "Map overview")
                self.assertContains(response, "No data")

    def test_italy_orga_level_map_allows_country_override(self):
        """Italy orga-level map still allows overriding country via query param."""
        response = self.client.get(
            reverse("waste-atlas-orga-level-italy-map"),
            {"country": "DE"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="DE" selected')

    def test_waste_atlas_overview_includes_italy_orga_level_entry(self):
        """Overview page lists all country-specific organizational-level maps."""
        response = self.client.get(reverse("waste-atlas-overview"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("waste-atlas-orga-level-italy-map"))
        self.assertContains(
            response,
            "Map 29 — Administrative level of waste collection (Italy, EN)",
        )
        self.assertContains(response, reverse("waste-atlas-orga-level-sweden-map"))
        self.assertContains(
            response,
            "Map 30 — Administrative level of waste collection (Sweden, EN)",
        )
        self.assertContains(
            response,
            f"{reverse('waste-atlas-collection-system-map')}?country=SE&amp;year=2023",
        )
        self.assertContains(
            response,
            f"{reverse('waste-atlas-connection-rate-map')}?country=SE&amp;year=2023",
        )
        self.assertContains(
            response,
            f"{reverse('waste-atlas-biowaste-collection-amount-map')}?country=SE&amp;year=2023",
        )
        self.assertContains(
            response,
            f"{reverse('waste-atlas-organic-waste-ratio-map')}?country=SE&amp;year=2023",
        )
        self.assertContains(response, reverse("waste-atlas-orga-level-denmark-map"))
        self.assertContains(
            response,
            "Map 31 — Administrative level of waste collection (Denmark, EN)",
        )
        self.assertContains(
            response,
            reverse("waste-atlas-orga-level-netherlands-map"),
        )
        self.assertContains(
            response,
            "Map 32 — Administrative level of waste collection (The Netherlands, EN)",
        )
        self.assertContains(response, reverse("waste-atlas-orga-level-belgium-map"))
        self.assertContains(
            response,
            "Map 33 — Administrative level of waste collection (Belgium, EN)",
        )


class CollectionCascadeMixinTestCase(TestCase):
    """Test the CollectionReviewActionCascadeMixin functionality."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        # Create users
        cls.owner = User.objects.create_user(
            username="owner", email="owner@test.com", password="pass"
        )
        cls.other_owner = User.objects.create_user(
            username="other", email="other@test.com", password="pass"
        )
        cls.staff = User.objects.create_user(
            username="staff", email="staff@test.com", password="pass", is_staff=True
        )

        # Create base data
        cls.catchment = CollectionCatchment.objects.create(
            name="Test Catchment", publication_status="published"
        )
        cls.system = CollectionSystem.objects.create(
            name="Test System", publication_status="published"
        )
        cls.category = WasteCategory.objects.create(
            name="Test Category", publication_status="published"
        )
        cls.stream = WasteStream.objects.create(
            category=cls.category, publication_status="published"
        )

        # Create properties
        cls.prop1 = Property.objects.create(
            name="Property 1", publication_status="published"
        )
        cls.prop2 = Property.objects.create(
            name="Property 2", publication_status="published"
        )
        cls.unit1 = Unit.objects.create(name="Unit 1", publication_status="published")
        cls.unit2 = Unit.objects.create(name="Unit 2", publication_status="published")
        cls.prop1.allowed_units.add(cls.unit1)
        cls.prop2.allowed_units.add(cls.unit2)

        cls.factory = RequestFactory()

    def _create_collection(self, name, owner, status="private", valid_from=None):
        """Helper to create a collection."""
        return Collection.objects.create(
            name=name,
            catchment=self.catchment,
            collection_system=self.system,
            waste_stream=self.stream,
            valid_from=valid_from or date(2020, 1, 1),
            publication_status=status,
            owner=owner,
        )

    def _create_cpv(self, collection, prop, unit, owner, status="private", **kwargs):
        """Helper to create a collection property value."""
        return CollectionPropertyValue.objects.create(
            collection=collection,
            property=prop,
            unit=unit,
            owner=owner,
            publication_status=status,
            year=kwargs.get("year", 2020),
            average=kwargs.get("average", 10.0),
        )

    def _create_acpv(self, collections, prop, unit, owner, status="private", **kwargs):
        """Helper to create an aggregated collection property value."""
        acpv = AggregatedCollectionPropertyValue.objects.create(
            property=prop,
            unit=unit,
            owner=owner,
            publication_status=status,
            year=kwargs.get("year", 2020),
            average=kwargs.get("average", 20.0),
        )
        acpv.collections.set(collections)
        return acpv


class SubmitCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test submit_for_review cascade via CollectionSubmitForReviewView."""

    def test_submit_cascades_to_owner_cpvs(self):
        """Submit cascades to owner's private and declined CPVs."""
        collection = self._create_collection("C1", self.owner, status="private")
        cpv_private = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "private", year=2020
        )
        cpv_declined = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "declined", year=2021
        )
        cpv_published = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "published", year=2022
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        cpv_private.refresh_from_db()
        cpv_declined.refresh_from_db()
        cpv_published.refresh_from_db()

        self.assertEqual(cpv_private.publication_status, "review")
        self.assertEqual(cpv_declined.publication_status, "review")
        self.assertEqual(cpv_published.publication_status, "published")  # Unchanged

    def test_submit_includes_collaborator_cpvs(self):
        """Submit cascades to all CPVs on owner's collection, including collaborators' CPVs."""
        collection = self._create_collection("C1", self.owner, status="private")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "private"
        )
        # Other user contributed a CPV to owner's collection
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "private"
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both cascade because they're on the owner's collection
        self.assertEqual(cpv_owner.publication_status, "review")
        self.assertEqual(cpv_other.publication_status, "review")  # Also cascaded

    def test_submit_cascades_to_acpvs(self):
        """Submit cascades to owner's aggregated property values."""
        collection = self._create_collection("C1", self.owner, status="private")
        acpv = self._create_acpv(
            [collection], self.prop1, self.unit1, self.owner, "private"
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        acpv.refresh_from_db()
        self.assertEqual(acpv.publication_status, "review")

    def test_submit_cascades_across_version_chain(self):
        """Submit cascades to CPVs on all versions in the chain."""
        v1 = self._create_collection("V1", self.owner, "published", date(2020, 1, 1))
        v2 = self._create_collection("V2", self.owner, "published", date(2021, 1, 1))
        v3 = self._create_collection("V3", self.owner, "private", date(2022, 1, 1))
        v2.predecessors.add(v1)
        v3.predecessors.add(v2)

        cpv1 = self._create_cpv(v1, self.prop1, self.unit1, self.owner, "private")
        cpv2 = self._create_cpv(v2, self.prop1, self.unit1, self.owner, "declined")
        cpv3 = self._create_cpv(v3, self.prop1, self.unit1, self.owner, "private")

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionSubmitForReviewView()
        view.request = request
        view.object = v3
        view.action_attr_name = "submit_for_review"
        view.post_action_hook(request, "private")

        cpv1.refresh_from_db()
        cpv2.refresh_from_db()
        cpv3.refresh_from_db()

        self.assertEqual(cpv1.publication_status, "review")
        self.assertEqual(cpv2.publication_status, "review")
        self.assertEqual(cpv3.publication_status, "review")


class WithdrawCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test withdraw_from_review cascade via CollectionWithdrawFromReviewView."""

    def test_withdraw_cascades_to_owner_cpvs(self):
        """Withdraw cascades to owner's CPVs in review."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv = self._create_cpv(collection, self.prop1, self.unit1, self.owner, "review")

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionWithdrawFromReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "withdraw_from_review"
        view.post_action_hook(request, "review")

        cpv.refresh_from_db()
        self.assertEqual(cpv.publication_status, "private")

    def test_withdraw_includes_collaborator_cpvs(self):
        """Withdraw cascades to all CPVs on owner's collection, including collaborators'."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "review"
        )
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.owner

        view = CollectionWithdrawFromReviewView()
        view.request = request
        view.object = collection
        view.action_attr_name = "withdraw_from_review"
        view.post_action_hook(request, "review")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both cascade because they're on the owner's collection
        self.assertEqual(cpv_owner.publication_status, "private")
        self.assertEqual(cpv_other.publication_status, "private")  # Also cascaded


class ApproveCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test approve cascade via CollectionApproveItemView."""

    def test_approve_cascades_to_all_cpvs_in_review(self):
        """Approve cascades to ALL CPVs in review, regardless of owner."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "review"
        )
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionApproveItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "approve"
        view.post_action_hook(request, "review")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both should be approved
        self.assertEqual(cpv_owner.publication_status, "published")
        self.assertEqual(cpv_other.publication_status, "published")
        self.assertEqual(cpv_owner.approved_by, self.staff)
        self.assertEqual(cpv_other.approved_by, self.staff)

    def test_approve_cascades_to_acpvs(self):
        """Approve cascades to aggregated property values."""
        collection = self._create_collection("C1", self.owner, status="review")
        acpv = self._create_acpv(
            [collection], self.prop1, self.unit1, self.owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionApproveItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "approve"
        view.post_action_hook(request, "review")

        acpv.refresh_from_db()
        self.assertEqual(acpv.publication_status, "published")
        self.assertEqual(acpv.approved_by, self.staff)

    def test_approve_cascades_across_version_chain(self):
        """Approve cascades to CPVs on all versions."""
        v1 = self._create_collection("V1", self.owner, "published", date(2020, 1, 1))
        v2 = self._create_collection("V2", self.owner, "review", date(2021, 1, 1))
        v2.predecessors.add(v1)

        cpv1 = self._create_cpv(v1, self.prop1, self.unit1, self.owner, "review")
        cpv2 = self._create_cpv(v2, self.prop1, self.unit1, self.owner, "review")

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionApproveItemView()
        view.request = request
        view.object = v2
        view.action_attr_name = "approve"
        view.post_action_hook(request, "review")

        cpv1.refresh_from_db()
        cpv2.refresh_from_db()

        self.assertEqual(cpv1.publication_status, "published")
        self.assertEqual(cpv2.publication_status, "published")


class RejectCascadeMixinTest(CollectionCascadeMixinTestCase):
    """Test reject cascade via CollectionRejectItemView."""

    def test_reject_cascades_to_all_cpvs_in_review(self):
        """Reject cascades to ALL CPVs in review, regardless of owner."""
        collection = self._create_collection("C1", self.owner, status="review")
        cpv_owner = self._create_cpv(
            collection, self.prop1, self.unit1, self.owner, "review"
        )
        cpv_other = self._create_cpv(
            collection, self.prop2, self.unit2, self.other_owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionRejectItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "reject"
        view.post_action_hook(request, "review")

        cpv_owner.refresh_from_db()
        cpv_other.refresh_from_db()

        # Both should be rejected
        self.assertEqual(cpv_owner.publication_status, "declined")
        self.assertEqual(cpv_other.publication_status, "declined")

    def test_reject_cascades_to_acpvs(self):
        """Reject cascades to aggregated property values."""
        collection = self._create_collection("C1", self.owner, status="review")
        acpv = self._create_acpv(
            [collection], self.prop1, self.unit1, self.owner, "review"
        )

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionRejectItemView()
        view.request = request
        view.object = collection
        view.action_attr_name = "reject"
        view.post_action_hook(request, "review")

        acpv.refresh_from_db()
        self.assertEqual(acpv.publication_status, "declined")

    def test_reject_cascades_across_version_chain(self):
        """Reject cascades to CPVs on all versions."""
        v1 = self._create_collection("V1", self.owner, "published", date(2020, 1, 1))
        v2 = self._create_collection("V2", self.owner, "review", date(2021, 1, 1))
        v2.predecessors.add(v1)

        cpv1 = self._create_cpv(v1, self.prop1, self.unit1, self.owner, "review")
        cpv2 = self._create_cpv(v2, self.prop1, self.unit1, self.owner, "review")

        request = self.factory.post("/")
        request.user = self.staff

        view = CollectionRejectItemView()
        view.request = request
        view.object = v2
        view.action_attr_name = "reject"
        view.post_action_hook(request, "review")

        cpv1.refresh_from_db()
        cpv2.refresh_from_db()

        self.assertEqual(cpv1.publication_status, "declined")
        self.assertEqual(cpv2.publication_status, "declined")


class ReviewSubmissionConsistencyTests(TestCase):
    """Regression tests ensuring review-submission timestamps remain consistent."""

    def setUp(self):
        self.owner = User.objects.create_user(username="owner")
        self.moderator = User.objects.create_user(username="moderator", is_staff=True)

        content_type = ContentType.objects.get_for_model(Collection)
        permission, _ = Permission.objects.get_or_create(
            codename="can_moderate_collection",
            content_type=content_type,
            defaults={"name": "Can moderate collections"},
        )
        self.moderator.user_permissions.add(permission)

        self.collection = Collection.objects.create(
            name="Test Collection",
            owner=self.owner,
            publication_status=Collection.STATUS_PRIVATE,
        )

    def test_initial_submission_consistency(self):
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        self.collection.refresh_from_db()
        self.assertIsNotNone(self.collection.submitted_at)

        actions = ReviewAction.objects.filter(
            content_type=ContentType.objects.get_for_model(Collection),
            object_id=self.collection.pk,
            action=ReviewAction.ACTION_SUBMITTED,
        )
        self.assertEqual(actions.count(), 1)

        latest_action = actions.order_by("-created_at", "-id").first()
        self.assertIsNotNone(latest_action)

        time_diff = abs(self.collection.submitted_at - latest_action.created_at)
        self.assertLess(time_diff.total_seconds(), 1.0)

    def test_resubmission_consistency(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()
        initial_submitted_at = self.collection.submitted_at
        initial_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )

        response = self.client.post(
            reverse(
                "object_management:withdraw_from_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()
        self.assertIsNone(self.collection.submitted_at)

        time.sleep(0.01)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()

        latest_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )

        self.assertIsNotNone(self.collection.submitted_at)
        self.assertIsNotNone(latest_action)
        time_diff = abs(self.collection.submitted_at - latest_action.created_at)
        self.assertLess(time_diff.total_seconds(), 1.0)

        self.assertGreater(self.collection.submitted_at, initial_submitted_at)
        self.assertGreater(latest_action.created_at, initial_action.created_at)

    def test_review_ui_shows_latest_submission(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        time.sleep(0.01)

        response = self.client.post(
            reverse(
                "object_management:withdraw_from_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        time.sleep(0.01)

        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()

        self.client.force_login(self.moderator)
        response = self.client.get(
            reverse(
                "object_management:review_item_detail",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            )
        )
        self.assertEqual(response.status_code, 200)

        self.assertIn("review_submitted_action", response.context)
        submitted_action = response.context["review_submitted_action"]
        self.assertIsNotNone(submitted_action)
        self.assertEqual(submitted_action.action, ReviewAction.ACTION_SUBMITTED)

        latest_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )
        self.assertEqual(submitted_action.pk, latest_action.pk)

        time_diff = abs(submitted_action.created_at - self.collection.submitted_at)
        self.assertLess(time_diff.total_seconds(), 1.0)

    def test_approval_preserves_submission_timestamp(self):
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse(
                "object_management:submit_for_review",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()
        submitted_at_before = self.collection.submitted_at

        self.client.force_login(self.moderator)
        response = self.client.post(
            reverse(
                "object_management:approve_item",
                kwargs={
                    "content_type_id": ContentType.objects.get_for_model(Collection).pk,
                    "object_id": self.collection.pk,
                },
            ),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.collection.refresh_from_db()

        self.assertEqual(self.collection.submitted_at, submitted_at_before)

        latest_action = (
            ReviewAction.objects.filter(
                content_type=ContentType.objects.get_for_model(Collection),
                object_id=self.collection.pk,
                action=ReviewAction.ACTION_SUBMITTED,
            )
            .order_by("-created_at", "-id")
            .first()
        )
        time_diff = abs(self.collection.submitted_at - latest_action.created_at)
        self.assertLess(time_diff.total_seconds(), 1.0)


@patch("case_studies.soilcom.tests.test_views.check_wasteflyer_urls.apply")
@patch("case_studies.soilcom.tests.test_views.chord")
class CheckWasteFlyerUrlsTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        with mute_signals(signals.post_save):
            for i in range(1, 5):
                WasteFlyer.objects.create(
                    title=f"Waste flyer {i}",
                    abbreviation=f"WF{i}",
                    url_valid=i % 2 == 0,
                )

    def setUp(self):
        self.flyer = WasteFlyer.objects.first

    def test_initial(self, mock_chord, mock_apply):
        mock_async_result = namedtuple("MockAsyncResult", ["status", "get"])
        mock_apply.return_value = mock_async_result(status="SUCCESS", get=lambda: None)
        self.assertEqual(4, WasteFlyer.objects.count())
        params = {
            "csrfmiddlewaretoken": [
                "Hm7MXB2NjRCOIpNbGaRKR87VCHM5KwpR1t4AdZFgaqKfqui1EJwhKKmkxFKDfL3h"
            ],
            "url_valid": ["False"],
            "page": ["2"],
        }
        qdict = QueryDict("", mutable=True)
        qdict.update(MultiValueDict(params))
        newparams = qdict.copy()
        newparams.pop("csrfmiddlewaretoken")
        newparams.pop("page")
        result = check_wasteflyer_urls.apply(args=[newparams])
        while result.status == "PENDING":
            self.assertEqual("PENDING", result.status)
        if result.status == "FAILURE":
            result.get()
        self.assertEqual("SUCCESS", result.status)

    def test_chord(self, mock_chord, mock_apply):
        mock_chord.return_value = lambda x: type(
            "task", (object,), {"task_id": "fake_task_id"}
        )
        mock_apply.side_effect = [
            type("task", (object,), {"status": "SUCCESS"})
            for _ in WasteFlyer.objects.all()
        ]
        callback = check_wasteflyer_urls_callback.s()
        header = [
            check_wasteflyer_url.s(flyer.pk) for flyer in WasteFlyer.objects.all()
        ]
        result = chord(header)(callback)
        self.assertEqual(result.task_id, "fake_task_id")


@patch("case_studies.soilcom.tasks.chord")
class CheckWasteFlyerUrlsScopeRegressionTestCase(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user("task_owner")
        self.other_owner = User.objects.create_user("task_other_owner")

        with mute_signals(signals.post_save):
            self.owner_flyers = [
                WasteFlyer.objects.create(
                    owner=self.owner,
                    title="Owner flyer 1",
                    abbreviation="OwnerWF1",
                ),
                WasteFlyer.objects.create(
                    owner=self.owner,
                    title="Owner flyer 2",
                    abbreviation="OwnerWF2",
                ),
            ]
            WasteFlyer.objects.create(
                owner=self.other_owner,
                title="Other owner flyer",
                abbreviation="OtherWF",
            )

    def test_private_scope_checks_only_requesting_users_flyers(self, mock_chord):
        captured_header = {}

        def fake_chord(header):
            captured_header["header"] = header
            return lambda _: type("task", (object,), {"task_id": "callback-task-id"})()

        mock_chord.side_effect = fake_chord

        params = QueryDict("scope=private", mutable=True)
        task_id = check_wasteflyer_urls.run(params, self.owner.pk)

        self.assertEqual(task_id, "callback-task-id")
        scheduled_ids = sorted(
            signature.args[0] for signature in captured_header["header"]
        )
        self.assertEqual(
            scheduled_ids,
            sorted(flyer.pk for flyer in self.owner_flyers),
        )


@patch("case_studies.soilcom.tasks.find_wayback_snapshot_for_year")
@patch("case_studies.soilcom.tasks.check_url")
class CheckWasteFlyerUrlWaybackFallbackTestCase(TestCase):
    def setUp(self):
        with mute_signals(signals.post_save):
            self.flyer = WasteFlyer.objects.create(
                title="Waste flyer",
                abbreviation="WF",
                url="https://example.com/dead-flyer.pdf",
            )

        self.collection = Collection.objects.create(valid_from=date(2021, 1, 1))
        self.collection.flyers.add(self.flyer)

    def test_replaces_broken_url_with_year_snapshot(self, mock_check_url, mock_wayback):
        original_url = self.flyer.url
        mock_check_url.return_value = False
        mock_wayback.return_value = (
            "https://web.archive.org/web/20211230153000/"
            "https://example.com/dead-flyer.pdf"
        )

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertTrue(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, mock_wayback.return_value)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_keeps_original_url_when_no_snapshot_exists(
        self, mock_check_url, mock_wayback
    ):
        original_url = self.flyer.url
        mock_check_url.return_value = False
        mock_wayback.return_value = None

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertFalse(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, original_url)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_replaces_live_url_with_year_snapshot(self, mock_check_url, mock_wayback):
        original_url = self.flyer.url
        mock_check_url.return_value = True
        mock_wayback.return_value = (
            "https://web.archive.org/web/20211230153000/"
            "https://example.com/dead-flyer.pdf"
        )

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertTrue(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, mock_wayback.return_value)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_keeps_live_url_when_no_snapshot_exists(self, mock_check_url, mock_wayback):
        original_url = self.flyer.url
        mock_check_url.return_value = True
        mock_wayback.return_value = None

        check_wasteflyer_url(self.flyer.pk)

        self.flyer.refresh_from_db()
        self.assertTrue(self.flyer.url_valid)
        self.assertEqual(self.flyer.url, original_url)
        mock_wayback.assert_called_once_with(original_url, 2021)

    def test_skips_wayback_lookup_when_url_is_already_archived(
        self, mock_check_url, mock_wayback
    ):
        with mute_signals(signals.post_save):
            self.flyer.url = (
                "https://web.archive.org/web/20211230153000/"
                "https://example.com/dead-flyer.pdf"
            )
            self.flyer.save()
        mock_check_url.return_value = True

        check_wasteflyer_url(self.flyer.pk)

        mock_wayback.assert_not_called()

    def test_returns_false_when_wasteflyer_was_deleted(
        self, mock_check_url, mock_wayback
    ):
        flyer_pk = self.flyer.pk
        self.flyer.delete()

        result = check_wasteflyer_url(flyer_pk)

        self.assertFalse(result)
        mock_check_url.assert_not_called()
        mock_wayback.assert_not_called()


class CollectionCSVRendererTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        MaterialCategory.objects.create(name="Biowaste component")
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed Material 1"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed Material 2"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden Material 1"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden Material 2"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=WasteCategory.objects.create(name="Test category"),
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            waste_flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123", url="https://www.test-flyer.org"
            )
        frequency = CollectionFrequency.objects.create(name="Test Frequency")
        nuts = NutsRegion.objects.create(
            name="Test NUTS", nuts_id="DE123", cntr_code="DE"
        )
        catchment = CollectionCatchment.objects.create(
            name="Test catchment", region=nuts.region_ptr
        )
        for i in range(1, 3):
            collection = Collection.objects.create(
                name=f"collection{i}",
                catchment=catchment,
                collector=Collector.objects.create(name=f"collector{1}"),
                collection_system=CollectionSystem.objects.create(name="Test system"),
                waste_stream=waste_stream,
                fee_system=FeeSystem.objects.create(name="Fixed fee"),
                frequency=frequency,
                valid_from=date(2020, 1, 1),
                description="This is a test case.",
            )
            collection.flyers.add(waste_flyer)

    def setUp(self):
        self.file = BytesIO()
        self.content = CollectionFlatSerializer(
            Collection.objects.all(), many=True
        ).data

    def test_fieldnames_in_right_order(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        fieldnames = [renderer.labels[key] for key in renderer.header]
        self.assertListEqual(fieldnames, list(reader.fieldnames))
        self.assertEqual(2, sum(1 for _ in reader))
        self.assertIn("Connection type", reader.fieldnames)
        self.assertEqual(
            renderer.header.index("connection_type"),
            reader.fieldnames.index("Connection type"),
        )

    def test_connection_type_field_exported(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        valid_labels = [
            "Compulsory",
            "Voluntary",
            "Mandatory",
            "Mandatory with exception for home composters",
            "Not specified",
            "",
        ]
        for row in reader:
            self.assertIn(row["Connection type"], valid_labels)

    def test_allowed_materials_formatted_as_comma_separated_list_in_one_field(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertEqual(
                "Allowed Material 1, Allowed Material 2", row["Allowed Materials"]
            )

    def test_forbidden_materials_formatted_as_comma_separated_list_in_one_field(self):
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        for row in reader:
            self.assertEqual(
                "Forbidden Material 1, Forbidden Material 2", row["Forbidden Materials"]
            )

    def test_regression_flyers_without_urls_dont_raise_type_error(self):
        defected_collection = Collection.objects.first()
        with mute_signals(signals.post_save):
            rogue_flyer = WasteFlyer.objects.create(
                title="Rogue flyer without url", abbreviation="RF"
            )
        defected_collection.flyers.add(rogue_flyer)
        renderer = CollectionCSVRenderer()
        renderer.render(self.file, self.content)
        self.file.seek(0)
        reader = csv.DictReader(codecs.getreader("utf-8")(self.file), delimiter="\t")
        self.assertEqual(Collection.objects.count(), len(list(reader)))


class CollectionXLSXRendererTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        User.objects.create(username="outsider")
        member = User.objects.create(username="member")
        content_type = ContentType.objects.get_for_model(Collection)
        permission, _ = Permission.objects.get_or_create(
            codename="add_collection",
            content_type=content_type,
            defaults={"name": "Can add collection"},
        )
        member.user_permissions.add(permission)

        MaterialCategory.objects.create(name="Biowaste component")
        cls.allowed_material_1 = WasteComponent.objects.create(
            name="Allowed Material 1"
        )
        cls.allowed_material_2 = WasteComponent.objects.create(
            name="Allowed Material 2"
        )
        cls.forbidden_material_1 = WasteComponent.objects.create(
            name="Forbidden Material 1"
        )
        cls.forbidden_material_2 = WasteComponent.objects.create(
            name="Forbidden Material 2"
        )
        waste_stream = WasteStream.objects.create(
            name="Test waste stream",
            category=WasteCategory.objects.create(name="Test category"),
        )
        waste_stream.allowed_materials.add(cls.allowed_material_1)
        waste_stream.allowed_materials.add(cls.allowed_material_2)
        waste_stream.forbidden_materials.add(cls.forbidden_material_1)
        waste_stream.forbidden_materials.add(cls.forbidden_material_2)
        with mute_signals(signals.post_save):
            waste_flyer = WasteFlyer.objects.create(
                abbreviation="WasteFlyer123", url="https://www.test-flyer.org"
            )
        frequency = CollectionFrequency.objects.create(name="Test Frequency")
        nuts = NutsRegion.objects.create(
            name="Test NUTS", nuts_id="DE123", cntr_code="DE"
        )
        catchment = CollectionCatchment.objects.create(
            name="Test catchment", region=nuts.region_ptr
        )
        for i in range(1, 3):
            collection = Collection.objects.create(
                name=f"collection{i}",
                catchment=catchment,
                collector=Collector.objects.create(name=f"collector{1}"),
                collection_system=CollectionSystem.objects.create(name="Test system"),
                waste_stream=waste_stream,
                frequency=frequency,
                description="This is a test case.",
            )
            collection.flyers.add(waste_flyer)

    def setUp(self):
        self.file = BytesIO()

    def test_contains_all_labels_in_right_order(self):
        renderer = CollectionXLSXRenderer()
        qs = Collection.objects.all()
        content = CollectionFlatSerializer(qs, many=True).data
        renderer.render(self.file, content)
        wb = load_workbook(self.file)
        ws = wb.active
        ordered_content = [
            {k: row.get(k) for k in list(renderer.labels.keys())} for row in content
        ]
        for column, (key, _value) in enumerate(ordered_content[0].items(), start=1):
            self.assertEqual(renderer.labels[key], ws.cell(row=1, column=column).value)


@override_settings(
    SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=None,
    SOILCOM_TOTAL_WASTE_PROPERTY_ID=None,
    SOILCOM_SPECIFIC_WASTE_UNIT_ID=None,
    SOILCOM_TOTAL_WASTE_UNIT_ID=None,
    SOILCOM_POPULATION_ATTRIBUTE_ID=None,
    SOILCOM_SPECIFIC_WASTE_PROPERTY_NAME="specific waste collected [test]",
    SOILCOM_TOTAL_WASTE_PROPERTY_NAME="total waste collected [test]",
    SOILCOM_SPECIFIC_WASTE_UNIT_NAME="kg/(cap.*a) [test]",
    SOILCOM_TOTAL_WASTE_UNIT_NAME="Mg/a [test]",
    SOILCOM_POPULATION_ATTRIBUTE_NAME="Population [test]",
)
class DerivedValuesTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.property_specific = Property.objects.create(
            name="specific waste collected [test]"
        )
        cls.property_total = Property.objects.create(
            name="total waste collected [test]"
        )
        cls.unit_specific = Unit.objects.create(name="kg/(cap.*a) [test]")
        cls.unit_total = Unit.objects.create(name="Mg/a [test]")
        cls.population_attribute = Attribute.objects.create(
            name="Population [test]",
            unit="cap",
        )

        cls.collection_system = CollectionSystem.objects.create(
            name="Collection system"
        )
        cls.waste_category = WasteCategory.objects.create(name="Waste category")
        cls.waste_stream = WasteStream.objects.create(
            name="Waste stream",
            category=cls.waste_category,
        )

    def setUp(self):
        clear_derived_value_config_cache()

    def tearDown(self):
        clear_derived_value_config_cache()

    def _create_collection(self, suffix, *, population=None):
        region = Region.objects.create(name=f"Region {suffix}", country="DE")
        catchment = CollectionCatchment.objects.create(
            name=f"Catchment {suffix}",
            region=region,
        )
        collection = Collection.objects.create(
            catchment=catchment,
            collection_system=self.collection_system,
            waste_stream=self.waste_stream,
            valid_from=date(2024, 1, 1),
            publication_status="published",
        )
        if population is not None:
            RegionAttributeValue.objects.create(
                name=f"Population {suffix}",
                region=region,
                attribute=self.population_attribute,
                date=date(2024, 1, 1),
                value=population,
            )
        return collection, catchment

    @staticmethod
    def _bulk_create_cpv(**kwargs):
        cpv = CollectionPropertyValue(**kwargs)
        CollectionPropertyValue.objects.bulk_create([cpv])
        return cpv

    def test_create_or_update_removes_stale_derived_when_manual_exists(self):
        collection, _catchment = self._create_collection("stale", population=2000)

        specific = self._bulk_create_cpv(
            name="specific source",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10,
            publication_status="published",
            is_derived=False,
        )

        _derived, action = create_or_update_derived_cpv(specific)
        self.assertEqual(action, "created")
        self.assertTrue(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

        self._bulk_create_cpv(
            name="manual total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=25,
            publication_status="published",
            is_derived=False,
        )

        _derived, action = create_or_update_derived_cpv(specific)
        self.assertEqual(action, "skipped")
        self.assertFalse(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

    def test_backfill_counts_created_updated_and_skipped(self):
        collection_with_population, _ = self._create_collection(
            "has-pop", population=1000
        )
        collection_without_population, _ = self._create_collection(
            "no-pop", population=None
        )

        self._bulk_create_cpv(
            name="create-source",
            collection=collection_with_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="update-source",
            collection=collection_with_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2025,
            average=20,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="existing-derived-total",
            collection=collection_with_population,
            property=self.property_total,
            unit=self.unit_total,
            year=2025,
            average=20,
            publication_status="published",
            is_derived=True,
        )
        self._bulk_create_cpv(
            name="skip-manual-specific",
            collection=collection_with_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2026,
            average=30,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="manual-total",
            collection=collection_with_population,
            property=self.property_total,
            unit=self.unit_total,
            year=2026,
            average=30,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="skip-no-pop",
            collection=collection_without_population,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=40,
            publication_status="published",
            is_derived=False,
        )

        dry_stats = backfill_derived_values(dry_run=True)
        self.assertEqual(dry_stats, {"created": 1, "updated": 1, "skipped": 3})

        write_stats = backfill_derived_values(dry_run=False)
        self.assertEqual(write_stats, {"created": 1, "updated": 1, "skipped": 3})

    def test_amounts_for_2024_falls_back_to_total_and_population(self):
        collection, catchment = self._create_collection(
            "atlas-fallback", population=2500
        )

        self._bulk_create_cpv(
            name="total-only",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=6.5,
            publication_status="published",
            is_derived=False,
        )

        amounts = _amounts_for_2024(
            year=2024,
            all_collection_ids={collection.pk},
            col_to_cid={collection.pk: catchment.pk},
            catchment_ids=[catchment.pk],
        )
        self.assertEqual(amounts[catchment.pk], round(6.5 * 1000 / 2500, 1))

    def test_compute_counterpart_value_returns_none_for_non_convertible_property(self):
        other_property = Property.objects.create(name="other property [test]")
        collection, _ = self._create_collection("other", population=1000)
        cpv = self._bulk_create_cpv(
            name="other-cpv",
            collection=collection,
            property=other_property,
            unit=self.unit_specific,
            year=2024,
            average=1.0,
            publication_status="published",
            is_derived=False,
        )
        self.assertIsNone(compute_counterpart_value(cpv))

    def test_compute_counterpart_value_returns_none_without_population(self):
        collection, _ = self._create_collection("no-pop-compute", population=None)
        cpv = self._bulk_create_cpv(
            name="specific-no-pop",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=12.3,
            publication_status="published",
            is_derived=False,
        )
        self.assertIsNone(compute_counterpart_value(cpv))

    def test_create_or_update_skips_when_input_is_already_derived(self):
        collection, _ = self._create_collection("already-derived", population=1000)
        derived_source = self._bulk_create_cpv(
            name="derived-source",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=True,
        )
        derived, action = create_or_update_derived_cpv(derived_source)
        self.assertIsNone(derived)
        self.assertEqual(action, "skipped")

    def test_conversion_helpers_guard_invalid_population(self):
        self.assertIsNone(convert_specific_to_total_mg(10, 0))
        self.assertIsNone(convert_specific_to_total_mg(10, -5))
        self.assertIsNone(convert_total_to_specific(10, 0))
        self.assertIsNone(convert_total_to_specific(10, -5))

    def test_get_derived_property_config_raises_for_invalid_configured_id(self):
        with override_settings(SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=999999):
            clear_derived_value_config_cache()
            with self.assertRaises(ImproperlyConfigured):
                get_derived_property_config()

    def test_get_derived_property_config_raises_for_ambiguous_names(self):
        Property.objects.create(name="specific waste collected [test]")
        with override_settings(SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=None):
            clear_derived_value_config_cache()
            with self.assertRaises(ImproperlyConfigured):
                get_derived_property_config()

    def test_signals_swallow_improperly_configured_and_do_not_raise(self):
        collection, _ = self._create_collection("signal-noise", population=1000)
        cpv = self._bulk_create_cpv(
            name="signal-cpv",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=15.0,
            publication_status="published",
            is_derived=False,
        )

        with override_settings(SOILCOM_SPECIFIC_WASTE_PROPERTY_ID=999999):
            clear_derived_value_config_cache()
            # Should not raise; handler catches ImproperlyConfigured.
            sync_derived_cpv_on_save(sender=CollectionPropertyValue, instance=cpv)
            sync_derived_cpv_on_delete(sender=CollectionPropertyValue, instance=cpv)

    def test_population_attribute_resolution_uses_fallback_when_misconfigured(self):
        with override_settings(SOILCOM_POPULATION_ATTRIBUTE_ID=999999):
            clear_derived_value_config_cache()
            self.assertEqual(
                _resolved_population_attribute_id(), POPULATION_ATTRIBUTE_ID
            )

    def test_convert_specific_to_total_mg_happy_path(self):
        self.assertEqual(convert_specific_to_total_mg(50, 2000), 100.0)
        self.assertEqual(convert_specific_to_total_mg(10, 500), 5.0)

    def test_convert_total_to_specific_happy_path(self):
        self.assertEqual(convert_total_to_specific(100, 2000), 50.0)
        self.assertEqual(convert_total_to_specific(5, 500), 10.0)

    def test_conversion_helpers_ndigits_none_returns_exact_float(self):
        result = convert_specific_to_total_mg(7, 3000, ndigits=None)
        self.assertEqual(result, 7 * 3000 / 1000)
        result = convert_total_to_specific(7, 3000, ndigits=None)
        self.assertEqual(result, 7 * 1000 / 3000)

    def test_conversion_helpers_respect_ndigits(self):
        self.assertEqual(convert_specific_to_total_mg(7, 3000, ndigits=1), 21.0)
        self.assertEqual(convert_total_to_specific(6.5, 2500, ndigits=1), 2.6)

    def test_compute_counterpart_specific_to_total(self):
        collection, _ = self._create_collection("s2t", population=5000)
        cpv = self._bulk_create_cpv(
            name="specific-s2t",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=80.0,
            publication_status="published",
            is_derived=False,
        )
        result = compute_counterpart_value(cpv)
        self.assertIsNotNone(result)
        target_prop_id, target_unit_id, computed = result
        self.assertEqual(target_prop_id, self.property_total.pk)
        self.assertEqual(target_unit_id, self.unit_total.pk)
        self.assertEqual(computed, 400.0)

    def test_compute_counterpart_total_to_specific(self):
        collection, _ = self._create_collection("t2s", population=4000)
        cpv = self._bulk_create_cpv(
            name="total-t2s",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=120.0,
            publication_status="published",
            is_derived=False,
        )
        result = compute_counterpart_value(cpv)
        self.assertIsNotNone(result)
        target_prop_id, target_unit_id, computed = result
        self.assertEqual(target_prop_id, self.property_specific.pk)
        self.assertEqual(target_unit_id, self.unit_specific.pk)
        self.assertEqual(computed, 30.0)

    def test_create_derived_total_stores_correct_average(self):
        collection, _ = self._create_collection("val-s2t", population=2000)
        cpv = self._bulk_create_cpv(
            name="source-specific",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        derived, action = create_or_update_derived_cpv(cpv)
        self.assertEqual(action, "created")
        self.assertIsNotNone(derived)
        self.assertEqual(derived.average, 20.0)
        self.assertEqual(derived.property_id, self.property_total.pk)
        self.assertEqual(derived.unit_id, self.unit_total.pk)
        self.assertTrue(derived.is_derived)

    def test_create_derived_specific_from_total_stores_correct_average(self):
        collection, _ = self._create_collection("val-t2s", population=5000)
        cpv = self._bulk_create_cpv(
            name="source-total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=250.0,
            publication_status="published",
            is_derived=False,
        )
        derived, action = create_or_update_derived_cpv(cpv)
        self.assertEqual(action, "created")
        self.assertIsNotNone(derived)
        self.assertEqual(derived.average, 50.0)
        self.assertEqual(derived.property_id, self.property_specific.pk)
        self.assertEqual(derived.unit_id, self.unit_specific.pk)
        self.assertTrue(derived.is_derived)

    def test_create_or_update_updates_existing_derived_value(self):
        collection, _ = self._create_collection("update-val", population=1000)
        cpv = self._bulk_create_cpv(
            name="source-update",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        derived1, action1 = create_or_update_derived_cpv(cpv)
        self.assertEqual(action1, "created")
        self.assertEqual(derived1.average, 10.0)

        cpv.average = 20.0
        cpv.save()
        derived2, action2 = create_or_update_derived_cpv(cpv)
        self.assertEqual(action2, "updated")
        self.assertEqual(derived2.average, 20.0)
        self.assertEqual(derived1.pk, derived2.pk)

    def test_delete_derived_cpv_removes_counterpart(self):
        collection, _ = self._create_collection("del", population=1000)
        source = self._bulk_create_cpv(
            name="source-del",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        create_or_update_derived_cpv(source)
        self.assertTrue(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

        count = delete_derived_cpv(source)
        self.assertEqual(count, 1)
        self.assertFalse(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=True,
            ).exists()
        )

    def test_delete_derived_cpv_does_not_remove_manual_counterpart(self):
        collection, _ = self._create_collection("del-manual", population=1000)
        source = self._bulk_create_cpv(
            name="source-del-m",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="manual-total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2024,
            average=99.0,
            publication_status="published",
            is_derived=False,
        )

        count = delete_derived_cpv(source)
        self.assertEqual(count, 0)
        self.assertTrue(
            CollectionPropertyValue.objects.filter(
                collection=collection,
                property=self.property_total,
                year=2024,
                is_derived=False,
            ).exists()
        )

    def test_delete_derived_cpv_skips_when_source_is_derived(self):
        collection, _ = self._create_collection("del-skip", population=1000)
        derived_source = self._bulk_create_cpv(
            name="derived-del",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=True,
        )
        self.assertEqual(delete_derived_cpv(derived_source), 0)

    def test_backfill_dry_run_does_not_write(self):
        collection, _ = self._create_collection("dryrun", population=1000)
        self._bulk_create_cpv(
            name="dry-source",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=10.0,
            publication_status="published",
            is_derived=False,
        )
        derived_before = CollectionPropertyValue.objects.filter(is_derived=True).count()
        backfill_derived_values(dry_run=True)
        derived_after = CollectionPropertyValue.objects.filter(is_derived=True).count()
        self.assertEqual(derived_before, derived_after)

    def test_backfill_writes_correct_values(self):
        collection, _ = self._create_collection("bf-val", population=2000)
        self._bulk_create_cpv(
            name="bf-specific",
            collection=collection,
            property=self.property_specific,
            unit=self.unit_specific,
            year=2024,
            average=60.0,
            publication_status="published",
            is_derived=False,
        )
        self._bulk_create_cpv(
            name="bf-total",
            collection=collection,
            property=self.property_total,
            unit=self.unit_total,
            year=2025,
            average=50.0,
            publication_status="published",
            is_derived=False,
        )

        backfill_derived_values(dry_run=False)

        derived_total = CollectionPropertyValue.objects.get(
            collection=collection,
            property=self.property_total,
            year=2024,
            is_derived=True,
        )
        self.assertEqual(derived_total.average, 120.0)

        derived_specific = CollectionPropertyValue.objects.get(
            collection=collection,
            property=self.property_specific,
            year=2025,
            is_derived=True,
        )
        self.assertEqual(derived_specific.average, 25.0)

    def test_get_population_for_collection_returns_exact_year(self):
        collection, _ = self._create_collection("pop-year", population=None)
        region = collection.catchment.region
        RegionAttributeValue.objects.create(
            name="Pop 2023",
            region=region,
            attribute=self.population_attribute,
            date=date(2023, 1, 1),
            value=3000,
        )
        RegionAttributeValue.objects.create(
            name="Pop 2024",
            region=region,
            attribute=self.population_attribute,
            date=date(2024, 6, 15),
            value=3500,
        )
        self.assertEqual(get_population_for_collection(collection, year=2024), 3500)
        self.assertEqual(get_population_for_collection(collection, year=2023), 3000)

    def test_get_population_for_collection_falls_back_to_most_recent(self):
        collection, _ = self._create_collection("pop-fallback", population=None)
        region = collection.catchment.region
        RegionAttributeValue.objects.create(
            name="Pop old",
            region=region,
            attribute=self.population_attribute,
            date=date(2020, 1, 1),
            value=1000,
        )
        RegionAttributeValue.objects.create(
            name="Pop newer",
            region=region,
            attribute=self.population_attribute,
            date=date(2022, 1, 1),
            value=2000,
        )
        self.assertEqual(get_population_for_collection(collection, year=2024), 2000)

    def test_get_population_for_collection_returns_none_without_data(self):
        collection, _ = self._create_collection("pop-none", population=None)
        self.assertIsNone(get_population_for_collection(collection, year=2024))


def dict_to_querydict(data):
    """Convert a dict to QueryDict for form testing."""
    from django.http import QueryDict

    qd = QueryDict(mutable=True)
    for key, value in data.items():
        if isinstance(value, list):
            qd.setlist(key, value)
        else:
            qd[key] = value
    return qd


class DeleteUnusedWasteStreamsSignalTestCase(TestCase):
    """Tests for the delete_unused_waste_streams signal handler."""

    @classmethod
    def setUpTestData(cls):
        cls.catchment = CollectionCatchment.objects.create(
            name="Catchment", publication_status="published"
        )
        cls.collector = Collector.objects.create(
            name="Collector", publication_status="published"
        )
        cls.system = CollectionSystem.objects.create(
            name="System", publication_status="published"
        )
        cls.category = WasteCategory.objects.create(
            name="Category", publication_status="published"
        )
        cls.waste_stream = WasteStream.objects.create(
            name="Used Stream", category=cls.category
        )
        cls.collection = Collection.objects.create(
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.system,
            waste_stream=cls.waste_stream,
            valid_from=date(2024, 1, 1),
        )

    def test_orphan_waste_streams_are_deleted_on_waste_stream_change(self):
        """Verify unused WasteStreams are cleaned up when waste_stream changes."""
        orphan_stream = WasteStream.objects.create(
            name="Orphan Stream", category=self.category
        )
        new_stream = WasteStream.objects.create(
            name="New Stream", category=self.category
        )

        with patch(
            "case_studies.soilcom.models.celery.current_app.send_task"
        ) as mock_send_task:
            self.collection.waste_stream = new_stream
            self.collection.save(update_fields=["waste_stream"])

        mock_send_task.assert_called_once_with("cleanup_orphaned_waste_streams")
        cleanup_orphaned_waste_streams()

        self.assertFalse(WasteStream.objects.filter(pk=orphan_stream.pk).exists())
        self.assertFalse(WasteStream.objects.filter(pk=self.waste_stream.pk).exists())
        self.assertTrue(WasteStream.objects.filter(pk=new_stream.pk).exists())

    def test_orphan_cleanup_not_scheduled_when_waste_stream_unchanged(self):
        """Verify cleanup is not scheduled when waste_stream is unchanged."""
        orphan_stream = WasteStream.objects.create(
            name="Orphan Stream 2", category=self.category
        )
        self.assertTrue(WasteStream.objects.filter(pk=orphan_stream.pk).exists())

        with patch(
            "case_studies.soilcom.models.celery.current_app.send_task"
        ) as mock_send_task:
            self.collection.description = "Updated"
            self.collection.save()

        mock_send_task.assert_not_called()
        self.assertTrue(WasteStream.objects.filter(pk=orphan_stream.pk).exists())

    def test_orphan_cleanup_not_scheduled_with_update_fields(self):
        """Verify cleanup is not scheduled when update_fields excludes waste_stream."""
        orphan_stream = WasteStream.objects.create(
            name="Orphan Stream 3", category=self.category
        )
        self.assertTrue(WasteStream.objects.filter(pk=orphan_stream.pk).exists())

        with patch(
            "case_studies.soilcom.models.celery.current_app.send_task"
        ) as mock_send_task:
            self.collection.valid_until = date(2024, 12, 31)
            self.collection.save(update_fields=["valid_until"])

        mock_send_task.assert_not_called()
        self.assertTrue(WasteStream.objects.filter(pk=orphan_stream.pk).exists())


class InvalidateCollectionCacheSignalTestCase(TestCase):
    """Tests for the invalidate_collection_geojson_cache signal handler."""

    @classmethod
    def setUpTestData(cls):
        cls.catchment = CollectionCatchment.objects.create(
            name="Catchment", publication_status="published"
        )
        cls.collector = Collector.objects.create(
            name="Collector", publication_status="published"
        )
        cls.system = CollectionSystem.objects.create(
            name="System", publication_status="published"
        )
        cls.category = WasteCategory.objects.create(
            name="Category", publication_status="published"
        )
        cls.waste_stream = WasteStream.objects.create(
            name="Stream", category=cls.category
        )
        cls.private_collection = Collection.objects.create(
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.system,
            waste_stream=cls.waste_stream,
            valid_from=date(2024, 1, 1),
            publication_status="private",
        )
        cls.published_collection = Collection.objects.create(
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.system,
            waste_stream=cls.waste_stream,
            valid_from=date(2024, 2, 1),
            publication_status="published",
        )

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    @patch("case_studies.soilcom.signals._schedule_cache_warmup")
    def test_cache_not_cleared_on_private_full_collection_save(
        self, mock_warmup, mock_clear
    ):
        """Private-only saves should not invalidate published GeoJSON cache."""
        self.private_collection.description = "Updated description"
        self.private_collection.save()

        mock_clear.assert_not_called()
        mock_warmup.assert_not_called()

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    @patch("case_studies.soilcom.signals._schedule_cache_warmup")
    def test_cache_cleared_on_published_full_collection_save(
        self, mock_warmup, mock_clear
    ):
        """Published saves must invalidate and warm the collection GeoJSON cache."""
        self.published_collection.description = "Updated description"
        self.published_collection.save()

        mock_clear.assert_called_once_with("collection_geojson:*")
        mock_warmup.assert_called_once()

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    @patch("case_studies.soilcom.signals._schedule_cache_warmup")
    def test_cache_cleared_when_status_changes_from_published_to_private(
        self, mock_warmup, mock_clear
    ):
        """Transition away from published must invalidate published cache entries."""
        self.published_collection.publication_status = "private"
        self.published_collection.save(update_fields=["publication_status"])

        mock_clear.assert_called_once_with("collection_geojson:*")
        mock_warmup.assert_called_once()

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    @patch("case_studies.soilcom.signals._schedule_cache_warmup")
    def test_cache_not_cleared_on_private_save_with_geojson_affecting_field(
        self, mock_warmup, mock_clear
    ):
        """Private GeoJSON-affecting updates should not flush global published cache."""
        self.private_collection.description = "Updated description"
        self.private_collection.save(update_fields=["description"])

        mock_clear.assert_not_called()
        mock_warmup.assert_not_called()

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    def test_cache_not_cleared_on_valid_until_update(self, mock_clear):
        """Verify cache is NOT cleared for valid_until updates."""
        self.published_collection.valid_until = date(2024, 12, 31)
        self.published_collection.save(update_fields=["valid_until"])

        mock_clear.assert_not_called()

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    def test_cache_not_cleared_on_name_only_update(self, mock_clear):
        """Verify cache is NOT cleared for name-only updates."""
        self.published_collection.name = "New Name"
        self.published_collection.save(update_fields=["name"])

        mock_clear.assert_not_called()


class UpdateCollectionNamesSignalTestCase(TestCase):
    """Tests for the update_collection_names signal handler."""

    @classmethod
    def setUpTestData(cls):
        cls.catchment = CollectionCatchment.objects.create(
            name="Catchment", publication_status="published"
        )
        cls.collector = Collector.objects.create(
            name="Collector", publication_status="published"
        )
        cls.system = CollectionSystem.objects.create(
            name="System", publication_status="published"
        )
        cls.category = WasteCategory.objects.create(
            name="Category", publication_status="published"
        )
        cls.waste_stream = WasteStream.objects.create(
            name="Stream", category=cls.category
        )
        cls.collection = Collection.objects.create(
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.system,
            waste_stream=cls.waste_stream,
            valid_from=date(2024, 1, 1),
        )

    def test_collection_name_updated_when_system_changes(self):
        """Verify Collection name is updated when CollectionSystem name changes."""
        original_name = self.collection.name
        self.system.name = "New System"
        self.system.save()

        self.collection.refresh_from_db()
        self.assertNotEqual(self.collection.name, original_name)
        self.assertIn("New System", self.collection.name)

    def test_collection_name_updated_when_catchment_changes(self):
        """Verify Collection name is updated when Catchment name changes."""
        original_name = self.collection.name
        self.catchment.name = "New Catchment"
        self.catchment.save()

        self.collection.refresh_from_db()
        self.assertNotEqual(self.collection.name, original_name)
        self.assertIn("New Catchment", self.collection.name)

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    def test_name_update_does_not_trigger_cache_invalidation(self, mock_clear):
        """Verify name updates via update_collection_names don't trigger cache clear."""
        Collection.objects.create(
            catchment=self.catchment,
            collector=self.collector,
            collection_system=self.system,
            waste_stream=self.waste_stream,
            valid_from=date(2023, 1, 1),
        )
        Collection.objects.create(
            catchment=self.catchment,
            collector=self.collector,
            collection_system=self.system,
            waste_stream=self.waste_stream,
            valid_from=date(2022, 1, 1),
        )

        mock_clear.reset_mock()

        self.system.name = "Updated System"
        self.system.save()

        self.assertEqual(mock_clear.call_count, 0)


class CollectionFormPredecessorSaveTestCase(TestCase):
    """Tests for predecessor handling in CollectionModelForm.save()."""

    @classmethod
    def setUpTestData(cls):
        MaterialCategory.objects.get_or_create(
            name="Biowaste component", publication_status="published"
        )
        cls.catchment = CollectionCatchment.objects.create(
            name="Catchment", publication_status="published"
        )
        cls.collector = Collector.objects.create(
            name="Collector", publication_status="published"
        )
        cls.system = CollectionSystem.objects.create(
            name="System", publication_status="published"
        )
        cls.category = WasteCategory.objects.create(
            name="Category", publication_status="published"
        )
        cls.allowed_material = WasteComponent.objects.create(
            name="Allowed", publication_status="published"
        )
        biowaste_cat = MaterialCategory.objects.get(name="Biowaste component")
        cls.allowed_material.categories.add(biowaste_cat)
        cls.waste_stream = WasteStream.objects.create(
            name="Stream", category=cls.category
        )
        cls.waste_stream.allowed_materials.add(cls.allowed_material)
        cls.frequency = CollectionFrequency.objects.create(
            name="Frequency", publication_status="published"
        )
        cls.predecessor = Collection.objects.create(
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.system,
            waste_stream=cls.waste_stream,
            valid_from=date(2023, 1, 1),
            publication_status="published",
        )
        cls.collection = Collection.objects.create(
            catchment=cls.catchment,
            collector=cls.collector,
            collection_system=cls.system,
            waste_stream=cls.waste_stream,
            valid_from=date(2024, 1, 1),
            publication_status="published",
        )
        cls.collection.predecessors.add(cls.predecessor)

    def test_predecessor_valid_until_updated_on_form_save(self):
        """Verify predecessors' valid_until is updated when form is saved."""
        self.assertIsNone(self.predecessor.valid_until)

        form_data = {
            "catchment": self.catchment.id,
            "collector": self.collector.id,
            "collection_system": self.system.id,
            "waste_category": self.category.id,
            "allowed_materials": [self.allowed_material.id],
            "forbidden_materials": [],
            "frequency": self.frequency.id,
            "valid_from": date(2024, 6, 1),
            "connection_type": "VOLUNTARY",
        }
        form = CollectionModelForm(
            instance=self.collection, data=dict_to_querydict(form_data)
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        self.predecessor.refresh_from_db()
        self.assertEqual(self.predecessor.valid_until, date(2024, 5, 31))

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    def test_predecessor_save_minimizes_cache_invalidation(self, mock_clear):
        """Verify predecessor updates minimize cache invalidation calls."""
        predecessor2 = Collection.objects.create(
            catchment=self.catchment,
            collector=self.collector,
            collection_system=self.system,
            waste_stream=self.waste_stream,
            valid_from=date(2022, 1, 1),
            publication_status="published",
        )
        self.collection.predecessors.add(predecessor2)

        mock_clear.reset_mock()

        form_data = {
            "catchment": self.catchment.id,
            "collector": self.collector.id,
            "collection_system": self.system.id,
            "waste_category": self.category.id,
            "allowed_materials": [self.allowed_material.id],
            "forbidden_materials": [],
            "frequency": self.frequency.id,
            "valid_from": date(2024, 6, 1),
            "connection_type": "VOLUNTARY",
        }
        form = CollectionModelForm(
            instance=self.collection, data=dict_to_querydict(form_data)
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        self.assertLessEqual(mock_clear.call_count, 1)


class WasteStreamSaveOptimizationTestCase(TestCase):
    """Tests for WasteStream save optimization in CollectionModelForm."""

    @classmethod
    def setUpTestData(cls):
        MaterialCategory.objects.get_or_create(
            name="Biowaste component", publication_status="published"
        )
        cls.catchment = CollectionCatchment.objects.create(
            name="Catchment", publication_status="published"
        )
        cls.collector = Collector.objects.create(
            name="Collector", publication_status="published"
        )
        cls.system = CollectionSystem.objects.create(
            name="System", publication_status="published"
        )
        cls.category = WasteCategory.objects.create(
            name="Category", publication_status="published"
        )
        cls.allowed_material = WasteComponent.objects.create(
            name="Allowed", publication_status="published"
        )
        biowaste_cat = MaterialCategory.objects.get(name="Biowaste component")
        cls.allowed_material.categories.add(biowaste_cat)
        cls.frequency = CollectionFrequency.objects.create(
            name="Frequency", publication_status="published"
        )

    def test_new_waste_stream_created_when_needed(self):
        """Verify a new WasteStream is created when no matching one exists."""
        initial_count = WasteStream.objects.count()

        form_data = {
            "catchment": self.catchment.id,
            "collector": self.collector.id,
            "collection_system": self.system.id,
            "waste_category": self.category.id,
            "allowed_materials": [self.allowed_material.id],
            "forbidden_materials": [],
            "frequency": self.frequency.id,
            "valid_from": date(2024, 1, 1),
            "connection_type": "VOLUNTARY",
        }
        form = CollectionModelForm(data=dict_to_querydict(form_data))
        self.assertTrue(form.is_valid(), form.errors)
        form.instance.owner_id = 1
        instance = form.save()

        self.assertEqual(WasteStream.objects.count(), initial_count + 1)
        self.assertIsNotNone(instance.waste_stream)

    def test_existing_waste_stream_reused(self):
        """Verify existing WasteStream is reused when a matching one exists."""
        existing_stream = WasteStream.objects.create(
            name="Existing", category=self.category
        )
        existing_stream.allowed_materials.add(self.allowed_material)

        initial_count = WasteStream.objects.count()

        form_data = {
            "catchment": self.catchment.id,
            "collector": self.collector.id,
            "collection_system": self.system.id,
            "waste_category": self.category.id,
            "allowed_materials": [self.allowed_material.id],
            "forbidden_materials": [],
            "frequency": self.frequency.id,
            "valid_from": date(2024, 1, 1),
            "connection_type": "VOLUNTARY",
        }
        form = CollectionModelForm(data=dict_to_querydict(form_data))
        self.assertTrue(form.is_valid(), form.errors)
        form.instance.owner_id = 1
        instance = form.save()

        self.assertEqual(WasteStream.objects.count(), initial_count)
        self.assertEqual(instance.waste_stream.pk, existing_stream.pk)

    @patch("case_studies.soilcom.signals.clear_geojson_cache_pattern")
    def test_reusing_waste_stream_minimizes_cache_clears(self, mock_clear):
        """Verify reusing WasteStream minimizes cache invalidation."""
        existing_stream = WasteStream.objects.create(
            name="Existing2", category=self.category
        )
        existing_stream.allowed_materials.add(self.allowed_material)

        Collection.objects.create(
            catchment=self.catchment,
            collector=self.collector,
            collection_system=self.system,
            waste_stream=existing_stream,
            valid_from=date(2023, 1, 1),
        )

        mock_clear.reset_mock()

        form_data = {
            "catchment": self.catchment.id,
            "collector": self.collector.id,
            "collection_system": self.system.id,
            "waste_category": self.category.id,
            "allowed_materials": [self.allowed_material.id],
            "forbidden_materials": [],
            "frequency": self.frequency.id,
            "valid_from": date(2024, 1, 1),
            "connection_type": "VOLUNTARY",
        }
        form = CollectionModelForm(data=dict_to_querydict(form_data))
        self.assertTrue(form.is_valid(), form.errors)
        form.instance.owner_id = 1
        form.save()

        self.assertLessEqual(mock_clear.call_count, 2)


class SortingMethodModelTestCase(TestCase):
    """Unit tests for the SortingMethod model."""

    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(username="sm-test-owner")

    def test_str_returns_name(self):
        method = SortingMethod(name="Separate bins")
        self.assertEqual(str(method), "Separate bins")

    def test_create_and_retrieve(self):
        method = SortingMethod.objects.create(
            name="Optical bag sorting",
            owner=self.owner,
            publication_status="private",
        )
        retrieved = SortingMethod.objects.get(pk=method.pk)
        self.assertEqual(retrieved.name, "Optical bag sorting")

    def test_description_optional(self):
        method = SortingMethod.objects.create(
            name="Two compartments bin",
            owner=self.owner,
            publication_status="private",
        )
        self.assertIsNone(method.description)


class CollectionSortingMethodFieldTestCase(TestCase):
    """Tests for Collection.sorting_method FK behaviour."""

    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(username="col-sm-owner")
        cls.catchment = CollectionCatchment.objects.create(name="SM Field Catchment")
        cls.sorting_method = SortingMethod.objects.create(
            name="Four compartments bin",
            owner=cls.owner,
            publication_status="private",
        )

    def test_collection_stores_sorting_method(self):
        col = Collection.objects.create(
            catchment=self.catchment,
            sorting_method=self.sorting_method,
        )
        col.refresh_from_db()
        self.assertEqual(col.sorting_method_id, self.sorting_method.pk)

    def test_sorting_method_null_by_default(self):
        col = Collection.objects.create(catchment=self.catchment)
        self.assertIsNone(col.sorting_method)

    def test_delete_sorting_method_sets_null_on_collection(self):
        method = SortingMethod.objects.create(
            name="Temporary method",
            owner=self.owner,
            publication_status="private",
        )
        col = Collection.objects.create(
            catchment=self.catchment,
            sorting_method=method,
        )
        method.delete()
        col.refresh_from_db()
        self.assertIsNone(col.sorting_method)


class CollectionEstablishedFieldTestCase(TestCase):
    """Tests for Collection.established field."""

    @classmethod
    def setUpTestData(cls):
        cls.catchment = CollectionCatchment.objects.create(name="Est Field Catchment")

    def test_established_stores_year(self):
        col = Collection.objects.create(
            catchment=self.catchment,
            established=2005,
        )
        col.refresh_from_db()
        self.assertEqual(col.established, 2005)

    def test_established_null_by_default(self):
        col = Collection.objects.create(catchment=self.catchment)
        self.assertIsNone(col.established)


class CollectionImporterSortingMethodTestCase(TestCase):
    """Integration tests for sorting_method and established in CollectionImporter."""

    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(username="imp-sm-owner")
        cls.catchment = CollectionCatchment.objects.create(name="Importer SM Catchment")
        cls.collection_system = CollectionSystem.objects.create(
            name="Door to door", owner=cls.owner
        )
        cls.waste_category = WasteCategory.objects.create(name="Food waste")
        cls.sorting_method = SortingMethod.objects.create(
            name="Separate bins",
            owner=cls.owner,
            publication_status="private",
        )
        cls.valid_from = date(2021, 1, 1)

    def _make_record(self, **overrides):
        base = {
            "nuts_or_lau_id": None,
            "catchment_name": self.catchment.name,
            "collection_system": self.collection_system.name,
            "waste_category": self.waste_category.name,
            "sorting_method": self.sorting_method.name,
            "established": 2015,
            "valid_from": self.valid_from,
            "valid_until": None,
            "collector_name": None,
            "fee_system": None,
            "frequency": None,
            "connection_type": None,
            "min_bin_size": None,
            "required_bin_capacity": None,
            "required_bin_capacity_reference": None,
            "allowed_materials": "",
            "forbidden_materials": "",
            "description": "",
            "property_values": [],
            "flyer_urls": [],
        }
        base.update(overrides)
        return base

    def test_import_sets_sorting_method(self):
        """Importer persists sorting_method on newly created collection."""
        importer = CollectionImporter(owner=self.owner, publication_status="private")
        stats = importer.run([self._make_record()])
        self.assertEqual(stats["created"], 1)
        self.assertEqual(stats["warnings"], [])

        col = Collection.objects.get(
            owner=self.owner,
            valid_from=self.valid_from,
            collection_system=self.collection_system,
        )
        self.assertEqual(col.sorting_method, self.sorting_method)
        col.delete()

    def test_import_sets_established(self):
        """Importer persists established year on newly created collection."""
        importer = CollectionImporter(owner=self.owner, publication_status="private")
        stats = importer.run([self._make_record()])
        self.assertEqual(stats["created"], 1)

        col = Collection.objects.get(
            owner=self.owner,
            valid_from=self.valid_from,
            collection_system=self.collection_system,
        )
        self.assertEqual(col.established, 2015)
        col.delete()

    def test_unknown_sorting_method_adds_warning_and_leaves_field_empty(self):
        """Unknown sorting_method name produces a warning and does not block import."""
        importer = CollectionImporter(owner=self.owner, publication_status="private")
        stats = importer.run([self._make_record(sorting_method="Nonexistent Method")])
        self.assertEqual(stats["created"], 1)
        self.assertTrue(
            any(
                "SortingMethod" in w and "Nonexistent Method" in w
                for w in stats["warnings"]
            ),
            msg=f"Expected SortingMethod warning, got: {stats['warnings']}",
        )

        col = Collection.objects.get(
            owner=self.owner,
            valid_from=self.valid_from,
            collection_system=self.collection_system,
        )
        self.assertIsNone(col.sorting_method)
        col.delete()

    def test_empty_sorting_method_skips_resolution_silently(self):
        """Empty sorting_method string does not produce a warning."""
        importer = CollectionImporter(owner=self.owner, publication_status="private")
        stats = importer.run([self._make_record(sorting_method="")])
        self.assertEqual(stats["created"], 1)
        self.assertFalse(
            any("SortingMethod" in w for w in stats["warnings"]),
            msg="Unexpected SortingMethod warning for empty value",
        )

        col = Collection.objects.get(
            owner=self.owner,
            valid_from=self.valid_from,
            collection_system=self.collection_system,
        )
        self.assertIsNone(col.sorting_method)
        col.delete()

    def test_import_updates_null_sorting_method_on_existing_collection(self):
        """Re-importing updates sorting_method when it was previously null."""
        valid_from = date(2021, 3, 1)

        importer = CollectionImporter(owner=self.owner, publication_status="private")
        importer.run([self._make_record(sorting_method="", valid_from=valid_from)])
        col = Collection.objects.get(
            owner=self.owner,
            valid_from=valid_from,
            collection_system=self.collection_system,
        )
        self.assertIsNone(col.sorting_method)

        importer2 = CollectionImporter(owner=self.owner, publication_status="private")
        importer2.run([self._make_record(valid_from=valid_from)])
        col.refresh_from_db()
        self.assertEqual(col.sorting_method, self.sorting_method)
        col.delete()

    def test_import_updates_null_established_on_existing_collection(self):
        """Re-importing updates established when it was previously null."""
        importer = CollectionImporter(owner=self.owner, publication_status="private")
        importer.run([self._make_record(established=None)])
        col = Collection.objects.get(
            owner=self.owner,
            valid_from=self.valid_from,
            collection_system=self.collection_system,
        )
        self.assertIsNone(col.established)

        importer2 = CollectionImporter(owner=self.owner, publication_status="private")
        importer2.run([self._make_record(established=2010)])
        col.refresh_from_db()
        self.assertEqual(col.established, 2010)
        col.delete()
