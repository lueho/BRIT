"""Management command to import Swedish waste collection data from Avfall Sverige reports.

Runs **locally** (outside the container) against a running BRIT instance via the
bulk-import API endpoint.  Parses the source files on the local machine — including
PDFs via pdfplumber — and POSTs the resulting records in batches.

Sources (local paths, configurable via CLI flags):
- 2021: Excel file  Sweden_Waste statistics2021.xlsx
- 2022: PDF         husha-llsavfall-i-siffror-2022.pdf
- 2023: PDF         husha-llsavfall-i-siffror-2023.pdf

Usage::

    python manage.py import_sweden_collections --api-url https://brit.example.com --token <TOKEN>
    python manage.py import_sweden_collections --api-url http://localhost:8000 \\
        --username staff --password secret --dry-run
    python manage.py import_sweden_collections --api-url http://localhost:8000 --token <TOKEN> \\
        --year 2022 --pdf-2022 /path/to/2022.pdf
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path
from urllib.request import Request, urlopen

import openpyxl
from django.core.management.base import BaseCommand, CommandError

# ---------------------------------------------------------------------------
# Default file paths (local, relative to cwd)
# ---------------------------------------------------------------------------
_EXCEL_FILE = Path("Sweden_Waste statistics2021.xlsx")
_PDF_2022 = Path("husha-llsavfall-i-siffror-2022.pdf")
_PDF_2023 = Path("husha-llsavfall-i-siffror-2023.pdf")

_BATCH_SIZE = 50
_VALID_STATUSES = ("private", "review")

# ---------------------------------------------------------------------------
# Property / unit constants (verified against DB)
# ---------------------------------------------------------------------------
_PROP_SPECIFIC = 1  # "specific waste collected"
_PROP_CONN_RATE = 4  # "Connection rate"
_UNIT_KG = "kg/(cap.*a)"
_UNIT_PCT_HH = "% of households"

# ---------------------------------------------------------------------------
# BRIT canonical names
# ---------------------------------------------------------------------------
_COLLECTION_SYSTEM_DOOR_TO_DOOR = "Door to door"
_COLLECTION_SYSTEM_NONE = "No separate collection"
_WASTE_CATEGORY_FOOD = "Food waste"
_WASTE_CATEGORY_RESIDUAL = "Residual waste"

# ---------------------------------------------------------------------------
# Value translation maps
# ---------------------------------------------------------------------------

# Swedish bag type → BRIT Material name
_BAG_MATERIAL_MAP = {
    "Papper": "Collection Support Item: Paper bags",
    "Plastpåse": "Collection Support Item: Plastic bags",
    "Bioplast": "Collection Support Item: Biodegradable plastic bags",
}

# Swedish fee type → BRIT FeeSystem name
_FEE_MAP = {
    "Fixed": "Flat fee",
    "PAYT": "Pay as you throw (PAYT)",
}

# Swedish connection type → BRIT connection_type key
_CONNECTION_MAP = {
    "Obl": "MANDATORY",
    "Friv": "VOLUNTARY",
}

# Swedish sorting system → BRIT SortingMethod name
_SORTING_METHOD_MAP = {
    "Separate bins": "Separate bins",
    "Separata kärl": "Separate bins",
    "Optical bag sorting": "Optical bag sorting",
    "Optisk sortering": "Optical bag sorting",
    "Four compartments bin": "Four compartments bin",
    "Fyrfackskärl": "Four compartments bin",
    "Two compartments bin": "Two compartments bin",
    "Tvådelade kärl": "Two compartments bin",
}

# System values that mean "no separate collection"
_NO_COLLECTION_SYSTEMS = {"Non", "Ingen utsortering", "-"}

# Non-standard municipalities → BRIT CollectionCatchment name
_CUSTOM_CATCHMENT_MAP = {
    "Motala/Vadstena": "Catchment of Motala-Vadstena vatten- och avfallsnämnd",
    "Kretslopp Sydost": "Catchment of Kretslopp Sydost",
    "Landskrona-Svalöv (LSR)": "Catchment of Landskrona–Svalövs Renhållnings AB (LSR)",
    "Avfall & Återvinning Skaraborg": "Catchment of Avfall & Återvinning Skaraborg (A&ÅS)",
    "VafabMiljö": "Catchment of Vafabmiljö kommunalförbund",
    "Gästrike Återvinnare": "Catchment of Gästrike Återvinnare",
}

# Swedish municipality name → 4-digit LAU ID (from BRIT CollectionCatchment DB)
_LAU_IDS = {
    "Ale": "1440",
    "Alingsås": "1489",
    "Alvesta": "0764",
    "Aneby": "0604",
    "Arboga": "1984",
    "Arjeplog": "2506",
    "Arvidsjaur": "2505",
    "Arvika": "1784",
    "Askersund": "1882",
    "Avesta": "2084",
    "Bengtsfors": "1460",
    "Berg": "2326",
    "Bjurholm": "2403",
    "Bjuv": "1260",
    "Boden": "2582",
    "Bollebygd": "1443",
    "Bollnäs": "2183",
    "Borgholm": "0885",
    "Borlänge": "2081",
    "Borås": "1490",
    "Botkyrka": "0127",
    "Boxholm": "0560",
    "Bromölla": "1272",
    "Bräcke": "2305",
    "Burlöv": "1231",
    "Båstad": "1278",
    "Dals-Ed": "1438",
    "Danderyd": "0162",
    "Degerfors": "1862",
    "Dorotea": "2425",
    "Eda": "1730",
    "Ekerö": "0125",
    "Eksjö": "0686",
    "Emmaboda": "0862",
    "Enköping": "0381",
    "Eskilstuna": "0484",
    "Eslöv": "1285",
    "Essunga": "1445",
    "Fagersta": "1982",
    "Falkenberg": "1382",
    "Falköping": "1499",
    "Falun": "2080",
    "Filipstad": "1782",
    "Finspång": "0562",
    "Flen": "0482",
    "Forshaga": "1763",
    "Färgelanda": "1439",
    "Gagnef": "2026",
    "Gislaved": "0662",
    "Gnesta": "0461",
    "Gnosjö": "0617",
    "Gotland": "0980",
    "Grums": "1764",
    "Grästorp": "1444",
    "Gullspång": "1447",
    "Gällivare": "2523",
    "Gävle": "2180",
    "Göteborg": "1480",
    "Götene": "1471",
    "Habo": "0643",
    "Hagfors": "1783",
    "Hallsberg": "1861",
    "Hallstahammar": "1961",
    "Halmstad": "1380",
    "Hammarö": "1761",
    "Haninge": "0136",
    "Haparanda": "2583",
    "Heby": "0331",
    "Hedemora": "2083",
    "Helsingborg": "1283",
    "Herrljunga": "1466",
    "Hjo": "1497",
    "Hofors": "2104",
    "Huddinge": "0126",
    "Hudiksvall": "2184",
    "Hultsfred": "0860",
    "Hylte": "1315",
    "Hällefors": "1863",
    "Härjedalen": "2361",
    "Härnösand": "2280",
    "Härryda": "1401",
    "Hässleholm": "1293",
    "Håbo": "0305",
    "Höganäs": "1284",
    "Högsby": "0821",
    "Hörby": "1266",
    "Höör": "1267",
    "Jokkmokk": "2510",
    "Järfälla": "0123",
    "Jönköping": "0680",
    "Kalix": "2514",
    "Kalmar": "0880",
    "Karlsborg": "1446",
    "Karlshamn": "1082",
    "Karlskoga": "1883",
    "Karlskrona": "1080",
    "Karlstad": "1780",
    "Katrineholm": "0483",
    "Kil": "1715",
    "Kinda": "0513",
    "Kiruna": "2584",
    "Klippan": "1276",
    "Knivsta": "0330",
    "Kramfors": "2282",
    "Kristianstad": "1290",
    "Kristinehamn": "1781",
    "Krokom": "2309",
    "Kumla": "1881",
    "Kungsbacka": "1384",
    "Kungsör": "1960",
    "Kungälv": "1482",
    "Kävlinge": "1261",
    "Köping": "1983",
    "Laholm": "1381",
    "Landskrona": "1282",
    "Laxå": "1860",
    "Lekeberg": "1814",
    "Leksand": "2029",
    "Lerum": "1441",
    "Lessebo": "0761",
    "Lidingö": "0186",
    "Lidköping": "1494",
    "Lilla Edet": "1462",
    "Lindesberg": "1885",
    "Linköping": "0580",
    "Ljungby": "0781",
    "Ljusdal": "2161",
    "Ljusnarsberg": "1864",
    "Lomma": "1262",
    "Ludvika": "2085",
    "Luleå": "2580",
    "Lund": "1281",
    "Lycksele": "2481",
    "Lysekil": "1484",
    "Malmö": "1280",
    "Malung-Sälen": "2023",
    "Malå": "2418",
    "Mariestad": "1493",
    "Mark": "1463",
    "Markaryd": "0767",
    "Mellerud": "1461",
    "Mjölby": "0586",
    "Mora": "2062",
    "Motala": "0583",
    "Mullsjö": "0642",
    "Munkedal": "1430",
    "Munkfors": "1762",
    "Mölndal": "1481",
    "Mönsterås": "0861",
    "Mörbylånga": "0840",
    "Nacka": "0182",
    "Nora": "1884",
    "Norberg": "1962",
    "Nordanstig": "2132",
    "Nordmaling": "2401",
    "Norrköping": "0581",
    "Norrtälje": "0188",
    "Norsjö": "2417",
    "Nybro": "0881",
    "Nykvarn": "0140",
    "Nyköping": "0480",
    "Nynäshamn": "0192",
    "Nässjö": "0682",
    "Ockelbo": "2101",
    "Olofström": "1060",
    "Orsa": "2034",
    "Orust": "1421",
    "Osby": "1273",
    "Oskarshamn": "0882",
    "Ovanåker": "2121",
    "Oxelösund": "0481",
    "Pajala": "2521",
    "Partille": "1402",
    "Perstorp": "1275",
    "Piteå": "2581",
    "Ragunda": "2303",
    "Robertsfors": "2409",
    "Ronneby": "1081",
    "Rättvik": "2031",
    "Sala": "1981",
    "Salem": "0128",
    "Sandviken": "2181",
    "Sigtuna": "0191",
    "Simrishamn": "1291",
    "Sjöbo": "1265",
    "Skara": "1495",
    "Skellefteå": "2482",
    "Skinnskatteberg": "1904",
    "Skurup": "1264",
    "Skövde": "1496",
    "Smedjebacken": "2061",
    "Sollefteå": "2283",
    "Sollentuna": "0163",
    "Solna": "0184",
    "Sorsele": "2422",
    "Sotenäs": "1427",
    "Staffanstorp": "1230",
    "Stenungsund": "1415",
    "Stockholm": "0180",
    "Storfors": "1760",
    "Storuman": "2421",
    "Strängnäs": "0486",
    "Strömstad": "1486",
    "Strömsund": "2313",
    "Sundbyberg": "0183",
    "Sundsvall": "2281",
    "Sunne": "1766",
    "Surahammar": "1907",
    "Svalöv": "1214",
    "Svedala": "1263",
    "Svenljunga": "1465",
    "Säffle": "1785",
    "Säter": "2082",
    "Sävsjö": "0684",
    "Söderhamn": "2182",
    "Söderköping": "0582",
    "Södertälje": "0181",
    "Sölvesborg": "1083",
    "Tanum": "1435",
    "Tibro": "1472",
    "Tidaholm": "1498",
    "Tierp": "0360",
    "Timrå": "2262",
    "Tingsryd": "0763",
    "Tjörn": "1419",
    "Tomelilla": "1270",
    "Torsby": "1737",
    "Torsås": "0834",
    "Tranemo": "1452",
    "Tranås": "0687",
    "Trelleborg": "1287",
    "Trollhättan": "1488",
    "Trosa": "0488",
    "Tyresö": "0138",
    "Täby": "0160",
    "Töreboda": "1473",
    "Uddevalla": "1485",
    "Ulricehamn": "1491",
    "Umeå": "2480",
    "Upplands Väsby": "0114",
    "Upplands-Bro": "0139",
    "Uppsala": "0380",
    "Uppvidinge": "0760",
    "Vadstena": "0584",
    "Vaggeryd": "0665",
    "Valdemarsvik": "0563",
    "Vallentuna": "0115",
    "Vansbro": "2021",
    "Vara": "1470",
    "Varberg": "1383",
    "Vaxholm": "0187",
    "Vellinge": "1233",
    "Vetlanda": "0685",
    "Vilhelmina": "2462",
    "Vimmerby": "0884",
    "Vindeln": "2404",
    "Vingåker": "0428",
    "Vänersborg": "1487",
    "Vännäs": "2460",
    "Värmdö": "0120",
    "Värnamo": "0683",
    "Västervik": "0883",
    "Västerås": "1980",
    "Växjö": "0780",
    "Vårgårda": "1442",
    "Ydre": "0512",
    "Ystad": "1286",
    "Älmhult": "0765",
    "Älvdalen": "2039",
    "Älvkarleby": "0319",
    "Älvsbyn": "2560",
    "Ängelholm": "1292",
    "Åmål": "1492",
    "Ånge": "2260",
    "Åre": "2321",
    "Årjäng": "1765",
    "Åsele": "2463",
    "Åstorp": "1277",
    "Åtvidaberg": "0561",
    "Öckerö": "1407",
    "Ödeshög": "0509",
    "Örebro": "1880",
    "Örkelljunga": "1257",
    "Örnsköldsvik": "2284",
    "Östersund": "2380",
    "Österåker": "0117",
    "Östhammar": "0382",
    "Östra Göinge": "1256",
    "Överkalix": "2513",
    "Övertorneå": "2518",
}

# PDF city-name cleaning: raw key → canonical key
_CITY_ALIAS = {
    "Avfall & Återvin-": "Avfall & Återvinning Skaraborg",
    "Gästrike Återvin-": "Gästrike Återvinnare",
    "Landskrona-": "Landskrona-Svalöv (LSR)",
    "Landskrona-Svalöv": "Landskrona-Svalöv (LSR)",
    "Avfall & Återvinning": "Avfall & Återvinning Skaraborg",
    "Dals-Ed -": "Dals-Ed",
    "Gällivare -": "Gällivare",
    "Habo -": "Habo",
    "Hultsfred -": "Hultsfred",
    "Högsby -": "Högsby",
    "Jönköping -": "Jönköping",
    "Karlshamn -": "Karlshamn",
    "Lund -": "Lund",
    "Malå -": "Malå",
    "Mullsjö -": "Mullsjö",
    "Norrtälje -": "Norrtälje",
    "Norrtälje DS": "Norrtälje",
    "Olofström -": "Olofström",
    "Piteå -": "Piteå",
    "Sölvesborg -": "Sölvesborg",
    "Tidaholm -": "Tidaholm",
    "Ydre -": "Ydre",
    "Vimmerby -": "Vimmerby",
    "Boxholm DS": "Boxholm",
    "Degerfors ET": "Degerfors",
    "Habo DS": "Habo",
    "Hultsfred DS": "Hultsfred",
    "Högsby DS": "Högsby",
    "Jönköping DS": "Jönköping",
    "Mullsjö DS": "Mullsjö",
    "Stockholm DS": "Stockholm",
    "Boxholm DS - Optisk sortering Plastpåse Obl": "Boxholm",
    "Färgelanda - - Separata kärl Plastpåse Obl -": "Färgelanda",
    "Norsjö - - Tvådelade kärl Papper Obl": "Norsjö",
    "Ödeshög - - Optisk sortering Bioplast Obl": "Ödeshög",
    "LSR Landskrona-Sva-": "Landskrona-Svalöv (LSR)",
}

# Rows to drop entirely (county aggregate rows, header fragments)
_DROP_KEYS = {
    "BLEKINGE",
    "DALARNA",
    "GOTLAND",
    "GÄVLEBORG",
    "HALLAND",
    "JÄMTLAND",
    "JÖNKÖPING",
    "KALMAR",
    "KRONOBERG",
    "NORRBOTTEN",
    "SKÅNE",
    "STOCKHOLM",
    "SÖDERMANLAND",
    "UPPSALA",
    "VÄRMLAND",
    "VÄSTERBOTTEN",
    "VÄSTERNORRLAND",
    "VÄSTMANLAND",
    "ÖREBRO",
    "ÖSTERGÖTLAND",
    "UPPSALA 1)",
    "VÄSTMANLAND 2)",
    "VÄSTRA GÖTA-",
    "VÄSTRA GÖTALAND",
    "miljöbalken",
    "Lan",
    "Total",
    "Kommun/ källsorterar och växtnäring matavfall",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _numeric(value) -> float | None:
    """Return float for a real numeric value, else None."""
    if isinstance(value, int | float) and not isinstance(value, bool):
        return float(value)
    return None


def _try_float_swedish(s: str) -> float | None:
    """Parse a Swedish-formatted number (comma as decimal separator)."""
    try:
        return float(str(s).replace(",", "."))
    except (ValueError, AttributeError):
        return None


def _valid_from_year(year: int) -> date:
    return date(year, 1, 1)


def _clean_city(raw: str) -> str | None:
    """Return canonical city name, or None if the row should be dropped."""
    raw = raw.strip()
    if raw in _DROP_KEYS:
        return None
    return _CITY_ALIAS.get(raw, raw)


def _resolve_catchment_key(city_name: str) -> dict:
    """Return the catchment-lookup keys for an importer record.

    Priority:
    1. LAU ID (covers all 290 standard Swedish municipalities).
    2. Custom catchment name (6 joint-authority entities).
    3. Raw name fallback.
    """
    lau_id = _LAU_IDS.get(city_name)
    if lau_id:
        return {"nuts_or_lau_id": lau_id, "catchment_name": ""}
    brit_name = _CUSTOM_CATCHMENT_MAP.get(city_name)
    if brit_name:
        return {"catchment_name": brit_name, "nuts_or_lau_id": ""}
    return {"catchment_name": city_name, "nuts_or_lau_id": ""}


# ---------------------------------------------------------------------------
# Record builders
# ---------------------------------------------------------------------------


def _build_food_waste_record(
    city: str,
    sorting_method_name: str | None,
    bag_material: str | None,
    fee_system: str | None,
    connection_type: str | None,
    established: int | None,
    data_year: int,
    food_kg: float | None,
    conn_rate: float | None,
    description: str = "",
) -> dict:
    """Build a CollectionImporter record for food waste collection."""
    allowed_materials = [bag_material] if bag_material else []
    pvs = []
    if food_kg is not None:
        pvs.append(
            {
                "property_id": _PROP_SPECIFIC,
                "unit_name": _UNIT_KG,
                "year": data_year,
                "average": food_kg,
            }
        )
    if conn_rate is not None:
        pvs.append(
            {
                "property_id": _PROP_CONN_RATE,
                "unit_name": _UNIT_PCT_HH,
                "year": data_year,
                "average": conn_rate,
            }
        )
    return {
        **_resolve_catchment_key(city),
        "collection_system": _COLLECTION_SYSTEM_DOOR_TO_DOOR,
        "sorting_method": sorting_method_name or "",
        "waste_category": _WASTE_CATEGORY_FOOD,
        "allowed_materials": ", ".join(allowed_materials),
        "forbidden_materials": "",
        "fee_system": fee_system or "",
        "connection_type": connection_type or "",
        "established": established,
        "valid_from": _valid_from_year(data_year),
        "valid_until": None,
        "description": description,
        "flyer_urls": [],
        "property_values": pvs,
    }


def _build_residual_waste_record(
    city: str,
    fee_system: str | None,
    data_year: int,
    residual_kg: float,
) -> dict:
    """Build a CollectionImporter record for residual waste collection."""
    return {
        **_resolve_catchment_key(city),
        "collection_system": _COLLECTION_SYSTEM_DOOR_TO_DOOR,
        "sorting_method": "",
        "waste_category": _WASTE_CATEGORY_RESIDUAL,
        "allowed_materials": "",
        "forbidden_materials": "",
        "fee_system": fee_system or "",
        "connection_type": "",
        "established": None,
        "valid_from": _valid_from_year(data_year),
        "valid_until": None,
        "description": "",
        "flyer_urls": [],
        "property_values": [
            {
                "property_id": _PROP_SPECIFIC,
                "unit_name": _UNIT_KG,
                "year": data_year,
                "average": residual_kg,
            }
        ],
    }


def _build_no_collection_record(city: str, data_year: int) -> dict:
    """Build a record for municipalities with no separate food waste collection."""
    return {
        **_resolve_catchment_key(city),
        "collection_system": _COLLECTION_SYSTEM_NONE,
        "sorting_method": "",
        "waste_category": _WASTE_CATEGORY_FOOD,
        "allowed_materials": "",
        "forbidden_materials": "",
        "fee_system": "",
        "connection_type": "",
        "established": None,
        "valid_from": _valid_from_year(data_year),
        "valid_until": None,
        "description": "",
        "flyer_urls": [],
        "property_values": [],
    }


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------


def _parse_excel_2021(path: Path) -> list[dict]:
    """Parse the 2021 Avfall Sverige Excel file and return importer records."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        city_raw = row[0]
        if not city_raw:
            continue
        city = _clean_city(str(city_raw))
        if city is None:
            continue

        food_kg = _numeric(row[3])
        residual_kg = _numeric(row[6])
        conn_rate = _numeric(row[8])
        bag_raw = str(row[9]).strip() if row[9] else ""
        fee_raw = str(row[10]).strip() if row[10] else ""
        system_raw = str(row[12]).strip() if row[12] else ""
        impl_year = row[13]

        bag_material = _BAG_MATERIAL_MAP.get(bag_raw)
        fee_system = _FEE_MAP.get(fee_raw)
        established = (
            int(impl_year)
            if isinstance(impl_year, int | float)
            and not isinstance(impl_year, bool)
            and 1900 < impl_year < 2100
            else None
        )

        if system_raw in _NO_COLLECTION_SYSTEMS:
            records.append(_build_no_collection_record(city, 2021))
        else:
            sorting_method = _SORTING_METHOD_MAP.get(system_raw, system_raw) or None
            description = (
                "Bag type reported as 'Annan' (other) in 2021 Avfall Sverige source data."
                if bag_raw == "Annan"
                else ""
            )
            records.append(
                _build_food_waste_record(
                    city=city,
                    sorting_method_name=sorting_method
                    if sorting_method not in ("-", "")
                    else None,
                    bag_material=bag_material,
                    fee_system=fee_system,
                    connection_type=None,
                    established=established,
                    data_year=2021,
                    food_kg=food_kg,
                    conn_rate=conn_rate,
                    description=description,
                )
            )

        if residual_kg is not None:
            records.append(
                _build_residual_waste_record(city, fee_system, 2021, residual_kg)
            )

    return records


def _parse_pdf(path: Path, data_year: int) -> list[dict]:
    """Parse an Avfall Sverige annual PDF and return importer records."""
    try:
        import pdfplumber  # noqa: PLC0415
    except ImportError as exc:
        raise CommandError(
            "pdfplumber is required to read PDF files. "
            "Install it inside the container: pip install pdfplumber"
        ) from exc

    amounts: dict[str, tuple] = {}
    details: dict[str, dict] = {}

    with pdfplumber.open(str(path)) as pdf:
        _extract_table3(pdf, amounts)
        _extract_table5(pdf, data_year, details)

    all_cities = set(amounts) | set(details)
    records = []
    for city in sorted(all_cities):
        amt = amounts.get(city, (None, None))
        det = details.get(city, {})
        food_kg = _numeric(amt[0])
        residual_kg = _numeric(amt[1])

        if det.get("no_collection"):
            records.append(_build_no_collection_record(city, data_year))
        else:
            records.append(
                _build_food_waste_record(
                    city=city,
                    sorting_method_name=det.get("sorting_method"),
                    bag_material=det.get("bag_material"),
                    fee_system=None,
                    connection_type=det.get("connection_type"),
                    established=det.get("established"),
                    data_year=data_year,
                    food_kg=food_kg,
                    conn_rate=_numeric(det.get("conn_rate")),
                )
            )

        if residual_kg is not None:
            records.append(
                _build_residual_waste_record(city, None, data_year, residual_kg)
            )

    return records


def _extract_table3(pdf, amounts: dict) -> None:
    """Extract food and residual waste kg/person from Table 3 using line-based parsing."""
    in_table3 = False
    for page in pdf.pages:
        text = page.extract_text() or ""
        # Match the exact page header of data pages (not the TOC)
        if text.startswith("Tabell 3 Insamlade mängder hushållsavfall"):
            in_table3 = True
        if text.startswith("Tabell 4") and in_table3:
            break
        if not in_table3:
            continue
        for line in text.split("\n"):
            parts = line.strip().split()
            if not parts:
                continue
            upper = line.upper()
            if any(kw in upper for kw in ("MEDEL", "RIKET", "TABELL")):
                continue
            if "Matavfall" in line or "Restavfall" in line:
                continue
            # Split line into city name (non-numeric prefix) and numeric values
            city_parts: list[str] = []
            numeric_vals: list[float] = []
            for p in parts:
                v = _try_float_swedish(p)
                if v is not None and city_parts:
                    numeric_vals.append(v)
                elif v is None and (city_parts or p not in ("-", "ET", "DS")):
                    city_parts.append(p)
            if not city_parts or len(numeric_vals) < 2:
                continue
            # Strip trailing dash/placeholder tokens from city name
            while city_parts and city_parts[-1] in ("-", "ET", "DS"):
                city_parts.pop()
            if not city_parts:
                continue
            city_raw = " ".join(city_parts)
            city = _clean_city(city_raw)
            if not city:
                continue
            # Layout: total, food, residual, ...
            food_kg = numeric_vals[1] if len(numeric_vals) > 1 else None
            residual_kg = numeric_vals[2] if len(numeric_vals) > 2 else None
            if food_kg is not None or residual_kg is not None:
                amounts[city] = (food_kg, residual_kg)


def _extract_table5(pdf, data_year: int, details: dict) -> None:
    """Extract collection details from Table 5 using line-based parsing."""
    in_table5 = False
    for page in pdf.pages:
        text = page.extract_text() or ""
        # Match the exact page header of data pages (not the TOC)
        if text.startswith("Tabell 5") and "matavfall" in text[:80].lower():
            in_table5 = True
        if text.startswith("Tabell 6") and in_table5:
            break
        if not in_table5:
            continue
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            if any(
                kw in line
                for kw in (
                    "Tabell",
                    "Andel",
                    "MEDEL",
                    "RIKET",
                    "Kommun",
                    "förbund",
                    "nyttjas",
                    "Villor",
                )
            ):
                continue
            parts = line.split()
            if len(parts) < 2:
                continue

            # Locate Obl/Friv token — required for a parseable detail row
            obl_friv_idx = next(
                (i for i, p in enumerate(parts) if p in ("Obl", "Friv")), None
            )

            if obl_friv_idx is None:
                # May be a "no collection" row
                if "Ingen" in line:
                    city_parts = []
                    for p in parts:
                        if _try_float_swedish(p) is not None:
                            break
                        city_parts.append(p)
                    if city_parts:
                        city = _clean_city(" ".join(city_parts))
                        if city:
                            details[city] = {
                                "no_collection": True,
                                "sorting_method": None,
                                "bag_material": None,
                                "connection_type": None,
                                "established": None,
                                "conn_rate": None,
                            }
                continue

            conn_type = _CONNECTION_MAP.get(parts[obl_friv_idx])
            established = None
            if obl_friv_idx + 1 < len(parts):
                try:
                    yr = int(parts[obl_friv_idx + 1])
                    if 1900 < yr < 2100:
                        established = yr
                except ValueError:
                    pass

            # Locate bag type token
            bag_raw = None
            bag_idx = None
            for i, p in enumerate(parts):
                if p in _BAG_MATERIAL_MAP:
                    bag_raw = p
                    bag_idx = i
                    break
            if bag_idx is None:
                continue

            bag_material = _BAG_MATERIAL_MAP[bag_raw]

            # City name: tokens before the first numeric / DS / ET / '-' token
            city_parts = []
            for p in parts:
                if _try_float_swedish(p) is not None or p in ("DS", "ET", "-"):
                    break
                city_parts.append(p)
            if not city_parts:
                continue
            city = _clean_city(" ".join(city_parts))
            if not city:
                continue

            # Sorting method: tokens between last numeric/DS/ET and bag type
            last_num_idx = -1
            for i in range(bag_idx):
                if _try_float_swedish(parts[i]) is not None or parts[i] in (
                    "DS",
                    "ET",
                    "-",
                ):
                    last_num_idx = i
            system_tokens = parts[last_num_idx + 1 : bag_idx]
            system_raw = " ".join(system_tokens).strip()
            if "Separata" in system_raw:
                system_raw = "Separata kärl"
            elif "Optisk" in system_raw:
                system_raw = "Optisk sortering"
            elif "Fyrfacks" in system_raw:
                system_raw = "Fyrfackskärl"
            elif "Tvådelade" in system_raw:
                system_raw = "Tvådelade kärl"
            sorting_method = _SORTING_METHOD_MAP.get(system_raw)

            # Connection rate: first numeric token after city
            conn_rate = None
            for p in parts[len(city_parts) : bag_idx]:
                v = _try_float_swedish(p)
                if v is not None:
                    conn_rate = v
                    break

            details[city] = {
                "no_collection": False,
                "sorting_method": sorting_method,
                "bag_material": bag_material,
                "connection_type": conn_type,
                "established": established,
                "conn_rate": conn_rate,
            }


# ---------------------------------------------------------------------------
# JSON serialisation helper
# ---------------------------------------------------------------------------


def _date_to_str(value) -> str | None:
    """Serialise a date to ISO 8601 string for the JSON payload."""
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _records_to_json_serialisable(records: list[dict]) -> list[dict]:
    """Convert Python date objects in records to ISO strings for JSON transport."""
    out = []
    for rec in records:
        r = dict(rec)
        r["valid_from"] = _date_to_str(r.get("valid_from"))
        r["valid_until"] = _date_to_str(r.get("valid_until"))
        out.append(r)
    return out


# ---------------------------------------------------------------------------
# Django management command
# ---------------------------------------------------------------------------


class Command(BaseCommand):
    """Import Swedish waste collection data from Avfall Sverige reports (2021–2023).

    Runs locally against a BRIT instance via the bulk-import API endpoint.
    Parses Excel and PDF source files on the local machine, then POSTs records
    in batches.  Does not require Django models or database access.
    """

    help = (
        "Import Swedish waste collection data (2021–2023) via the BRIT API. "
        "Runs locally; parses Excel/PDF source files and POSTs to the import endpoint."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--api-url",
            type=str,
            required=True,
            help="Base URL of the BRIT instance, e.g. https://brit.example.com",
        )
        parser.add_argument(
            "--token",
            type=str,
            default=None,
            help="DRF auth token. Omit to obtain one via --username/--password.",
        )
        parser.add_argument(
            "--username",
            type=str,
            default=None,
            help="Username to obtain a DRF token (requires --password).",
        )
        parser.add_argument(
            "--password",
            type=str,
            default=None,
            help="Password to obtain a DRF token (requires --username).",
        )
        parser.add_argument(
            "--publication-status",
            type=str,
            default="private",
            choices=_VALID_STATUSES,
            help="Publication status for created records (default: private).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Pass dry_run=true to the API — no records are written.",
        )
        parser.add_argument(
            "--year",
            type=int,
            choices=(2021, 2022, 2023),
            default=None,
            help="Import only this specific year (default: all years).",
        )
        parser.add_argument(
            "--excel",
            type=str,
            default=str(_EXCEL_FILE),
            help=f"Path to 2021 Excel file (default: {_EXCEL_FILE}).",
        )
        parser.add_argument(
            "--pdf-2022",
            type=str,
            default=str(_PDF_2022),
            help=f"Path to 2022 PDF (default: {_PDF_2022}).",
        )
        parser.add_argument(
            "--pdf-2023",
            type=str,
            default=str(_PDF_2023),
            help=f"Path to 2023 PDF (default: {_PDF_2023}).",
        )
        parser.add_argument(
            "--batch-size",
            type=int,
            default=_BATCH_SIZE,
            help=f"Number of records per API request (default: {_BATCH_SIZE}).",
        )

    # ------------------------------------------------------------------
    # Auth helpers
    # ------------------------------------------------------------------

    def _get_token(self, api_url: str, username: str, password: str) -> str:
        """Obtain a DRF token via the token-auth endpoint."""
        try:
            url = f"{api_url.rstrip('/')}/api/auth/token/"
            payload = json.dumps({"username": username, "password": password}).encode()
            req = Request(
                url,
                data=payload,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
            token = data.get("token")
            if not token:
                raise CommandError(f"Token endpoint returned no token: {data}")
            return token
        except CommandError:
            raise
        except Exception as exc:
            raise CommandError(f"Could not obtain token: {exc}") from exc

    def _post_batch(
        self,
        api_url: str,
        token: str,
        records: list[dict],
        publication_status: str,
        dry_run: bool,
    ) -> dict:
        """POST one batch of records to the import endpoint; return stats dict."""
        try:
            url = f"{api_url.rstrip('/')}/waste_collection/api/collection/import/"
            payload = json.dumps(
                {
                    "records": records,
                    "publication_status": publication_status,
                    "dry_run": dry_run,
                }
            ).encode()
            req = Request(
                url,
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Token {token}",
                },
                method="POST",
            )
            with urlopen(req, timeout=120) as resp:
                return json.loads(resp.read())
        except Exception as exc:
            raise CommandError(f"API request failed: {exc}") from exc

    # ------------------------------------------------------------------
    # Entry point
    # ------------------------------------------------------------------

    def handle(self, *args, **options):
        api_url = options["api_url"]
        dry_run = options["dry_run"]
        publication_status = options["publication_status"]
        batch_size = options["batch_size"]
        only_year = options["year"]
        years_to_run = [2021, 2022, 2023] if only_year is None else [only_year]

        paths = {
            2021: Path(options["excel"]),
            2022: Path(options["pdf_2022"]),
            2023: Path(options["pdf_2023"]),
        }

        # Resolve auth token
        token = options.get("token")
        if not token:
            username = options.get("username")
            password = options.get("password")
            if not username or not password:
                raise CommandError("Provide --token or both --username and --password.")
            token = self._get_token(api_url, username, password)
            self.stdout.write("Token obtained.\n")

        if dry_run:
            self.stdout.write("DRY RUN — no records will be written.\n")

        totals = {
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "predecessor_links": 0,
            "cpv_created": 0,
            "cpv_skipped": 0,
            "flyers_created": 0,
            "warnings": [],
            "changes": [],
        }

        for year in years_to_run:
            path = paths[year]
            if not path.exists():
                raise CommandError(f"Source file not found: {path}")

            self.stdout.write(f"\n--- Parsing {year} from {path.name} ---")
            try:
                if year == 2021:
                    records = _parse_excel_2021(path)
                else:
                    records = _parse_pdf(path, year)
            except CommandError:
                raise
            except Exception as exc:
                raise CommandError(f"Error parsing {year} data: {exc}") from exc

            records_json = _records_to_json_serialisable(records)
            self.stdout.write(f"  {len(records_json)} records parsed")

            batches = [
                records_json[i : i + batch_size]
                for i in range(0, len(records_json), batch_size)
            ]
            for idx, batch in enumerate(batches, start=1):
                self.stdout.write(
                    f"  Posting batch {idx}/{len(batches)} ({len(batch)} records)…"
                )
                sys.stdout.flush()
                result = self._post_batch(
                    api_url, token, batch, publication_status, dry_run
                )
                stats = result.get("stats", {})
                self._merge_stats(totals, stats)
                self.stdout.write(
                    f" created={stats.get('created', 0)}"
                    f" updated={stats.get('updated', 0)}"
                    f" skipped={stats.get('skipped', 0)}\n"
                )

        self.stdout.write("\n=== Import Summary ===\n")
        self.stdout.write(f"  Collections created:  {totals['created']}\n")
        self.stdout.write(f"  Collections updated:  {totals['updated']}\n")
        self.stdout.write(f"  Collections skipped:  {totals['skipped']}\n")
        self.stdout.write(f"  Predecessor links:    {totals['predecessor_links']}\n")
        self.stdout.write(f"  CPVs created:         {totals['cpv_created']}\n")
        self.stdout.write(f"  CPVs skipped:         {totals['cpv_skipped']}\n")
        self.stdout.write(f"  Flyers created:       {totals['flyers_created']}\n")
        if totals["warnings"]:
            self.stdout.write(f"\n  Warnings ({len(totals['warnings'])}):\n")
            for w in totals["warnings"]:
                self.stdout.write(f"    {w}\n")
        if totals["changes"] and dry_run:
            self.stdout.write(
                f"\n  Changes that would be made ({len(totals['changes'])}):\n"
            )
            for change in totals["changes"][:20]:  # Show first 20 changes
                self.stdout.write(f"    {change}\n")
            if len(totals["changes"]) > 20:
                self.stdout.write(f"    ... and {len(totals['changes']) - 20} more\n")

    @staticmethod
    def _merge_stats(totals: dict, stats: dict) -> None:
        """Accumulate per-batch stats into running totals."""
        for key in (
            "created",
            "updated",
            "skipped",
            "predecessor_links",
            "cpv_created",
            "cpv_skipped",
            "flyers_created",
        ):
            totals[key] = totals.get(key, 0) + stats.get(key, 0)
        totals["warnings"].extend(stats.get("warnings", []))
        totals["changes"].extend(stats.get("changes", []))
