"""Excel renderers for materials app exports."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.file_export.renderers import BaseCSVRenderer, BaseXLSXRenderer

# Excel column headers matching the import format (order matches input files)
MEASUREMENT_HEADERS = [
    "Parameter",
    "Abbr., acronym",
    "Parameter group",
    "Value",
    "Standard deviation",
    "n",
    "Unit",
    "Reference parameter, Base",
    "Method",
    "Source",
    "Comments",
]

METADATA_LABELS = [
    "Material type",
    "Sample name",
    "Sample info (e.g. structure, harvesting, storing)",
    "Sample origin (e.g. location, region)",
    "Sample campaign (e.g. season, project)",
    "Sample date",
    "Analysis date",
    "other (e.g. analysis objective)",
    "Analysis laboratorium",
    "other (e.g. lab accreditation)",
    "Source(s)",
    "DOI",
    "Entry done by:",
]

# Labels that should be styled in red (matching input template)
RED_LABELS = {
    "Analysis date",
    "other (e.g. analysis objective)",
    "Analysis laboratorium",
    "other (e.g. lab accreditation)",
}

# Style definitions matching input Excel template
STYLE_BOLD = Font(bold=True, size=11)
STYLE_BOLD_RED = Font(bold=True, size=11, color="C00000")
STYLE_HEADER_FILL = PatternFill(
    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
)
STYLE_HEADER_BORDER = Border(bottom=Side(style="medium"))
STYLE_LEFT_ALIGN = Alignment(horizontal="left")


class SampleCSVRenderer(BaseCSVRenderer):
    labels = {
        "id": "ID",
        "name": "Sample",
        "material": "Material",
        "series": "Series",
        "timestep": "Timestep",
        "datetime": "Date/time",
        "standalone": "Standalone",
        "publication_status": "Publication status",
        "owner": "Owner",
        "created_at": "Created at",
        "description": "Description",
    }
    header = labels


class SampleXLSXRenderer(BaseXLSXRenderer):
    labels = SampleCSVRenderer.labels


# Parameter group colors (approximating Excel theme colors with tints)
# Based on analysis of input Excel files
GROUP_COLORS = {
    # From input Excel analysis
    "Organic sum parameters": "B4C6A4",  # theme 6, tint 0.4 (greenish)
    "Inorganic sum parameters": "BFBFBF",  # theme 0, tint -0.25 (gray)
    "Chemical Elements": "F4B183",  # theme 4, tint 0.4 (orange)
    "Nitrogen": "D6DCE4",  # theme 3, tint 0.8 (light blue-gray)
    "Carbon": "D6DCE4",  # theme 3, tint 0.8 (light blue-gray)
    "chemical parameters": "A9D08E",  # theme 8, tint 0.4 (green)
    "Anions": "E2EFDA",  # theme 5, tint 0.8 (light green)
    "degradation properties (anaerobic )": "C5E0B3",  # theme 5, tint 0.4
    "degradation properties": "C5E0B3",  # theme 5, tint 0.4
    "Physical parameters": "9DC3E6",  # theme 7, tint 0.4 (blue)
    "Fractions of impurities": "FCE4D6",  # theme 9, tint 0.8 (light orange)
    "basic parameters": "FFFFFF",  # no fill (white)
    "basic parameters - chemical parameters": "FFFFFF",
    "basic parameters - Physical parameters": "FFFFFF",
    # Database group names (mapped to similar colors)
    "Biochemical Composition": "A9D08E",  # green
    "Carbon Fractions": "D6DCE4",  # light blue-gray
    "Macro Components": "B4C6A4",  # greenish
    "Nitrogen fractions": "D6DCE4",  # light blue-gray
    "Organic/Inorganic": "BFBFBF",  # gray
    "Solids/Water": "9DC3E6",  # blue
}

# Fallback colors for unknown groups (cycle through these)
FALLBACK_COLORS = [
    "FCE4D6",  # light orange
    "DDEBF7",  # light blue
    "E2EFDA",  # light green
    "FFF2CC",  # light yellow
    "F8CBAD",  # peach
    "D9E1F2",  # lavender
]

# Column widths for the measurements table (matches header order)
COLUMN_WIDTHS = [30, 15, 25, 12, 18, 6, 10, 25, 30, 40, 40]

# Sample-property worksheet definition. Captures non-mass sample properties
# such as moisture, density, pH that live on MaterialPropertyValue rather
# than ComponentMeasurement.
PROPERTY_HEADERS = [
    "Property",
    "Value",
    "Standard deviation",
    "Unit",
    "Reference parameter, Base",
    "Method",
    "Source",
    "Comments",
]
PROPERTY_COLUMN_WIDTHS = [30, 12, 18, 10, 25, 30, 40, 40]


class SampleMeasurementsXLSXRenderer:
    """Renderer for exporting sample measurements to Excel format matching the import template."""

    def __init__(self, sample, measurements, progress_callback=None):
        """
        Initialize the renderer.

        Args:
            sample: Sample model instance
            measurements: QuerySet of ComponentMeasurement objects
            progress_callback: Optional callable(percent, status) for progress reporting
        """
        self.sample = sample
        self.measurements = measurements
        self.progress_callback = progress_callback
        self._unknown_group_colors = {}

    def _report_progress(self, current, total, status):
        """Report progress via callback if available."""
        if self.progress_callback:
            percent = int((current / total) * 100) if total > 0 else 100
            self.progress_callback(percent, status)

    def _build_metadata_values(self):
        """Build dictionary of metadata label -> value mappings."""
        sample = self.sample
        sample_sources = sample.sources.all()

        return {
            "Material type": sample.material.name if sample.material else "",
            "Sample name": sample.name or "",
            "Sample info (e.g. structure, harvesting, storing)": sample.description
            or "",
            "Sample origin (e.g. location, region)": sample.location or "",
            "Sample campaign (e.g. season, project)": sample.series.name
            if sample.series
            else "",
            "Sample date": sample.datetime.strftime("%Y-%m-%d")
            if sample.datetime
            else "",
            "Analysis date": sample.analysis_date.strftime("%Y-%m-%d")
            if sample.analysis_date
            else "",
            "other (e.g. analysis objective)": sample.analysis_objective or "",
            "Analysis laboratorium": sample.analysis_laboratory or "",
            "other (e.g. lab accreditation)": sample.lab_accreditation or "",
            "Source(s)": "; ".join(s.abbreviation for s in sample_sources)
            if sample_sources
            else "",
            "DOI": "; ".join(s.doi for s in sample_sources if s.doi) or "",
            "Entry done by:": sample.owner.username if sample.owner else "",
        }

    def _get_group_color(self, group_name):
        """Get the fill color hex for a parameter group."""
        if not group_name:
            return "FFFFFF"

        if group_name in GROUP_COLORS:
            return GROUP_COLORS[group_name]

        if group_name in self._unknown_group_colors:
            return self._unknown_group_colors[group_name]

        # Assign a fallback color for unknown groups
        fallback_idx = len(self._unknown_group_colors) % len(FALLBACK_COLORS)
        color_hex = FALLBACK_COLORS[fallback_idx]
        self._unknown_group_colors[group_name] = color_hex
        return color_hex

    def _write_metadata_section(self, ws):
        """Write the metadata section to the worksheet. Returns the next row number."""
        metadata_values = self._build_metadata_values()

        row_num = 1
        for label in METADATA_LABELS:
            label_cell = ws.cell(row=row_num, column=1, value=label)
            label_cell.font = STYLE_BOLD_RED if label in RED_LABELS else STYLE_BOLD
            label_cell.alignment = STYLE_LEFT_ALIGN

            value_cell = ws.cell(
                row=row_num, column=2, value=metadata_values.get(label, "")
            )
            value_cell.alignment = STYLE_LEFT_ALIGN
            row_num += 1

        return row_num + 1  # Skip one row before data table

    def _write_header_row(self, ws, row_num):
        """Write the header row for the measurements table. Returns the next row number."""
        for col_num, header in enumerate(MEASUREMENT_HEADERS, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = STYLE_BOLD
            cell.fill = STYLE_HEADER_FILL
            cell.border = STYLE_HEADER_BORDER

        return row_num + 1

    def _write_measurement_row(self, ws, row_num, measurement):
        """Write a single measurement row with appropriate styling."""
        group_name = measurement.group.name if measurement.group else ""
        color_hex = self._get_group_color(group_name)

        # Create fill style (skip if white/no fill)
        row_fill = None
        if color_hex and color_hex != "FFFFFF":
            row_fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

        # Sources
        measurement_sources = measurement.sources.all()
        source_value = (
            "; ".join(s.abbreviation for s in measurement_sources)
            if measurement_sources
            else ""
        )

        # Build cell data matching header order:
        # Parameter, Abbr., Parameter group, Value, Std dev, n, Unit, Ref param, Method, Source, Comments
        cells_data = [
            (1, measurement.component.name if measurement.component else ""),
            (2, measurement.component.abbreviation if measurement.component else ""),
            (3, group_name),
            (4, float(measurement.average) if measurement.average is not None else ""),
            (
                5,
                float(measurement.standard_deviation)
                if measurement.standard_deviation is not None
                else "",
            ),
            (6, measurement.sample_size if measurement.sample_size is not None else ""),
            (7, measurement.unit.name if measurement.unit else ""),
            (
                8,
                measurement.basis_component.name if measurement.basis_component else "",
            ),
            (
                9,
                measurement.analytical_method.name
                if measurement.analytical_method
                else "",
            ),
            (10, source_value),
            (11, measurement.comment or ""),
        ]

        for col, value in cells_data:
            cell = ws.cell(row=row_num, column=col, value=value)
            if row_fill:
                cell.fill = row_fill

    def _set_column_widths(self, ws):
        """Set column widths for better readability."""
        for col_num, width in enumerate(COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

    def _iter_property_values(self):
        """Yield MaterialPropertyValue rows for the current sample.

        Imported lazily to avoid any risk of an import cycle between
        ``materials.models`` and the Celery task module that pulls in
        ``materials.renderers``.
        """
        from .models import MaterialPropertyValue

        return (
            MaterialPropertyValue.objects.filter(sample=self.sample)
            .select_related("property", "basis_component", "analytical_method", "unit")
            .prefetch_related("sources")
            .order_by("property__name")
        )

    @staticmethod
    def _format_sources(sources):
        if not sources:
            return ""
        return "; ".join(
            s.abbreviation for s in sources if s.abbreviation
        ) or "; ".join(str(s) for s in sources)

    def _write_property_header(self, ws, row_num):
        for col_num, header in enumerate(PROPERTY_HEADERS, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = STYLE_BOLD
            cell.fill = STYLE_HEADER_FILL
            cell.border = STYLE_HEADER_BORDER
        return row_num + 1

    def _write_property_row(self, ws, row_num, value):
        sources = list(value.sources.all())
        cells_data = [
            (1, value.property.name if value.property else ""),
            (
                2,
                float(value.average) if value.average is not None else "",
            ),
            (
                3,
                float(value.standard_deviation)
                if value.standard_deviation is not None
                else "",
            ),
            (4, value.unit.name if value.unit else ""),
            (5, value.basis_component.name if value.basis_component else ""),
            (
                6,
                value.analytical_method.name if value.analytical_method else "",
            ),
            (7, self._format_sources(sources)),
            (8, getattr(value, "comment", "") or ""),
        ]
        for col, cell_value in cells_data:
            ws.cell(row=row_num, column=col, value=cell_value)

    def _set_property_column_widths(self, ws):
        for col_num, width in enumerate(PROPERTY_COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

    def _write_properties_sheet(self, wb):
        """Append a second worksheet listing sample-level properties.

        Uses the same metadata header + data layout as the measurements
        sheet so the two read as siblings rather than strangers.
        """
        property_values = list(self._iter_property_values())
        ws = wb.create_sheet(title="Sample Properties")

        # Reuse the metadata block so both sheets are self-contained.
        row_num = self._write_metadata_section(ws)
        row_num = self._write_property_header(ws, row_num)

        for value in property_values:
            self._write_property_row(ws, row_num, value)
            row_num += 1

        self._set_property_column_widths(ws)
        return len(property_values)

    def render(self):
        """
        Render the sample measurements to an Excel workbook.

        Returns:
            BytesIO buffer containing the Excel file
        """
        total = self.measurements.count()
        self._report_progress(0, total, "Creating workbook...")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Measurements"[:31]

        # Write metadata section
        row_num = self._write_metadata_section(ws)

        # Write header row
        row_num = self._write_header_row(ws, row_num)

        # Write measurement data
        for idx, measurement in enumerate(self.measurements):
            self._write_measurement_row(ws, row_num, measurement)
            row_num += 1

            # Report progress
            self._report_progress(idx + 1, total, "Writing measurements...")

        # Adjust column widths
        self._set_column_widths(ws)

        # Append a second worksheet for sample-level properties.
        self._report_progress(total, total, "Writing sample properties...")
        self._write_properties_sheet(wb)

        self._report_progress(total, total, "Saving file...")

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
