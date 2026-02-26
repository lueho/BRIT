from collections import OrderedDict, namedtuple
from io import BytesIO
from unittest.mock import MagicMock, patch

from django.contrib.auth.models import User
from django.test import SimpleTestCase, TestCase
from openpyxl import load_workbook

from ..export_registry import (
    EXPORT_REGISTRY,
    ExportSpec,
    get_export_spec,
    register_export,
)
from ..generic_tasks import BATCH_SIZE, export_user_created_object_to_file
from ..renderers import BaseCSVRenderer, BaseXLSXRenderer


class ExportRegistryTestCase(SimpleTestCase):
    """Tests for the export registry functions."""

    def setUp(self):
        self._original_registry = EXPORT_REGISTRY.copy()

    def tearDown(self):
        EXPORT_REGISTRY.clear()
        EXPORT_REGISTRY.update(self._original_registry)

    def test_register_and_retrieve_spec(self):
        """Registered spec can be retrieved by label."""
        filterset = MagicMock()
        serializer = MagicMock()
        renderers = {"csv": MagicMock()}

        register_export("auth.User", filterset, serializer, renderers)

        spec = get_export_spec("auth.User")
        self.assertIsInstance(spec, ExportSpec)
        self.assertIs(spec.model, User)
        self.assertIs(spec.filterset, filterset)
        self.assertIs(spec.serializer, serializer)
        self.assertEqual(spec.renderers, renderers)

    def test_get_unknown_label_raises_key_error(self):
        """Requesting an unregistered label should raise KeyError."""
        with self.assertRaises(KeyError):
            get_export_spec("nonexistent.Model")

    def test_register_overwrites_existing_entry(self):
        """Re-registering the same label should overwrite the previous entry."""
        register_export("auth.User", MagicMock(), MagicMock(), {})
        new_serializer = MagicMock()
        register_export("auth.User", MagicMock(), new_serializer, {})

        spec = get_export_spec("auth.User")
        self.assertIs(spec.serializer, new_serializer)


class BaseXLSXRendererTestCase(SimpleTestCase):
    def setUp(self):
        self.renderer = BaseXLSXRenderer()
        self.renderer.workbook_options = {"in_memory": True}
        self.content = [OrderedDict({"column_1": "content", "column_2": "content"})]
        self.file = BytesIO()

    def test_creates_empty_xlsx_sheet_if_handed_empty_data_dictionary(self):
        self.renderer.render(self.file, {})
        wb = load_workbook(self.file)
        self.assertEqual(1, len(wb.sheetnames))
        self.assertEqual("sheet 1", wb.sheetnames[0])

    def test_uses_dictionary_keys_for_header_if_labels_are_missing(self):
        self.renderer.render(self.file, self.content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("column_1", ws.cell(row=1, column=1).value)
        self.assertEqual("column_2", ws.cell(row=1, column=2).value)

    def test_maintains_order_of_dict_keys(self):
        self.renderer.labels = {"column_2": "Column 2", "column_1": "Column 1"}
        self.renderer.render(self.file, self.content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("Column 2", ws.cell(row=1, column=1).value)
        self.assertEqual("Column 1", ws.cell(row=1, column=2).value)

    def test_writes_data_values_to_cells(self):
        content = [OrderedDict({"col_a": "value_a", "col_b": 42})]
        self.renderer.render(self.file, content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("value_a", ws.cell(row=2, column=1).value)
        self.assertEqual(42, ws.cell(row=2, column=2).value)

    def test_writes_multiple_rows(self):
        content = [
            OrderedDict({"col": "row1"}),
            OrderedDict({"col": "row2"}),
            OrderedDict({"col": "row3"}),
        ]
        self.renderer.render(self.file, content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("row1", ws.cell(row=2, column=1).value)
        self.assertEqual("row2", ws.cell(row=3, column=1).value)
        self.assertEqual("row3", ws.cell(row=4, column=1).value)

    def test_preset_column_order_controls_output_columns(self):
        self.renderer.column_order = ["column_2", "column_1"]
        self.renderer.render(self.file, self.content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("column_2", ws.cell(row=1, column=1).value)
        self.assertEqual("column_1", ws.cell(row=1, column=2).value)
        self.assertEqual("content", ws.cell(row=2, column=1).value)
        self.assertEqual("content", ws.cell(row=2, column=2).value)

    def test_labels_applied_with_preset_column_order(self):
        self.renderer.column_order = ["column_2", "column_1"]
        self.renderer.labels = {"column_1": "First", "column_2": "Second"}
        self.renderer.render(self.file, self.content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("Second", ws.cell(row=1, column=1).value)
        self.assertEqual("First", ws.cell(row=1, column=2).value)

    def test_missing_key_in_data_row_renders_none(self):
        content = [OrderedDict({"col_a": "present"})]
        self.renderer.column_order = ["col_a", "col_b"]
        self.renderer.render(self.file, content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("present", ws.cell(row=2, column=1).value)
        self.assertIsNone(ws.cell(row=2, column=2).value)

    def test_empty_list_creates_sheet_with_no_data_rows(self):
        self.renderer.render(self.file, [])
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertIsNone(ws.cell(row=1, column=1).value)


class BaseCSVRendererTestCase(SimpleTestCase):
    def setUp(self):
        self.renderer = BaseCSVRenderer()
        self.file = BytesIO()

    def test_writes_csv_content_to_file(self):
        content = [OrderedDict({"col_a": "val1", "col_b": "val2"})]
        self.renderer.render(self.file, content)
        output = self.file.getvalue().decode("utf-8")
        self.assertIn("col_a", output)
        self.assertIn("val1", output)
        self.assertIn("col_b", output)
        self.assertIn("val2", output)

    def test_writes_multiple_rows_to_csv(self):
        content = [
            OrderedDict({"col": "row1"}),
            OrderedDict({"col": "row2"}),
        ]
        self.renderer.render(self.file, content)
        lines = self.file.getvalue().decode("utf-8").strip().splitlines()
        self.assertEqual(3, len(lines))

    def test_empty_data_produces_empty_output(self):
        self.renderer.render(self.file, [])
        output = self.file.getvalue()
        self.assertEqual(b"", output)


class DummyFilterSet:
    """Minimal filterset stand-in that passes through the queryset unchanged."""

    def __init__(self, data, queryset):
        self.qs = queryset


class DummySerializer:
    """Minimal serializer stand-in that returns dicts with a pk key."""

    def __init__(self, instances, many=False):
        self.data = [OrderedDict({"pk": obj.pk}) for obj in instances]


TaskExportSpec = namedtuple(
    "TaskExportSpec", ["model", "filterset", "serializer", "renderers"]
)


class ExportTaskTestCase(TestCase):
    """Tests for export_user_created_object_to_file."""

    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(username="export_owner")
        cls.other = User.objects.create_user(username="export_other")

    def _run_task(self, model_label, file_format, query_params, context):
        """Invoke the task's run() with a mock self and return (result, mock_self)."""
        mock_self = MagicMock()
        mock_self.request.id = "fake-request-id"
        run_fn = export_user_created_object_to_file.run.__func__
        result = run_fn(mock_self, model_label, file_format, query_params, context)
        return result, mock_self

    def _make_spec(self, model=User):
        """Build a task export spec using User as the model."""
        return TaskExportSpec(
            model=model,
            filterset=DummyFilterSet,
            serializer=DummySerializer,
            renderers={"csv": MagicMock(), "xlsx": MagicMock()},
        )

    @patch(
        "utils.file_export.generic_tasks.utils.file_export.storages.write_file_for_download"
    )
    @patch("utils.file_export.generic_tasks.get_export_spec")
    def test_public_scope_returns_all_for_model_without_publication_status(
        self, mock_get_spec, mock_write
    ):
        """User model has no publication_status, so public scope should return all."""
        spec = self._make_spec()
        mock_get_spec.return_value = spec
        mock_write.return_value = "https://example.com/file.csv"

        result, _mock_self = self._run_task(
            "auth.User", "csv", {}, {"user_id": self.owner.pk, "list_type": "public"}
        )

        self.assertEqual(result, "https://example.com/file.csv")
        mock_write.assert_called_once()
        call_args, _ = mock_write.call_args
        self.assertEqual(call_args[0], "user_fake-request-id.csv")

    @patch(
        "utils.file_export.generic_tasks.utils.file_export.storages.write_file_for_download"
    )
    @patch("utils.file_export.generic_tasks.get_export_spec")
    def test_private_scope_filters_by_owner(self, mock_get_spec, mock_write):
        """Private scope should filter by owner_id."""
        mock_model = MagicMock()
        mock_qs = MagicMock()
        mock_model.objects.filter.return_value = mock_qs
        mock_model._meta.get_fields.return_value = []
        mock_model._meta.model_name = "mockmodel"

        mock_filtered = MagicMock()
        mock_filtered.qs = mock_qs
        mock_qs.count.return_value = 0

        spec = TaskExportSpec(
            model=mock_model,
            filterset=lambda data, queryset: mock_filtered,
            serializer=DummySerializer,
            renderers={"csv": MagicMock()},
        )
        mock_get_spec.return_value = spec
        mock_write.return_value = "url"

        self._run_task("test.Model", "csv", {}, {"user_id": 42, "list_type": "private"})

        mock_model.objects.filter.assert_called_once_with(owner_id=42)

    @patch(
        "utils.file_export.generic_tasks.utils.file_export.storages.write_file_for_download"
    )
    @patch("utils.file_export.generic_tasks.get_export_spec")
    def test_review_scope_returns_empty_for_model_without_publication_status(
        self, mock_get_spec, mock_write
    ):
        """Review scope on a model without publication_status should yield empty qs."""
        spec = self._make_spec()
        mock_get_spec.return_value = spec
        mock_write.return_value = "https://example.com/file.csv"

        _, _mock_self = self._run_task(
            "auth.User", "csv", {}, {"user_id": self.owner.pk, "list_type": "review"}
        )

        call_args, _ = mock_write.call_args
        data = call_args[1]
        self.assertEqual(data, [])

    @patch(
        "utils.file_export.generic_tasks.utils.file_export.storages.write_file_for_download"
    )
    @patch("utils.file_export.generic_tasks.get_export_spec")
    def test_reports_initial_progress_with_zero(self, mock_get_spec, mock_write):
        """Task should report initial progress with current=0."""
        spec = self._make_spec()
        mock_get_spec.return_value = spec
        mock_write.return_value = "url"

        _, mock_self = self._run_task(
            "auth.User", "csv", {}, {"user_id": self.owner.pk, "list_type": "public"}
        )

        first_call = mock_self.update_state.call_args_list[0]
        self.assertEqual(first_call[1]["state"], "PROGRESS")
        self.assertEqual(first_call[1]["meta"]["current"], 0)
        self.assertEqual(first_call[1]["meta"]["percent"], 0)

    @patch(
        "utils.file_export.generic_tasks.utils.file_export.storages.write_file_for_download"
    )
    @patch("utils.file_export.generic_tasks.get_export_spec")
    def test_empty_queryset_reports_progress_correctly(self, mock_get_spec, mock_write):
        """Empty queryset should report initial progress and not enter the batch loop."""
        spec = self._make_spec()
        mock_get_spec.return_value = spec
        mock_write.return_value = "url"

        _, mock_self = self._run_task(
            "auth.User", "csv", {}, {"user_id": self.owner.pk, "list_type": "review"}
        )

        mock_self.update_state.assert_called_once_with(
            state="PROGRESS", meta={"current": 0, "total": 0, "percent": 0}
        )

    @patch(
        "utils.file_export.generic_tasks.utils.file_export.storages.write_file_for_download"
    )
    @patch("utils.file_export.generic_tasks.get_export_spec")
    def test_uses_correct_renderer_for_format(self, mock_get_spec, mock_write):
        """Task should select the renderer matching the requested format."""
        spec = self._make_spec()
        mock_get_spec.return_value = spec
        mock_write.return_value = "url"

        _, _mock_self = self._run_task(
            "auth.User", "xlsx", {}, {"user_id": self.owner.pk, "list_type": "public"}
        )

        call_args, _ = mock_write.call_args
        self.assertIs(call_args[2], spec.renderers["xlsx"])
        self.assertTrue(call_args[0].endswith(".xlsx"))

    @patch(
        "utils.file_export.generic_tasks.utils.file_export.storages.write_file_for_download"
    )
    @patch("utils.file_export.generic_tasks.get_export_spec")
    def test_batch_progress_reporting(self, mock_get_spec, mock_write):
        """Verify progress is reported for each batch when total > BATCH_SIZE."""
        users_needed = BATCH_SIZE + 10
        existing = User.objects.count()
        for i in range(users_needed - existing):
            User.objects.create_user(username=f"batch_user_{i}")

        spec = self._make_spec()
        mock_get_spec.return_value = spec
        mock_write.return_value = "url"

        _, mock_self = self._run_task(
            "auth.User", "csv", {}, {"user_id": self.owner.pk, "list_type": "public"}
        )

        total = User.objects.count()
        expected_calls = 1 + ((total + BATCH_SIZE - 1) // BATCH_SIZE)
        self.assertEqual(mock_self.update_state.call_count, expected_calls)

        last_call = mock_self.update_state.call_args_list[-1]
        self.assertEqual(last_call[1]["meta"]["percent"], 100)
        self.assertEqual(last_call[1]["meta"]["current"], total)
