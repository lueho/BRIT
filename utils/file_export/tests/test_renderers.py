from collections import OrderedDict
from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import load_workbook

from utils.file_export.renderers import BaseXLSXRenderer


class BaseXLSXRendererTestCase(SimpleTestCase):
    def setUp(self):
        self.renderer = BaseXLSXRenderer()
        self.renderer.workbook_options = {"in_memory": True}
        self.content = [OrderedDict({"column_1": "content", "column_2": "content"})]
        self.file = BytesIO()

    def test_creates_empty_xlsx_sheet_if_handed_empty_data_dictionary(self):
        self.renderer.render(self.file, {})
        wb = load_workbook(self.file)
        self.assertEqual(1, len(wb.sheetnames))
        self.assertEqual("sheet 1", wb.sheetnames[0])

    def test_uses_dictionary_keys_for_header_if_labels_are_missing(self):
        self.renderer.render(self.file, self.content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("column_1", ws.cell(row=1, column=1).value)
        self.assertEqual("column_2", ws.cell(row=1, column=2).value)

    def test_maintains_order_of_dict_keys(self):
        self.renderer.labels = {"column_2": "Column 2", "column_1": "Column 1"}
        self.renderer.render(self.file, self.content)
        wb = load_workbook(self.file)
        ws = wb.active
        self.assertEqual("Column 2", ws.cell(row=1, column=1).value)
        self.assertEqual("Column 1", ws.cell(row=1, column=2).value)
