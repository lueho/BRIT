from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from django.test import SimpleTestCase

from sources.waste_collection.management.commands import (
    import_denmark_affaldsstatistik_cpvs as importer,
)


class DenmarkAffaldsstatistikImportTests(SimpleTestCase):
    def _write_workbook(self):
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Municipal Waste - Primær mængde"
        ws.append(
            [
                "Affaldstype",
                None,
                None,
                None,
                None,
                None,
                "Behandling",
                None,
                None,
                "Affaldets oprindelse",
                None,
                None,
                "Indsamlet affaldsmængde pr. år",
            ]
        )
        ws.append(
            [
                "Erhverv [E] eller husholdningsaffald [H]",
                "Affaldsfraktionskode",
                "Affaldsfraktion",
                "Ny affaldsfraktion",
                "EAKLevel3Kode",
                "EAKLevel3Navn",
                "RecoveryDisposalKode",
                "RecoveryDisposalNavn",
                "Ny behandling",
                "Producenttype",
                "Kommune",
                "Region",
                "2021 [ton]",
                "2022 [ton]",
            ]
        )
        ws.append(
            [
                "H",
                "H01",
                "Madaffald",
                "Madaffald",
                "20 01 08",
                "Bionedbrydeligt køkken- og kantineaffald",
                "R3",
                "Genanvendelse",
                "Genanvendelse",
                "Kommune",
                "Albertslund",
                "Region Hovedstaden",
                25,
                "INTET",
            ]
        )
        ws.append(
            [
                "H",
                "H02",
                "Haveaffald",
                "Haveaffald",
                "20 02 01",
                "Bionedbrydeligt affald",
                "R3",
                "Genanvendelse",
                "Genanvendelse",
                "Kommune",
                "Albertslund",
                "Region Hovedstaden",
                75,
                None,
            ]
        )
        ws.append(
            [
                "H",
                "H03",
                "Glas",
                "Glas",
                "20 01 02",
                "Glas",
                "R5",
                "Genanvendelse",
                "Genanvendelse",
                "Kommune",
                "Albertslund",
                "Region Hovedstaden",
                999,
                None,
            ]
        )

        recycling = workbook.create_sheet("Kommunal MW reel genanvendelse")
        recycling.append([None, "Husholdningsaffald, indsamlet total", None])
        recycling.append(["Kommune", "1000 tons", "kg pr indbygger"])
        recycling.append(["Albertslund", 10, 100])

        temp = NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp.close()
        workbook.save(temp.name)
        return Path(temp.name)

    def test_parse_primary_measurements_maps_food_and_green_waste(self):
        path = self._write_workbook()
        self.addCleanup(path.unlink)

        measurements = importer.parse_primary_measurements(path, years={2021})

        self.assertEqual(len(measurements), 2)
        by_category = {
            measurement.waste_category: measurement for measurement in measurements
        }
        self.assertEqual(by_category["Food waste"].source_fraction, "Madaffald")
        self.assertEqual(by_category["Food waste"].total_mg, 25)
        self.assertEqual(by_category["Green waste"].source_fraction, "Haveaffald")
        self.assertEqual(by_category["Green waste"].total_mg, 75)

    def test_build_cpv_records_matches_danish_catchment_names(self):
        measurements = [
            importer.DenmarkMeasurement(
                municipality="Århus",
                waste_category="Food waste",
                source_fraction="Madaffald",
                year=2021,
                total_mg=100,
                source_urls=("https://example.com/source.xlsx",),
            )
        ]
        collection_index = importer.build_collection_index(
            [
                {
                    "id": 123,
                    "country": "Denmark",
                    "catchment": "Aarhus (751)",
                    "waste_category": "Food waste",
                    "collection_system": "Door to door",
                }
            ]
        )

        records, warnings = importer.build_cpv_records(measurements, collection_index)

        self.assertEqual(warnings, [])
        self.assertEqual(len(records), 1)
        self.assertEqual({record.collection_id for record in records}, {123})
        self.assertEqual(
            {record.property_id for record in records}, {importer._PROP_TOTAL}
        )

    def test_build_collection_index_accepts_collection_api_payload_without_country(
        self,
    ):
        collection_index = importer.build_collection_index(
            [
                {
                    "id": 123,
                    "catchment": "Albertslund",
                    "waste_category": "Food waste",
                    "collection_system": "Door to door",
                },
                {
                    "id": 456,
                    "catchment": "Aalborg",
                    "waste_category": "Food waste",
                    "collection_system": "Door to door",
                },
            ],
            allowed_municipalities={
                importer._normalize_municipality_name("Albertslund")
            },
        )

        self.assertEqual(
            list(collection_index.keys()),
            [("albertslund", "Food waste", "Door to door")],
        )
        self.assertEqual(
            collection_index[("albertslund", "Food waste", "Door to door")][0]["id"],
            123,
        )

    def test_fetch_collections_reads_published_review_and_private_scopes(self):
        class FakeClient:
            def __init__(self):
                self.calls = []

            def get_json(self, _path, params):
                self.calls.append(params.copy())
                scope = params["scope"]
                results_by_scope = {
                    "published": [{"id": 1}],
                    "review": [{"id": 2}],
                    "private": [{"id": 2}, {"id": 3}],
                }
                return {"results": results_by_scope[scope], "next": None}

        client = FakeClient()

        collections = importer.fetch_collections(client)

        self.assertEqual([collection["id"] for collection in collections], [1, 2, 3])
        self.assertEqual(
            [call["scope"] for call in client.calls],
            ["published", "review", "private"],
        )

    def test_parse_primary_measurements_returns_empty_without_primary_sheet(self):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Kommunal MW reel genanvendelse"
        temp = NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp.close()
        path = Path(temp.name)
        workbook.save(path)
        self.addCleanup(path.unlink)

        self.assertEqual(importer.parse_primary_measurements(path, years={2023}), [])

    def test_parse_statbank_laby24_measurements_maps_target_fractions(self):
        csv_text = "\n".join(
            [
                "KOMGRP;BEHANDLING;AFFFRAK;TID;INDHOLD",
                "Hele landet;I alt;MADAFFALD;2023;278129",
                "Albertslund;Forbrænding;DAGRENOVATION OG LIGNENDE;2023;4853",
                "Albertslund;I alt;DAGRENOVATION OG LIGNENDE;2023;4853",
                "Albertslund;I alt;ORGANISK AFFALD, INKL. HAVEAFFALD;2023;1862",
                "Albertslund;I alt;MADAFFALD;2023;1365",
                "Albertslund;I alt;HUSHOLDNINGSAFFALD I ALT;2023;12962",
                "Albertslund;I alt;MADAFFALD;2022;..",
            ]
        )

        measurements = importer.parse_statbank_laby24_measurements(csv_text)

        self.assertEqual(len(measurements), 3)
        by_category = {
            measurement.waste_category: measurement for measurement in measurements
        }
        self.assertEqual(
            by_category["Residual waste"].source_fraction,
            "DAGRENOVATION OG LIGNENDE",
        )
        self.assertEqual(by_category["Residual waste"].total_mg, 4853)
        self.assertEqual(
            by_category["Biowaste"].source_fraction,
            "ORGANISK AFFALD, INKL. HAVEAFFALD",
        )
        self.assertEqual(by_category["Biowaste"].total_mg, 1862)
        self.assertEqual(by_category["Food waste"].source_fraction, "MADAFFALD")
        self.assertEqual(by_category["Food waste"].total_mg, 1365)
        self.assertEqual(
            by_category["Food waste"].source_urls,
            (importer._STATBANK_LABY24_URL,),
        )

    def test_build_cpv_records_matches_statbank_target_categories(self):
        measurements = [
            importer.DenmarkMeasurement(
                municipality="Albertslund",
                waste_category="Residual waste",
                source_fraction="DAGRENOVATION OG LIGNENDE",
                year=2023,
                total_mg=4853,
                source_urls=(importer._STATBANK_LABY24_URL,),
            ),
            importer.DenmarkMeasurement(
                municipality="Albertslund",
                waste_category="Biowaste",
                source_fraction="ORGANISK AFFALD, INKL. HAVEAFFALD",
                year=2023,
                total_mg=1862,
                source_urls=(importer._STATBANK_LABY24_URL,),
            ),
            importer.DenmarkMeasurement(
                municipality="Albertslund",
                waste_category="Food waste",
                source_fraction="MADAFFALD",
                year=2023,
                total_mg=1365,
                source_urls=(importer._STATBANK_LABY24_URL,),
            ),
        ]
        collection_index = importer.build_collection_index(
            [
                {
                    "id": 10,
                    "catchment": "Albertslund",
                    "waste_category": "Residual waste",
                    "collection_system": "Door to door",
                },
                {
                    "id": 20,
                    "catchment": "Albertslund",
                    "waste_category": "Biowaste",
                    "collection_system": "Door to door",
                },
                {
                    "id": 30,
                    "catchment": "Albertslund",
                    "waste_category": "Food waste",
                    "collection_system": "Door to door",
                },
            ]
        )

        records, warnings = importer.build_cpv_records(measurements, collection_index)

        self.assertEqual(warnings, [])
        self.assertEqual(
            {record.waste_category: record.collection_id for record in records},
            {
                "Residual waste": 10,
                "Biowaste": 20,
                "Food waste": 30,
            },
        )
        self.assertEqual(
            {record.property_id for record in records}, {importer._PROP_TOTAL}
        )
        self.assertEqual({record.unit_name for record in records}, {importer._UNIT_MG})
