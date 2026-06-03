from datetime import date, datetime
from io import StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase, override_settings

from bibliography.models import Source
from sources.waste_collection.models import (
    Collection,
    CollectionCatchment,
    CollectionPropertyValue,
    CollectionSystem,
    WasteCategory,
    WasteFlyer,
)
from utils.object_management.models import ReviewAction
from utils.properties.models import Property, Unit

User = get_user_model()


def _resolve_date(value) -> date | None:
    """Convert various date formats to date object."""
    if value is None:
        return None
    if isinstance(value, date):
        # If it's already a date but not a datetime, return it
        if not isinstance(value, datetime):
            return value
        # If it's a datetime, convert to date
        return value.date()
    # Try to parse string
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common formats
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(value, fmt)
                return parsed.date()
            except ValueError:
                continue
    return None


def _split_source_cell(raw: str) -> tuple[list[str], list[str]]:
    """Split a source cell into URLs and notes."""
    if not raw:
        return [], []
    urls = []
    notes = []
    for item in raw.split(","):
        item = item.strip()
        if not item:
            continue
        if item.startswith("http"):
            urls.append(item)
        else:
            notes.append(item)
    return urls, notes


def _load_records(file_path: Path) -> tuple[list[dict], list[str], int]:
    """Load a generic workbook into importer-compatible records for testing.

    Returns:
        Tuple of (valid_records, preflight_warnings, total_data_row_count).
    """
    workbook = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    headers = next(rows, None)

    if not headers:
        return [], ["No headers found"], 0

    records: list[dict] = []
    warnings: list[str] = []
    row_count = 0

    for row_number, row in enumerate(rows, start=2):
        row_count += 1
        row_data = dict(zip(headers, row, strict=True))

        # Check for required fields
        missing = []
        if not row_data.get("Collection System"):
            missing.append("collection_system")
        valid_from_raw = row_data.get("Valid from")
        valid_from = _resolve_date(valid_from_raw)
        if not valid_from:
            missing.append("valid_from")

        if missing:
            catchment = row_data.get("Catchment", "unknown")
            warnings.append(
                f"Row {row_number} ({catchment!r}): "
                f"skipped — missing required field(s): {', '.join(missing)}"
            )
            continue

        # Convert to API payload format
        sources_raw = row_data.get("Sources_new") or ""
        sources_urls, sources_notes = _split_source_cell(sources_raw)
        weblinks_raw = row_data.get("Weblinks") or ""
        weblinks_urls, weblinks_notes = _split_source_cell(weblinks_raw)
        sources = sources_urls + sources_notes
        weblinks = weblinks_urls + weblinks_notes

        record = {
            "catchment_name": row_data.get("Catchment"),
            "collection_system": row_data.get("Collection System"),
            "waste_category": row_data.get("Waste Category"),
            "valid_from": valid_from,
            "allowed_materials": row_data.get("Allowed Materials") or "",
            "forbidden_materials": row_data.get("Forbidden Materials") or "",
            "weblinks": weblinks,
            "sources": sources,
        }

        # Add property values if present
        for header in headers:
            if header and "Specific Waste Collected" in header:
                year_match = header.split()[-1]  # e.g., "2020"
                if year_match.isdigit():
                    year = int(year_match)
                    unit_header = f"{header} Unit"
                    unit = row_data.get(unit_header)
                    value = row_data.get(header)
                    if value is not None:
                        record[f"property_{year}"] = value
                        if unit:
                            record[f"property_{year}_unit"] = unit

        records.append(record)

    return records, warnings, row_count


@override_settings(AUTO_ENQUEUE_URL_CHECKS=False)
class SubmitImportedWorkbookForReviewCommandTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user(username="tobi")
        cls.other_owner = User.objects.create_user(username="other-owner")
        cls.catchment = CollectionCatchment.objects.create(name="Thueringen Catchment")
        cls.collection_system = CollectionSystem.objects.create(name="Brown Bin")
        cls.waste_category = WasteCategory.objects.create(name="Biowaste")
        cls.property, _ = Property.objects.get_or_create(
            id=1,
            defaults={"name": "Specific waste collected"},
        )
        cls.unit, _ = Unit.objects.get_or_create(name="kg/(cap.*a)")

    def setUp(self):
        self.workbook_path = self._create_workbook()

    def tearDown(self):
        if self.workbook_path.exists():
            self.workbook_path.unlink()

    def _create_workbook(self) -> Path:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "Catchment",
                "Collection System",
                "Waste Category",
                "Valid from",
                "Allowed Materials",
                "Forbidden Materials",
                "Weblinks",
                "Sources_new",
                "Specific Waste Collected 2020",
                "Specific Waste Collected 2020 Unit",
            ]
        )
        sheet.append(
            [
                self.catchment.name,
                self.collection_system.name,
                self.waste_category.name,
                date(2024, 1, 1),
                "",
                "",
                "https://thueringen.example.org/imported-flyer.pdf",
                "Thüringen Imported Note",
                12.3,
                self.unit.name,
            ]
        )

        handle = NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        return Path(handle.name)

    def _create_matching_collection(self, owner):
        return Collection.objects.create(
            owner=owner,
            publication_status="private",
            catchment=self.catchment,
            collection_system=self.collection_system,
            waste_category=self.waste_category,
            valid_from=date(2024, 1, 1),
        )

    def test_command_submits_only_matching_tobi_owned_import_objects(self):
        collection = self._create_matching_collection(self.owner)
        imported_source = Source.objects.create(
            owner=self.owner,
            publication_status="private",
            type="custom",
            title="Thüringen Imported Note",
        )
        extra_source = Source.objects.create(
            owner=self.owner,
            publication_status="private",
            type="custom",
            title="Not In Workbook",
        )
        imported_flyer = WasteFlyer.objects.create(
            owner=self.owner,
            publication_status="private",
            title="thueringen.example.org",
            url="https://thueringen.example.org/imported-flyer.pdf",
        )
        extra_flyer = WasteFlyer.objects.create(
            owner=self.owner,
            publication_status="private",
            title="thueringen.example.org",
            url="https://thueringen.example.org/not-imported.pdf",
        )
        collection.sources.add(imported_source, extra_source)
        collection.flyers.add(imported_flyer, extra_flyer)

        cpv = CollectionPropertyValue.objects.create(
            name="Thueringen CPV",
            owner=self.owner,
            publication_status="private",
            collection=collection,
            property=self.property,
            unit=self.unit,
            year=2020,
            average=12.3,
        )

        other_collection = self._create_matching_collection(self.other_owner)
        other_cpv = CollectionPropertyValue.objects.create(
            name="Other Owner CPV",
            owner=self.other_owner,
            publication_status="private",
            collection=other_collection,
            property=self.property,
            unit=self.unit,
            year=2020,
            average=12.3,
        )

        stdout = StringIO()
        call_command(
            "submit_imported_workbook_for_review",
            file=str(self.workbook_path),
            owner=self.owner.username,
            stdout=stdout,
        )

        collection.refresh_from_db()
        imported_source.refresh_from_db()
        extra_source.refresh_from_db()
        imported_flyer.refresh_from_db()
        extra_flyer.refresh_from_db()
        cpv.refresh_from_db()
        other_collection.refresh_from_db()
        other_cpv.refresh_from_db()

        self.assertEqual(collection.publication_status, "review")
        self.assertEqual(imported_source.publication_status, "review")
        self.assertEqual(imported_flyer.publication_status, "review")
        self.assertEqual(cpv.publication_status, "review")

        self.assertEqual(extra_source.publication_status, "private")
        self.assertEqual(extra_flyer.publication_status, "private")
        self.assertEqual(other_collection.publication_status, "private")
        self.assertEqual(other_cpv.publication_status, "private")

        self.assertTrue(
            ReviewAction.for_object(collection)
            .filter(action=ReviewAction.ACTION_SUBMITTED)
            .exists()
        )
        self.assertTrue(
            ReviewAction.for_object(imported_source)
            .filter(action=ReviewAction.ACTION_SUBMITTED)
            .exists()
        )
        self.assertTrue(
            ReviewAction.for_object(imported_flyer)
            .filter(action=ReviewAction.ACTION_SUBMITTED)
            .exists()
        )
        self.assertTrue(
            ReviewAction.for_object(cpv)
            .filter(action=ReviewAction.ACTION_SUBMITTED)
            .exists()
        )
        self.assertFalse(
            ReviewAction.for_object(extra_source)
            .filter(action=ReviewAction.ACTION_SUBMITTED)
            .exists()
        )
        self.assertFalse(
            ReviewAction.for_object(extra_flyer)
            .filter(action=ReviewAction.ACTION_SUBMITTED)
            .exists()
        )

        self.assertIn("Collections matched:      1", stdout.getvalue())
