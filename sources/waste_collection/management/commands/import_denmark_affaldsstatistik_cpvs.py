from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from csv import DictReader
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl
from django.core.management.base import BaseCommand, CommandError

_PROP_TOTAL = 9
_UNIT_MG = "Mg/a"
_COLLECTION_SYSTEM_DOOR_TO_DOOR = "Door to door"
_SOURCE_PAGE_URL = "https://mst.dk/erhverv/groen-produktion-og-affald/affald-og-genanvendelse/affaldshaandtering/affaldsdata-og-affaldsdatasystemet/find-affaldsstatistikker-og-kortlaegning"
_SOURCE_URLS = {
    2021: "https://mst.dk/media/ahmpl3eq/raadata_as2021.xlsx",
    2022: "https://mst.dk/media/g0upwp3a/raadata.xlsx",
    2023: "https://mst.dk/media/evsnaiqx/raadata_as_2023_til_brug_i_as2023_rettet_format_mw.xlsx",
}
_STATBANK_LABY24_URL = "https://www.statbank.dk/LABY24"
_STATBANK_API_URL = "https://api.statbank.dk/v1/data"
_DEFAULT_FILES = {
    2021: Path("denmark_affaldsstatistik_2021_raadata.xlsx"),
    2022: Path("denmark_affaldsstatistik_2022_raadata.xlsx"),
    2023: Path("denmark_affaldsstatistik_2023_raadata.xlsx"),
}
_PRIMARY_SHEET_KEY = "primær mængde"
_MUNICIPAL_RECYCLING_SHEET = "Kommunal MW reel genanvendelse"
_FRACTION_CATEGORY_MAP = {
    "Madaffald": "Food waste",
    "Haveaffald": "Green waste",
}
_STATBANK_LABY24_FRACTION_CATEGORY_MAP = {
    "DAGRENOVATION OG LIGNENDE": "Residual waste",
    "ORGANISK AFFALD, INKL. HAVEAFFALD": "Biowaste",
    "MADAFFALD": "Food waste",
}
_STATBANK_LABY24_FRACTION_CODES = {
    "A": "DAGRENOVATION OG LIGNENDE",
    "B": "ORGANISK AFFALD, INKL. HAVEAFFALD",
    "I": "MADAFFALD",
}
_STATBANK_LABY24_TOTAL_TREATMENT = "I alt"
_VALID_STATUSES = ("private", "review")
_BATCH_SIZE = 50
_URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)
_CATCHMENT_CODE_SUFFIX_RE = re.compile(r"\s*\(\d+\)\s*$")


@dataclass(frozen=True)
class DenmarkMeasurement:
    municipality: str
    waste_category: str
    source_fraction: str
    year: int
    total_mg: float
    source_urls: tuple[str, ...]


@dataclass(frozen=True)
class DenmarkCPVRecord:
    collection_id: int
    municipality: str
    waste_category: str
    source_fraction: str
    property_id: int
    unit_name: str
    year: int
    average: float
    source_urls: tuple[str, ...]


class ImportHttpClient:
    def __init__(self, api_url: str, token: str, timeout: int = 120):
        self.api_url = api_url.rstrip("/")
        self.token = token
        self.timeout = timeout

    def get_json(self, path: str, params: dict | None = None) -> dict | list:
        url = f"{self.api_url}{path}"
        if params:
            url = f"{url}?{urlencode(params, doseq=True)}"
        req = Request(url, headers=self._headers(), method="GET")
        return self._request_json(req)

    def post_json(self, path: str, payload: dict) -> dict:
        req = Request(
            f"{self.api_url}{path}",
            data=json.dumps(payload).encode(),
            headers={**self._headers(), "Content-Type": "application/json"},
            method="POST",
        )
        return self._request_json(req)

    def _headers(self) -> dict[str, str]:
        return {"Authorization": f"Token {self.token}"}

    def _request_json(self, req: Request):
        try:
            with urlopen(req, timeout=self.timeout) as resp:
                body = resp.read()
        except HTTPError as exc:
            detail = exc.read().decode(errors="replace")
            raise CommandError(f"API request failed ({exc.code}): {detail}") from exc
        except Exception as exc:
            raise CommandError(f"API request failed: {exc}") from exc
        return json.loads(body or b"{}")


def _get_token(api_url: str, username: str, password: str) -> str:
    payload = json.dumps({"username": username, "password": password}).encode()
    for path in ("/api/auth/token/", "/api-token-auth/"):
        req = Request(
            f"{api_url.rstrip('/')}{path}",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
        except Exception:
            continue
        token = data.get("token")
        if token:
            return token
    raise CommandError("Could not obtain token from known token endpoints.")


def _normalize_text(value) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text or None


def _normalize_municipality_name(value: str) -> str:
    text = _normalize_text(value) or ""
    text = _CATCHMENT_CODE_SUFFIX_RE.sub("", text)
    text = text.casefold()
    replacements = {
        "å": "aa",
        "ä": "ae",
        "æ": "ae",
        "ö": "oe",
        "ø": "oe",
        "ü": "ue",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[^a-z0-9]+", "", text)


def _safe_float(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip()
    if not text or text.upper() == "INTET":
        return None
    text = text.replace(" ", "")
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _header_index(header_row) -> dict[str, int]:
    return {name: index for index, name in enumerate(header_row) if name}


def _year_columns(header: dict[str, int]) -> dict[int, int]:
    result = {}
    for name, index in header.items():
        match = re.fullmatch(r"(20\d{2}) \[ton\]", str(name).strip())
        if match:
            result[int(match.group(1))] = index
    return result


def _source_urls_for_year(year: int) -> tuple[str, ...]:
    urls = [_SOURCE_PAGE_URL]
    source_url = _SOURCE_URLS.get(year)
    if source_url:
        urls.append(source_url)
    return tuple(urls)


def _statbank_source_urls() -> tuple[str, ...]:
    return (_STATBANK_LABY24_URL,)


def _find_primary_sheet(workbook) -> str | None:
    for sheet_name in workbook.sheetnames:
        if _PRIMARY_SHEET_KEY in sheet_name.casefold():
            return sheet_name
    return None


def parse_primary_measurements(
    file_path: Path, *, years: set[int] | None = None
) -> list[DenmarkMeasurement]:
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_name = _find_primary_sheet(workbook)
    if sheet_name is None:
        return []
    ws = workbook[sheet_name]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    header_row = next(rows, None)
    if header_row is None:
        return []
    header = _header_index(header_row)
    required = (
        "Ny affaldsfraktion",
        "Kommune",
    )
    if any(name not in header for name in required):
        return []
    municipality_col = header["Kommune"]
    fraction_col = header["Ny affaldsfraktion"]
    year_columns = _year_columns(header)
    if years is not None:
        year_columns = {
            year: col for year, col in year_columns.items() if year in years
        }
    totals = defaultdict(float)
    display_names = {}
    for row in rows:
        municipality = _normalize_text(
            row[municipality_col] if municipality_col < len(row) else None
        )
        fraction = _normalize_text(
            row[fraction_col] if fraction_col < len(row) else None
        )
        waste_category = _FRACTION_CATEGORY_MAP.get(fraction or "")
        if not municipality or not fraction or waste_category is None:
            continue
        normalized_municipality = _normalize_municipality_name(municipality)
        display_names[normalized_municipality] = municipality
        for year, column in year_columns.items():
            amount = _safe_float(row[column] if column < len(row) else None)
            if amount is None:
                continue
            totals[(normalized_municipality, waste_category, fraction, year)] += amount
    measurements = []
    for (municipality_key, waste_category, fraction, year), total_mg in sorted(
        totals.items()
    ):
        if total_mg <= 0:
            continue
        measurements.append(
            DenmarkMeasurement(
                municipality=display_names[municipality_key],
                waste_category=waste_category,
                source_fraction=fraction,
                year=year,
                total_mg=total_mg,
                source_urls=_source_urls_for_year(year),
            )
        )
    return measurements


def _municipal_recycling_column(
    header_row, starts_with: str, unit_fragment: str
) -> int | None:
    for index, value in enumerate(header_row):
        text = _normalize_text(value)
        if text and text.startswith(starts_with) and unit_fragment in text:
            return index
    return None


def parse_municipal_garden_measurements(
    file_path: Path, *, year: int
) -> list[DenmarkMeasurement]:
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if _MUNICIPAL_RECYCLING_SHEET not in workbook.sheetnames:
        return []
    ws = workbook[_MUNICIPAL_RECYCLING_SHEET]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    header_row = next(rows, None)
    if header_row is None:
        return []
    municipality_col = 0
    garden_col = _municipal_recycling_column(
        header_row,
        "Haveaffald fra husholdninger",
        "1000 tons",
    )
    if garden_col is None:
        return []
    measurements = []
    for row in rows:
        municipality = _normalize_text(
            row[municipality_col] if municipality_col < len(row) else None
        )
        total_kt = _safe_float(row[garden_col] if garden_col < len(row) else None)
        if not municipality or total_kt is None or total_kt <= 0:
            continue
        total_mg = total_kt * 1000
        measurements.append(
            DenmarkMeasurement(
                municipality=municipality,
                waste_category="Green waste",
                source_fraction="Haveaffald fra husholdninger",
                year=year,
                total_mg=total_mg,
                source_urls=_source_urls_for_year(year),
            )
        )
    return measurements


def parse_statbank_laby24_measurements(csv_text: str) -> list[DenmarkMeasurement]:
    measurements = []
    reader = DictReader(StringIO(csv_text), delimiter=";")
    for row in reader:
        municipality = _normalize_text(row.get("KOMGRP"))
        treatment = _normalize_text(row.get("BEHANDLING"))
        fraction = _normalize_text(row.get("AFFFRAK"))
        year = _safe_float(row.get("TID"))
        total_mg = _safe_float(row.get("INDHOLD"))
        waste_category = _STATBANK_LABY24_FRACTION_CATEGORY_MAP.get(fraction or "")
        if (
            not municipality
            or municipality == "Hele landet"
            or treatment != _STATBANK_LABY24_TOTAL_TREATMENT
            or waste_category is None
            or year is None
            or total_mg is None
            or total_mg <= 0
        ):
            continue
        measurements.append(
            DenmarkMeasurement(
                municipality=municipality,
                waste_category=waste_category,
                source_fraction=fraction,
                year=int(year),
                total_mg=total_mg,
                source_urls=_statbank_source_urls(),
            )
        )
    return measurements


def fetch_statbank_laby24_measurements(years: list[int]) -> list[DenmarkMeasurement]:
    payload = {
        "table": "LABY24",
        "format": "BULK",
        "lang": "da",
        "variables": [
            {"code": "KOMGRP", "values": ["*"]},
            {"code": "BEHANDLING", "values": ["TOT"]},
            {
                "code": "AFFFRAK",
                "values": list(_STATBANK_LABY24_FRACTION_CODES),
            },
            {"code": "Tid", "values": [str(year) for year in years]},
        ],
    }
    req = Request(
        _STATBANK_API_URL,
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urlopen(req, timeout=120) as resp:
            csv_text = resp.read().decode("utf-8-sig")
    except HTTPError as exc:
        detail = exc.read().decode(errors="replace")
        raise CommandError(f"StatBank request failed ({exc.code}): {detail}") from exc
    except Exception as exc:
        raise CommandError(f"StatBank request failed: {exc}") from exc
    return parse_statbank_laby24_measurements(csv_text)


def _extract_results(response: dict | list) -> list[dict]:
    if isinstance(response, list):
        return response
    if isinstance(response, dict):
        results = response.get("results")
        if isinstance(results, list):
            return results
    return []


def fetch_collections(client: ImportHttpClient) -> list[dict]:
    collections = []
    seen_ids = set()
    for scope in ("published", "review", "private"):
        page = 1
        while True:
            response = client.get_json(
                "/waste_collection/api/collection/",
                {"scope": scope, "page": page, "page_size": 100},
            )
            results = _extract_results(response)
            if not results:
                break
            for collection in results:
                collection_id = collection.get("id")
                if collection_id in seen_ids:
                    continue
                seen_ids.add(collection_id)
                collections.append(collection)
            if not isinstance(response, dict) or not response.get("next"):
                break
            page += 1
    return collections


def build_collection_index(
    collections: list[dict],
    *,
    allowed_municipalities: set[str] | None = None,
) -> dict[tuple[str, str, str], list[dict]]:
    index = defaultdict(list)
    for collection in collections:
        catchment = _normalize_text(collection.get("catchment")) or ""
        category = _normalize_text(collection.get("waste_category")) or ""
        system = _normalize_text(collection.get("collection_system")) or ""
        country = _normalize_text(collection.get("country"))
        if country and country not in {"Denmark", "DK"}:
            continue
        municipality_key = _normalize_municipality_name(catchment)
        if (
            allowed_municipalities is not None
            and municipality_key not in allowed_municipalities
        ):
            continue
        key = (municipality_key, category, system)
        index[key].append(collection)
    return index


def build_cpv_records(
    measurements: list[DenmarkMeasurement],
    collection_index: dict[tuple[str, str, str], list[dict]],
) -> tuple[list[DenmarkCPVRecord], list[str]]:
    records = []
    warnings = []
    for measurement in measurements:
        key = (
            _normalize_municipality_name(measurement.municipality),
            measurement.waste_category,
            _COLLECTION_SYSTEM_DOOR_TO_DOOR,
        )
        matches = collection_index.get(key, [])
        if not matches:
            warnings.append(
                f"No Denmark collection match for {measurement.municipality} / {measurement.waste_category} / {_COLLECTION_SYSTEM_DOOR_TO_DOOR}."
            )
            continue
        if len(matches) > 1:
            warnings.append(
                f"Multiple Denmark collection matches for {measurement.municipality} / {measurement.waste_category}; skipped."
            )
            continue
        collection_id = int(matches[0]["id"])
        records.append(
            DenmarkCPVRecord(
                collection_id=collection_id,
                municipality=measurement.municipality,
                waste_category=measurement.waste_category,
                source_fraction=measurement.source_fraction,
                property_id=_PROP_TOTAL,
                unit_name=_UNIT_MG,
                year=measurement.year,
                average=measurement.total_mg,
                source_urls=measurement.source_urls,
            )
        )
    return records, warnings


def _record_payload(record: DenmarkCPVRecord, submit_for_review: bool) -> dict:
    return {
        "collection": record.collection_id,
        "property_id": record.property_id,
        "unit_name": record.unit_name,
        "year": record.year,
        "average": record.average,
        "flyer_urls": list(record.source_urls),
        "submit_for_review": submit_for_review,
    }


class Command(BaseCommand):
    help = "Import Denmark Affaldsstatistik municipal CPVs via the BRIT API."

    def add_arguments(self, parser):
        parser.add_argument("--api-url", required=True)
        parser.add_argument("--token", default=None)
        parser.add_argument("--username", default=None)
        parser.add_argument("--password", default=None)
        parser.add_argument(
            "--year", type=int, action="append", choices=(2021, 2022, 2023)
        )
        parser.add_argument("--file-2021", default=str(_DEFAULT_FILES[2021]))
        parser.add_argument("--file-2022", default=str(_DEFAULT_FILES[2022]))
        parser.add_argument("--file-2023", default=str(_DEFAULT_FILES[2023]))
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--include-municipal-garden-sheet", action="store_true")
        parser.add_argument("--include-statbank-laby24", action="store_true")
        parser.add_argument(
            "--publication-status", default="review", choices=_VALID_STATUSES
        )
        parser.add_argument("--batch-size", type=int, default=_BATCH_SIZE)

    def handle(self, *args, **options):
        years = sorted(set(options.get("year") or (2021, 2022, 2023)))
        paths = {
            2021: Path(options["file_2021"]),
            2022: Path(options["file_2022"]),
            2023: Path(options["file_2023"]),
        }
        measurements = []
        for year in years:
            path = paths[year]
            if not path.exists():
                raise CommandError(f"Source file not found for {year}: {path}")
            parsed = parse_primary_measurements(path, years={year})
            if not parsed and options["include_municipal_garden_sheet"]:
                parsed = parse_municipal_garden_measurements(path, year=year)
            if not parsed:
                self.stdout.write(
                    f"No detailed municipal primary measurements parsed for {year} from {path.name}."
                )
            measurements.extend(parsed)
        if options["include_statbank_laby24"]:
            parsed = fetch_statbank_laby24_measurements(years)
            measurements.extend(parsed)
            self.stdout.write(f"Parsed {len(parsed)} StatBank LABY24 measurements.")
        self.stdout.write(f"Parsed {len(measurements)} measurements.")
        if measurements:
            summary = defaultdict(int)
            for measurement in measurements:
                summary[(measurement.year, measurement.waste_category)] += 1
            for (year, waste_category), count in sorted(summary.items()):
                self.stdout.write(f"  Measurements {year} / {waste_category}: {count}")

        token = options.get("token")
        if not token:
            username = options.get("username")
            password = options.get("password")
            if not username or not password:
                raise CommandError("Provide --token or both --username and --password.")
            token = _get_token(options["api_url"], username, password)
            self.stdout.write("Token obtained.")

        client = ImportHttpClient(options["api_url"], token)
        collections = fetch_collections(client)
        allowed_municipalities = {
            _normalize_municipality_name(measurement.municipality)
            for measurement in measurements
        }
        collection_index = build_collection_index(
            collections,
            allowed_municipalities=allowed_municipalities,
        )
        records, warnings = build_cpv_records(measurements, collection_index)
        self.stdout.write(f"Prepared {len(records)} CPV payloads.")
        if records:
            summary = defaultdict(int)
            for record in records:
                summary[(record.year, record.waste_category)] += 1
            for (year, waste_category), count in sorted(summary.items()):
                self.stdout.write(f"  CPVs {year} / {waste_category}: {count}")
        if warnings:
            self.stdout.write(f"Pre-flight warnings: {len(warnings)}")
            for warning in warnings[:50]:
                self.stdout.write(f"  {warning}")
            if len(warnings) > 50:
                self.stdout.write(f"  ... and {len(warnings) - 50} more")

        if options["dry_run"]:
            self.stdout.write("DRY RUN — no CPVs posted.")
            return

        submit_for_review = options["publication_status"] == "review"
        totals = defaultdict(int)
        batches = [
            records[i : i + options["batch_size"]]
            for i in range(0, len(records), options["batch_size"])
        ]
        for batch_index, batch in enumerate(batches, start=1):
            self.stdout.write(
                f"Posting batch {batch_index}/{len(batches)} ({len(batch)} CPVs)…"
            )
            sys.stdout.flush()
            for record in batch:
                result = client.post_json(
                    "/waste_collection/api/collection/property-value/create/",
                    _record_payload(record, submit_for_review),
                )
                for key in (
                    "cpv_created",
                    "cpv_unchanged",
                    "cpv_skipped",
                    "flyers_created",
                ):
                    totals[key] += int(result.get(key, 0))

        self.stdout.write("\n=== Import Summary ===")
        self.stdout.write(f"  CPVs created:    {totals['cpv_created']}")
        self.stdout.write(f"  CPVs unchanged:  {totals['cpv_unchanged']}")
        self.stdout.write(f"  CPVs skipped:    {totals['cpv_skipped']}")
        self.stdout.write(f"  Flyers created:  {totals['flyers_created']}")
